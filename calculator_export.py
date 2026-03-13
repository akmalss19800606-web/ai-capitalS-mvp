"""
Calculator History & Export Router — CALC-EXPORT-001
Endpoints: GET /history, GET /history/{id}, DELETE /history/{id}, GET /history/{id}/pdf, GET /history/{id}/xlsx
Based on TZ v3.0 calculator export requirements.
"""
import io
import json
import logging
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel

from app.db.session import get_db
from app.api.v1.routers.auth import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/calculator", tags=["calculator-history"])


# ── Pydantic Models ──
class CalcHistoryItem(BaseModel):
    id: int
    calc_type: str
    inputs: dict
    results: dict
    currency: str = "USD"
    created_at: Optional[str] = None


# ── DB Model (add to app/db/models/ if not exists) ──
# CREATE TABLE calculator_history (
#   id SERIAL PRIMARY KEY,
#   user_id INTEGER NOT NULL REFERENCES users(id),
#   calc_type VARCHAR(50) NOT NULL,
#   inputs JSONB NOT NULL,
#   results JSONB NOT NULL,
#   currency VARCHAR(3) DEFAULT 'USD',
#   created_at TIMESTAMP DEFAULT NOW()
# );


def _get_calc_model(db: Session):
    """Try to import CalculatorHistory model, fallback to raw SQL."""
    try:
        from app.db.models.calculator_history import CalculatorHistory
        return CalculatorHistory
    except ImportError:
        return None


@router.get("/history", summary="List calculator history")
async def list_history(
    calc_type: Optional[str] = Query(None),
    limit: int = Query(50, le=200),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    Model = _get_calc_model(db)
    if Model is None:
        return []
    q = db.query(Model).filter(Model.user_id == user.id)
    if calc_type:
        q = q.filter(Model.calc_type == calc_type)
    items = q.order_by(Model.created_at.desc()).limit(limit).all()
    return [CalcHistoryItem(
        id=i.id, calc_type=i.calc_type, inputs=i.inputs,
        results=i.results, currency=getattr(i, "currency", "USD"),
        created_at=str(i.created_at) if i.created_at else None,
    ) for i in items]


@router.get("/history/{calc_id}", summary="Get single calculator result")
async def get_history_item(
    calc_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)
):
    Model = _get_calc_model(db)
    if Model is None:
        raise HTTPException(status_code=404, detail="History table not available")
    item = db.query(Model).filter(Model.id == calc_id, Model.user_id == user.id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Calculation not found")
    return CalcHistoryItem(
        id=item.id, calc_type=item.calc_type, inputs=item.inputs,
        results=item.results, currency=getattr(item, "currency", "USD"),
        created_at=str(item.created_at) if item.created_at else None,
    )


@router.delete("/history/{calc_id}", summary="Delete calculator result")
async def delete_history_item(
    calc_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)
):
    Model = _get_calc_model(db)
    if Model is None:
        raise HTTPException(status_code=404, detail="History table not available")
    item = db.query(Model).filter(Model.id == calc_id, Model.user_id == user.id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Calculation not found")
    db.delete(item)
    db.commit()
    return {"detail": "Deleted"}


@router.get("/history/{calc_id}/pdf", summary="Export calculator result as PDF")
async def export_calc_pdf(
    calc_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)
):
    Model = _get_calc_model(db)
    if Model is None:
        raise HTTPException(status_code=404, detail="History table not available")
    item = db.query(Model).filter(Model.id == calc_id, Model.user_id == user.id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Calculation not found")
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.colors import HexColor

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=25*mm, rightMargin=25*mm)
        styles = getSampleStyleSheet()
        title_s = ParagraphStyle("CT", parent=styles["Title"], fontSize=16, spaceAfter=10)
        body_s = ParagraphStyle("CB", parent=styles["BodyText"], fontSize=10, spaceAfter=6)
        story = []
        story.append(Paragraph(f"AI Capital — {item.calc_type.upper()} Calculator Report", title_s))
        story.append(Paragraph(f"Date: {item.created_at}", body_s))
        story.append(Spacer(1, 10))
        story.append(Paragraph("<b>Inputs</b>", body_s))
        for k, v in item.inputs.items():
            story.append(Paragraph(f"{k}: {v}", body_s))
        story.append(Spacer(1, 10))
        story.append(Paragraph("<b>Results</b>", body_s))
        for k, v in item.results.items():
            if isinstance(v, (int, float)):
                story.append(Paragraph(f"{k}: {v:,.2f}", body_s))
            elif isinstance(v, str):
                story.append(Paragraph(f"{k}: {v}", body_s))
        doc.build(story)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="calc-{calc_id}.pdf"'})
    except ImportError:
        raise HTTPException(status_code=501, detail="reportlab not installed")


@router.get("/history/{calc_id}/xlsx", summary="Export calculator result as XLSX")
async def export_calc_xlsx(
    calc_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)
):
    Model = _get_calc_model(db)
    if Model is None:
        raise HTTPException(status_code=404, detail="History table not available")
    item = db.query(Model).filter(Model.id == calc_id, Model.user_id == user.id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Calculation not found")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Calculator Result"
        ws.append(["AI Capital Calculator", item.calc_type.upper()])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append(["Date", str(item.created_at)])
        ws.append([])
        ws.append(["— INPUTS —"])
        ws["A4"].font = Font(bold=True)
        for k, v in item.inputs.items():
            ws.append([k, str(v)])
        ws.append([])
        ws.append(["— RESULTS —"])
        ws["A" + str(ws.max_row)].font = Font(bold=True)
        for k, v in item.results.items():
            if isinstance(v, (list, dict)):
                ws.append([k, json.dumps(v, ensure_ascii=False)[:500]])
            else:
                ws.append([k, v])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="calc-{calc_id}.xlsx"'})
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl not installed")
