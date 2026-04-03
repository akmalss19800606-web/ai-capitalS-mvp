"""
E2-03: Парсер 1С Excel (10-листовый формат).

Поддерживает выгрузки 1С:Бухгалтерия для Узбекистана с листами:
📋 Реквизиты, 📊 ОСВ_Годовая, 🏗️ Основные_Средства, 📦 Запасы,
👥 Дебиторы_Кредиторы, 💰 Денежные_Средства, 📈 Доходы_Расходы,
🏦 Кредиты_Займы, 💼 Капитал, 🧾 Налог_на_прибыль
"""
import io
import re
import openpyxl
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Optional
from datetime import date


@dataclass
class BalanceEntry:
    account_code: str
    account_name: str
    debit_start: Decimal = Decimal(0)
    credit_start: Decimal = Decimal(0)
    debit_end: Decimal = Decimal(0)
    credit_end: Decimal = Decimal(0)


@dataclass
class FinancialStatementRow:
    line_code: str
    name: str
    value_start: Decimal = Decimal(0)
    value_end: Decimal = Decimal(0)


@dataclass
class ParsedFinancialData:
    organization_name: str = ""
    inn: str = ""
    address: str = ""
    activity: str = ""
    director: str = ""
    accountant: str = ""
    charter_capital: str = ""
    registration_date: str = ""
    accounting_system: str = ""
    unit: str = ""
    period_from: Optional[date] = None
    period_to: Optional[date] = None
    previous_period_from: Optional[date] = None
    previous_period_to: Optional[date] = None
    osv_entries: list[BalanceEntry] = field(default_factory=list)
    fixed_assets: list[dict] = field(default_factory=list)
    inventory: list[dict] = field(default_factory=list)
    receivables: list[dict] = field(default_factory=list)
    payables: list[dict] = field(default_factory=list)
    cash: list[dict] = field(default_factory=list)
    cashflow: list[dict] = field(default_factory=list)
    income_expenses: list[dict] = field(default_factory=list)
    loans: list[dict] = field(default_factory=list)
    capital_rows: list[dict] = field(default_factory=list)
    tax_rows: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    sheets_found: list[str] = field(default_factory=list)
    sheets_parsed: list[str] = field(default_factory=list)


# Mapping: canonical category -> possible substrings in real sheet names
_SHEET_ALIASES: dict[str, list[str]] = {
    "requisites": ["Реквизиты", "Реквизит", "Requisites", "Общие"],
    "osv": ["ОСВ", "Оборотно-сальдовая", "OSV", "Обороты"],
    "fixed_assets": ["Основные_Средства", "Основные средства", "Fixed Assets"],
    "inventory": ["Запасы", "ТМЗ", "Inventory", "Материалы"],
    "debtors_creditors": [
        "Дебиторы_Кредиторы", "Дебиторы", "Кредиторы",
        "Дебиторская", "Кредиторская", "Receivables", "Payables",
    ],
    "cash": ["Денежные_Средства", "Денежные средства", "Cash"],
    "income_expenses": ["Доходы_Расходы", "Доходы и расходы", "Доходы", "ОПиУ", "P&L"],
    "loans": ["Кредиты_Займы", "Кредиты и займы", "Кредиты", "Займы", "Loans"],
    "capital": ["Капитал", "Equity", "Capital"],
    "tax": ["Налог", "Tax", "НП", "Налог_на_прибыль"],
}


class Excel1CParser:
    """
    Парсер выгрузок 1С в формате Excel.
    Поддерживает стандартные форматы 1С:Бухгалтерия для Узбекистана.
    """

    def parse(self, file_bytes: bytes, filename: str = "") -> ParsedFinancialData:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), data_only=True, read_only=True,
        )
        result = ParsedFinancialData()
        result.sheets_found = list(wb.sheetnames)

        # Parse each category
        self._parse_requisites(wb, result)
        self._parse_osv(wb, result)
        self._parse_fixed_assets(wb, result)
        self._parse_inventory(wb, result)
        self._parse_debtors_creditors(wb, result)
        self._parse_cash(wb, result)
        self._parse_income_expenses(wb, result)
        self._parse_loans(wb, result)
        self._parse_capital(wb, result)
        self._parse_tax(wb, result)

        wb.close()
        return result

    # ── Sheet finder ──────────────────────────────────────────

    def _find_sheet(self, wb: openpyxl.Workbook, category: str):
        aliases = _SHEET_ALIASES.get(category, [])
        for name in wb.sheetnames:
            # Strip emoji prefixes for matching
            clean = re.sub(r'[^\w\s]', '', name).strip()
            for alias in aliases:
                alias_clean = re.sub(r'[^\w\s]', '', alias).strip()
                if alias_clean.lower() in clean.lower():
                    return wb[name]
        return None

    # ── Реквизиты ─────────────────────────────────────────────

    def _parse_requisites(self, wb, result: ParsedFinancialData):
        ws = self._find_sheet(wb, "requisites")
        if not ws:
            result.warnings.append("Лист 'Реквизиты' не найден")
            return

        result.sheets_parsed.append(self._sheet_name(ws))

        # The requisites sheet has label in col B and value in col C
        label_map = {
            "Полное наименование": "organization_name",
            "ИНН": "inn",
            "Юридический адрес": "address",
            "Вид деятельности": "activity",
            "Уставный капитал": "charter_capital",
            "Дата государственной регистрации": "registration_date",
            "Директор": "director",
            "Главный бухгалтер": "accountant",
            "Единица измерения": "unit",
            "Учётная система": "accounting_system",
        }

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 3:
                continue
            label = str(row[1] or "").strip().rstrip(":")
            value = str(row[2] or "").strip() if row[2] is not None else ""
            for key, attr in label_map.items():
                if key.lower() in label.lower() and value:
                    setattr(result, attr, value)

            # Parse periods
            if "Отчётный период" in label and value:
                result.period_from, result.period_to = self._parse_period(value)
            elif "Предыдущий период" in label and value:
                result.previous_period_from, result.previous_period_to = self._parse_period(value)

    # ── ОСВ ───────────────────────────────────────────────────

    def _parse_osv(self, wb, result: ParsedFinancialData):
        ws = self._find_sheet(wb, "osv")
        if not ws:
            result.warnings.append("Лист 'ОСВ' не найден")
            return

        result.sheets_parsed.append(self._sheet_name(ws))
        # Structure: row 5 is header (Код, Наименование, Дт нач, Кт нач, Дт кон, Кт кон)
        # Data rows start from row 6, col B=code, C=name, D=debit_start, E=credit_start, F=debit_end, G=credit_end
        # Section headers have no account code
        # Skip ИТОГО / КОНТРОЛЬНАЯ rows
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 7:
                continue
            code_raw = row[1]
            name_raw = row[2]
            if code_raw is None or name_raw is None:
                continue
            code = str(code_raw).strip()
            # Must look like an account code (digits, possibly with dot)
            if not re.match(r'^\d{4}', code):
                continue
            # Skip control/total rows
            name = str(name_raw).strip()
            if any(kw in name.upper() for kw in ["ИТОГО", "КОНТРОЛЬ", "ВСЕГО"]):
                continue

            result.osv_entries.append(BalanceEntry(
                account_code=code,
                account_name=name,
                debit_start=self._safe_decimal(row[3]),
                credit_start=self._safe_decimal(row[4]),
                debit_end=self._safe_decimal(row[5]),
                credit_end=self._safe_decimal(row[6]),
            ))

    # ── Основные средства ─────────────────────────────────────

    def _parse_fixed_assets(self, wb, result: ParsedFinancialData):
        ws = self._find_sheet(wb, "fixed_assets")
        if not ws:
            result.warnings.append("Лист 'Основные средства' не найден")
            return

        result.sheets_parsed.append(self._sheet_name(ws))

        # Section A: rows 5(header)..11 — original cost
        # Section B: rows 14(header)..20 — accumulated depreciation
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        section = None
        for row in rows:
            if not row or len(row) < 3:
                continue
            col_b = str(row[1] or "").strip()

            # Detect sections
            if "ПЕРВОНАЧАЛЬНАЯ СТОИМОСТЬ" in col_b.upper():
                section = "cost"
                continue
            if "НАКОПЛЕННАЯ АМОРТИЗАЦИЯ" in col_b.upper() or "АМОРТИЗАЦИЯ" in col_b.upper():
                section = "depreciation"
                continue
            if "БАЛАНСОВАЯ СТОИМОСТЬ" in col_b.upper():
                section = "net"
                continue

            # Skip headers and totals
            if col_b.startswith("Категория") or col_b.startswith("ИТОГО") or not col_b:
                continue
            if section == "cost" and col_b and not col_b.startswith("ИТОГО"):
                entry = {
                    "section": "cost",
                    "category": col_b,
                    "balance_start": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                    "inflow": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                    "disposal": self._safe_decimal_val(row[4] if len(row) > 4 else None),
                    "revaluation": self._safe_decimal_val(row[5] if len(row) > 5 else None),
                    "balance_end": self._safe_decimal_val(row[6] if len(row) > 6 else None),
                    "useful_life": self._safe_decimal_val(row[7] if len(row) > 7 else None),
                }
                if entry["balance_start"] or entry["balance_end"]:
                    result.fixed_assets.append(entry)
            elif section == "depreciation" and col_b and not col_b.startswith("ИТОГО"):
                entry = {
                    "section": "depreciation",
                    "category": col_b,
                    "accum_start": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                    "charged": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                    "disposal_accum": self._safe_decimal_val(row[4] if len(row) > 4 else None),
                    "accum_end": self._safe_decimal_val(row[6] if len(row) > 6 else None),
                    "rate": self._safe_decimal_val(row[7] if len(row) > 7 else None),
                }
                if entry["accum_start"] or entry["accum_end"]:
                    result.fixed_assets.append(entry)

    # ── Запасы ────────────────────────────────────────────────

    def _parse_inventory(self, wb, result: ParsedFinancialData):
        ws = self._find_sheet(wb, "inventory")
        if not ws:
            result.warnings.append("Лист 'Запасы' не найден")
            return

        result.sheets_parsed.append(self._sheet_name(ws))

        # Structure: row 4 = header, data from row 5
        # col B=name, C=balance_start, D=inflow, E=outflow, F=balance_end, G=nrv
        rows = list(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True))
        current_group = ""
        for row in rows:
            if not row or len(row) < 3:
                continue
            col_b = str(row[1] or "").strip()
            if not col_b or col_b.startswith("Вид запасов"):
                continue

            # Group headers (all caps or contain "сч.")
            if col_b.isupper() or (col_b.startswith("ИТОГО") or col_b.startswith("ПРОВЕРКА")):
                if "ИТОГО" in col_b or "ПРОВЕРКА" in col_b:
                    continue
                current_group = col_b
                continue

            entry = {
                "group": current_group,
                "name": col_b,
                "balance_start": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                "inflow": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                "outflow": self._safe_decimal_val(row[4] if len(row) > 4 else None),
                "balance_end": self._safe_decimal_val(row[5] if len(row) > 5 else None),
                "nrv": self._safe_decimal_val(row[6] if len(row) > 6 else None),
            }
            if entry["balance_start"] or entry["balance_end"] or entry["inflow"]:
                result.inventory.append(entry)

    # ── Дебиторы и Кредиторы ──────────────────────────────────

    def _parse_debtors_creditors(self, wb, result: ParsedFinancialData):
        ws = self._find_sheet(wb, "debtors_creditors")
        if not ws:
            result.warnings.append("Лист 'Дебиторы_Кредиторы' не найден")
            return

        result.sheets_parsed.append(self._sheet_name(ws))

        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        section = None  # "receivables", "ecl", "payables"
        for row in rows:
            if not row or len(row) < 3:
                continue
            col_b = str(row[1] or "").strip()
            if not col_b:
                continue

            # Detect sections
            upper = col_b.upper()
            if "ДЕБИТОРСКАЯ ЗАДОЛЖЕННОСТЬ" in upper and "РЕЗЕРВ" not in upper:
                section = "receivables"
                continue
            if "РЕЗЕРВ ECL" in upper:
                section = "ecl"
                continue
            if "КРЕДИТОРСКАЯ ЗАДОЛЖЕННОСТЬ" in upper:
                section = "payables"
                continue

            # Skip headers
            if col_b.startswith("Группа") or col_b.startswith("Наименование"):
                continue
            if "ИТОГО" in upper:
                continue

            if section == "receivables":
                entry = {
                    "type": "receivable",
                    "name": col_b,
                    "balance_start": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                    "turnover_debit": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                    "turnover_credit": self._safe_decimal_val(row[4] if len(row) > 4 else None),
                    "balance_end": self._safe_decimal_val(row[5] if len(row) > 5 else None),
                }
                if entry["balance_start"] or entry["balance_end"]:
                    result.receivables.append(entry)
            elif section == "ecl":
                entry = {
                    "type": "ecl_reserve",
                    "name": col_b,
                    "gross_amount": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                    "ecl_rate": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                    "ecl_reserve": self._safe_decimal_val(row[4] if len(row) > 4 else None),
                    "net_amount": self._safe_decimal_val(row[5] if len(row) > 5 else None),
                }
                if entry["gross_amount"]:
                    result.receivables.append(entry)
            elif section == "payables":
                entry = {
                    "type": "payable",
                    "name": col_b,
                    "balance_start": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                    "turnover_debit": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                    "turnover_credit": self._safe_decimal_val(row[4] if len(row) > 4 else None),
                    "balance_end": self._safe_decimal_val(row[5] if len(row) > 5 else None),
                }
                if entry["balance_start"] or entry["balance_end"]:
                    result.payables.append(entry)

    # ── Денежные средства ─────────────────────────────────────

    def _parse_cash(self, wb, result: ParsedFinancialData):
        ws = self._find_sheet(wb, "cash")
        if not ws:
            result.warnings.append("Лист 'Денежные средства' не найден")
            return

        result.sheets_parsed.append(self._sheet_name(ws))

        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        section = None  # "balances" or "cashflow"
        current_group = ""
        for row in rows:
            if not row or len(row) < 3:
                continue
            col_b = str(row[1] or "").strip()
            if not col_b:
                continue

            upper = col_b.upper()
            # Detect top-level sections
            if "ДВИЖЕНИЕ ДС ПО ВИДАМ" in upper:
                section = "cashflow"
                continue
            if "ДЕНЕЖНЫЕ СРЕДСТВА" in upper and "ДВИЖЕНИЕ" not in upper and "ИТОГО" not in upper:
                section = "balances"
                continue

            # Group headers (e.g. КАССА, РАСЧЁТНЫЕ СЧЕТА)
            if upper.startswith("КАССА") or upper.startswith("РАСЧЁТНЫЕ") or \
               upper.startswith("ВАЛЮТНЫЕ") or upper.startswith("ПРОЧИЕ") or \
               upper.startswith("ОПЕРАЦИОННАЯ") or upper.startswith("ИНВЕСТИЦИОННАЯ") or \
               upper.startswith("ФИНАНСОВАЯ"):
                current_group = col_b
                continue

            # Skip headers/totals
            if col_b.startswith("Счёт") or col_b.startswith("Вид деятельности") or "ИТОГО" in upper:
                continue

            if section == "cashflow":
                entry = {
                    "section": "cashflow",
                    "group": current_group,
                    "name": col_b,
                    "inflow": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                    "outflow": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                    "net": self._safe_decimal_val(row[4] if len(row) > 4 else None),
                    "previous_inflow": self._safe_decimal_val(row[5] if len(row) > 5 else None),
                    "previous_outflow": self._safe_decimal_val(row[6] if len(row) > 6 else None),
                    "previous_net": self._safe_decimal_val(row[7] if len(row) > 7 else None),
                    "source": str(row[8] or "") if len(row) > 8 else "",
                }
                if entry["inflow"] or entry["outflow"]:
                    result.cashflow.append(entry)
            else:
                entry = {
                    "section": "balance",
                    "group": current_group,
                    "name": col_b,
                    "balance_start": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                    "inflow": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                    "outflow": self._safe_decimal_val(row[4] if len(row) > 4 else None),
                    "balance_end": self._safe_decimal_val(row[5] if len(row) > 5 else None),
                    "currency_info": str(row[6] or "") if len(row) > 6 else "",
                }
                if entry["balance_start"] or entry["balance_end"] or entry["inflow"]:
                    result.cash.append(entry)

    # ── Доходы и расходы ──────────────────────────────────────

    def _parse_income_expenses(self, wb, result: ParsedFinancialData):
        ws = self._find_sheet(wb, "income_expenses")
        if not ws:
            result.warnings.append("Лист 'Доходы и расходы' не найден")
            return

        result.sheets_parsed.append(self._sheet_name(ws))

        # Structure: row 4 = header, data from row 5
        # col B = account + name, C = current year, D = previous year, E = source
        rows = list(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True))
        current_section = ""
        for row in rows:
            if not row or len(row) < 3:
                continue
            col_b = str(row[1] or "").strip()
            if not col_b or col_b.startswith("Статья"):
                continue

            upper = col_b.upper()
            # Section headers (Roman numerals or ИТОГО)
            if upper.startswith("I.") or upper.startswith("II.") or \
               upper.startswith("III.") or upper.startswith("IV."):
                current_section = col_b
                continue
            if "ИТОГО" in upper or "ПРИБЫЛЬ" in upper:
                continue

            entry = {
                "section": current_section,
                "name": col_b,
                "current_year": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                "previous_year": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                "source": str(row[4] or "") if len(row) > 4 else "",
            }
            if entry["current_year"] or entry["previous_year"]:
                result.income_expenses.append(entry)

    # ── Кредиты и займы ───────────────────────────────────────

    def _parse_loans(self, wb, result: ParsedFinancialData):
        ws = self._find_sheet(wb, "loans")
        if not ws:
            result.warnings.append("Лист 'Кредиты и займы' не найден")
            return

        result.sheets_parsed.append(self._sheet_name(ws))

        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        current_group = ""
        section = "principal"  # "principal" or "interest"
        for row in rows:
            if not row or len(row) < 3:
                continue
            col_b = str(row[1] or "").strip()
            if not col_b:
                continue

            upper = col_b.upper()
            # Detect interest section
            if "РАСХОДЫ ПО ФИНАНСИРОВАНИЮ" in upper:
                section = "interest"
                continue

            # Group headers
            if "ДОЛГОСРОЧНЫЕ" in upper or "КРАТКОСРОЧНЫЕ" in upper or "ЛИЗИНГОВЫЕ" in upper:
                current_group = col_b
                continue

            # Skip headers/totals
            if col_b.startswith("Банк") or "ИТОГО" in upper:
                continue

            if section == "principal":
                entry = {
                    "section": "principal",
                    "group": current_group,
                    "name": col_b,
                    "balance_start": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                    "received": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                    "repaid": self._safe_decimal_val(row[4] if len(row) > 4 else None),
                    "balance_end": self._safe_decimal_val(row[5] if len(row) > 5 else None),
                    "rate": self._safe_decimal_val(row[6] if len(row) > 6 else None),
                }
                if entry["balance_start"] or entry["balance_end"]:
                    result.loans.append(entry)
            else:
                entry = {
                    "section": "interest",
                    "name": col_b,
                    "charged": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                    "paid": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                    "outstanding": self._safe_decimal_val(row[4] if len(row) > 4 else None),
                }
                if entry["charged"] or entry["paid"]:
                    result.loans.append(entry)

    # ── Капитал ───────────────────────────────────────────────

    def _parse_capital(self, wb, result: ParsedFinancialData):
        ws = self._find_sheet(wb, "capital")
        if not ws:
            result.warnings.append("Лист 'Капитал' не найден")
            return

        result.sheets_parsed.append(self._sheet_name(ws))

        # Structure: row 4 = header, data from row 5
        # col B = name, C = balance_start, D = movement, E = balance_end
        rows = list(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True))
        for row in rows:
            if not row or len(row) < 4:
                continue
            col_b = str(row[1] or "").strip()
            if not col_b or col_b.startswith("Статья") or "ИТОГО" in col_b.upper():
                continue

            entry = {
                "name": col_b,
                "balance_start": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                "movement": self._safe_decimal_val(row[3] if len(row) > 3 else None),
                "balance_end": self._safe_decimal_val(row[4] if len(row) > 4 else None),
            }
            if entry["balance_start"] or entry["balance_end"]:
                result.capital_rows.append(entry)

    # ── Налог на прибыль ──────────────────────────────────────

    def _parse_tax(self, wb, result: ParsedFinancialData):
        ws = self._find_sheet(wb, "tax")
        if not ws:
            result.warnings.append("Лист 'Налог на прибыль' не найден")
            return

        result.sheets_parsed.append(self._sheet_name(ws))

        # Structure: row 4 = header, data from row 5
        # col B = name, C = current year, D = previous year
        rows = list(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True))
        current_section = ""
        for row in rows:
            if not row or len(row) < 3:
                continue
            col_b = str(row[1] or "").strip()
            if not col_b or col_b.startswith("Показатель"):
                continue

            upper = col_b.upper()
            if upper.startswith("I.") or upper.startswith("II."):
                current_section = col_b
                continue

            entry = {
                "section": current_section,
                "name": col_b,
                "current_year": self._safe_decimal_val(row[2] if len(row) > 2 else None),
                "previous_year": self._safe_decimal_val(row[3] if len(row) > 3 else None),
            }
            if entry["current_year"] or entry["previous_year"]:
                result.tax_rows.append(entry)

    # ── Helpers ───────────────────────────────────────────────

    @staticmethod
    def _sheet_name(ws) -> str:
        return ws.title

    @staticmethod
    def _safe_decimal(value) -> Decimal:
        if value is None:
            return Decimal(0)
        try:
            s = str(value).replace(" ", "").replace(",", ".").replace("\xa0", "")
            # Filter out formulas
            if s.startswith("="):
                return Decimal(0)
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return Decimal(0)

    @staticmethod
    def _safe_decimal_val(value) -> float:
        """Return float for JSON-serializable dict entries."""
        if value is None:
            return 0.0
        try:
            s = str(value).replace(" ", "").replace(",", ".").replace("\xa0", "")
            if s.startswith("="):
                return 0.0
            return float(s)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _parse_period(text: str) -> tuple[Optional[date], Optional[date]]:
        """Parse period string like '01.01.2025 — 31.12.2025'."""
        # Find all date patterns dd.mm.yyyy
        dates = re.findall(r'(\d{2})\.(\d{2})\.(\d{4})', text)
        result_dates: list[Optional[date]] = [None, None]
        for i, m in enumerate(dates[:2]):
            try:
                result_dates[i] = date(int(m[2]), int(m[1]), int(m[0]))
            except ValueError:
                pass
        return result_dates[0], result_dates[1]
