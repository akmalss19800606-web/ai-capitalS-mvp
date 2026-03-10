"""
Модуль загрузки и AI-анализа документов.

Функционал:
  - Загрузка PDF, DOCX, XLSX файлов
  - Извлечение текста (PDF → PyPDF2/pdfplumber, DOCX → python-docx, XLSX → openpyxl)
  - AI-извлечение полей через мультипровайдерный ai_service
  - Хранение метаданных документов в БД
  - Локальное файловое хранилище (/app/uploads/)
"""

import os
import logging
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Tuple

logger = logging.getLogger(__name__)

# Директория хранения загруженных файлов
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", "/app/uploads"))
try:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
except PermissionError:
    UPLOAD_DIR = Path("/tmp/app_uploads")
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# Допустимые типы файлов
ALLOWED_EXTENSIONS = {
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".doc": "application/msword",
    ".xls": "application/vnd.ms-excel",
    ".txt": "text/plain",
    ".csv": "text/csv",
}

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB


# ─── Извлечение текста ───────────────────────────────────────────────

def extract_text_from_pdf(file_path: str) -> str:
    """Извлечение текста из PDF."""
    text_parts = []

    # Попытка 1: PyPDF2
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(file_path)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
        if text_parts:
            return "\n\n".join(text_parts)
    except ImportError:
        logger.debug("PyPDF2 не установлен")
    except Exception as e:
        logger.warning("PyPDF2 не смог извлечь текст: %s", e)

    # Попытка 2: pdfplumber (лучше для таблиц)
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        if text_parts:
            return "\n\n".join(text_parts)
    except ImportError:
        logger.debug("pdfplumber не установлен")
    except Exception as e:
        logger.warning("pdfplumber не смог извлечь текст: %s", e)

    return ""


def extract_text_from_docx(file_path: str) -> str:
    """Извлечение текста из DOCX."""
    try:
        from docx import Document
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except ImportError:
        logger.warning("python-docx не установлен")
        return ""
    except Exception as e:
        logger.error("Ошибка извлечения текста из DOCX: %s", e)
        return ""


def extract_text_from_xlsx(file_path: str) -> str:
    """Извлечение текста из XLSX."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        text_parts = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                text_parts.append(f"--- Лист: {ws.title} ---\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(text_parts)
    except ImportError:
        logger.warning("openpyxl не установлен")
        return ""
    except Exception as e:
        logger.error("Ошибка извлечения текста из XLSX: %s", e)
        return ""


def extract_text_from_txt(file_path: str) -> str:
    """Чтение TXT/CSV."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="cp1251") as f:
            return f.read()
    except Exception as e:
        logger.error("Ошибка чтения текстового файла: %s", e)
        return ""


def extract_text(file_path: str) -> str:
    """Автоматическое извлечение текста по расширению."""
    ext = Path(file_path).suffix.lower()

    if ext == ".pdf":
        return extract_text_from_pdf(file_path)
    elif ext == ".docx":
        return extract_text_from_docx(file_path)
    elif ext in (".xlsx", ".xls"):
        return extract_text_from_xlsx(file_path)
    elif ext in (".txt", ".csv"):
        return extract_text_from_txt(file_path)
    else:
        logger.warning("Неподдерживаемый формат файла: %s", ext)
        return ""


# ─── Хранение файлов ─────────────────────────────────────────────────

def save_uploaded_file(
    file_data: bytes,
    original_filename: str,
    user_id: int,
) -> Tuple[str, str, int]:
    """
    Сохранение загруженного файла.

    Args:
        file_data: Байты файла.
        original_filename: Исходное имя файла.
        user_id: ID пользователя.

    Returns:
        (stored_path, file_hash, file_size)
    """
    ext = Path(original_filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Недопустимый формат файла: {ext}. Разрешены: {', '.join(ALLOWED_EXTENSIONS.keys())}")

    file_size = len(file_data)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(f"Файл слишком большой: {file_size/1024/1024:.1f} MB. Максимум: {MAX_FILE_SIZE/1024/1024:.0f} MB")

    # Хеш для дедупликации
    file_hash = hashlib.sha256(file_data).hexdigest()[:16]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # user_id/YYYYMMDD_HHMMSS_hash.ext
    user_dir = UPLOAD_DIR / str(user_id)
    user_dir.mkdir(parents=True, exist_ok=True)

    safe_name = f"{timestamp}_{file_hash}{ext}"
    stored_path = user_dir / safe_name

    with open(stored_path, "wb") as f:
        f.write(file_data)

    logger.info("Файл сохранён: %s (%d KB)", stored_path, file_size // 1024)
    return str(stored_path), file_hash, file_size


def delete_uploaded_file(stored_path: str) -> bool:
    """Удаление загруженного файла."""
    try:
        path = Path(stored_path)
        if path.exists() and str(path).startswith(str(UPLOAD_DIR)):
            path.unlink()
            logger.info("Файл удалён: %s", stored_path)
            return True
    except Exception as e:
        logger.error("Ошибка удаления файла: %s", e)
    return False


# ─── AI-анализ документа ─────────────────────────────────────────────

async def analyze_document_with_ai(
    text: str,
    analysis_type: str = "summary",
) -> dict:
    """
    AI-анализ текста документа.

    Args:
        text: Извлечённый текст документа.
        analysis_type: Тип анализа —
            summary (краткое содержание),
            extract_fields (извлечение полей),
            dd_analysis (Due Diligence анализ),
            risk_assessment (оценка рисков).

    Returns:
        Результат анализа от AI-провайдера.
    """
    if not text or len(text.strip()) < 50:
        return {"error": "Текст слишком короткий для анализа", "result": ""}

    # Обрезаем до 4000 символов для AI
    truncated = text[:4000]

    prompts = {
        "summary": (
            "Ты — финансовый аналитик. Дай краткое содержание документа (3-5 пунктов). "
            "Выдели ключевые цифры, даты, участников."
        ),
        "extract_fields": (
            "Извлеки из документа следующие поля в формате JSON: "
            "company_name, inn, director, address, revenue, profit, employees, "
            "industry, registration_date. Если поле не найдено — null."
        ),
        "dd_analysis": (
            "Проведи Due Diligence анализ документа. Оцени: "
            "1) Финансовое состояние. 2) Юридические риски. "
            "3) Операционные риски. 4) Репутация. "
            "Дай оценку от 1 до 10 по каждому критерию."
        ),
        "risk_assessment": (
            "Оцени риски, описанные в документе. Классифицируй: "
            "Высокий/Средний/Низкий. Перечисли red flags если есть."
        ),
    }

    system_prompt = prompts.get(analysis_type, prompts["summary"])

    try:
        from app.services.ai_service import _call_with_fallback, AIProvider
        providers = [AIProvider.GROQ, AIProvider.GEMINI, AIProvider.OLLAMA]
        result, provider = await _call_with_fallback(
            providers, system_prompt, truncated, max_tokens=500,
        )
        return {
            "result": result or "AI не смог проанализировать документ",
            "provider": provider.value if provider else None,
            "analysis_type": analysis_type,
            "text_length": len(text),
        }
    except Exception as e:
        logger.error("Ошибка AI-анализа документа: %s", e)
        return {
            "result": "AI-сервис временно недоступен",
            "error": str(e),
            "analysis_type": analysis_type,
        }
