"""
Balance Sheet Parser for NSBU (Uzbekistan) format.
Parses Excel files with balance sheet data (lines 010-780).
"""
import io
import re
import logging
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)


@dataclass
class BalanceLineItem:
    code: str
    name_ru: str
    begin_period: float = 0.0
    end_period: float = 0.0


@dataclass
class ParsedBalance:
    company_name: str = ""
    period_date: str = ""
    lines: Dict[str, BalanceLineItem] = field(default_factory=dict)
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def total_assets_begin(self) -> float:
        return self._get_value("400", "begin")

    @property
    def total_assets_end(self) -> float:
        return self._get_value("400", "end")

    @property
    def total_liabilities_begin(self) -> float:
        return self._get_value("780", "begin")

    @property
    def total_liabilities_end(self) -> float:
        return self._get_value("780", "end")

    @property
    def is_balanced(self) -> bool:
        return abs(self.total_assets_end - self.total_liabilities_end) < 0.01

    def _get_value(self, code: str, period: str) -> float:
        item = self.lines.get(code)
        if not item:
            return 0.0
        return item.begin_period if period == "begin" else item.end_period

    def get_section(self, section: str) -> Dict[str, BalanceLineItem]:
        ranges = {
            "long_term_assets": ("010", "130"),
            "current_assets": ("140", "390"),
            "equity": ("410", "480"),
            "long_term_liabilities": ("490", "590"),
            "current_liabilities": ("600", "770"),
        }
        if section not in ranges:
            return {}
        start, end = ranges[section]
        return {
            k: v for k, v in self.lines.items()
            if start <= k <= end
        }


# NSBU Balance Sheet line codes mapping
NSBU_LINES: Dict[str, str] = {
    "010": "Основные средства (первоначальная стоимость)",
    "011": "износ",
    "012": "остаточная стоимость",
    "020": "Нематериальные активы (первоначальная стоимость)",
    "021": "амортизация",
    "022": "остаточная стоимость",
    "030": "Долгосрочные инвестиции",
    "040": "Оборудование к установке",
    "050": "Капитальные вложения",
    "060": "Долгосрочные отложенные расходы",
    "070": "Прочие основные средства",
    "080": "Долгосрочная дебиторская задолженность",
    "090": "Долгосрочные отложенные расходы",
    "100": "Прочие долгосрочные активы",
    "110": "Долгосрочные расходы будущих периодов",
    "120": "Прочие",
    "130": "ИТОГО по разделу I",
    "140": "Товарно-материальные запасы",
    "150": "Незавершённое производство",
    "160": "Готовая продукция",
    "170": "Товары",
    "180": "Расходы будущих периодов",
    "190": "Прочие текущие активы",
    "200": "Авансы выданные",
    "210": "Дебиторская задолженность",
    "220": "Денежные средства",
    "240": "Краткосрочные инвестиции",
    "250": "Прочие денежные средства",
    "260": "Краткосрочные финансовые инвестиции",
    "270": "Прочие текущие активы",
    "280": "Прочие текущие активы",
    "290": "НДС к возмещению",
    "300": "Авансовые платежи по налогам",
    "310": "Отложенные расходы",
    "320": "Дебиторская задолженность по выданным авансам",
    "330": "Задолженность учредителей",
    "340": "Дебиторская задолженность дочерних предприятий",
    "350": "Дебиторская задолженность персонала",
    "360": "Прочая дебиторская задолженность",
    "370": "Прочие текущие расходы",
    "380": "Прочие",
    "390": "ИТОГО по разделу II",
    "400": "ВСЕГО по активу баланса",
    "410": "Уставный капитал",
    "420": "Добавленный капитал",
    "430": "Резервный капитал",
    "440": "Выкупленные собственные акции",
    "450": "Нераспределённая прибыль",
    "460": "Целевые поступления",
    "470": "Резервы предстоящих расходов",
    "480": "ИТОГО по разделу I (пассив)",
    "490": "Долгосрочные обязательства",
    "500": "Долгосрочные кредиты банков",
    "520": "Долгосрочные займы",
    "530": "Прочие долгосрочные кредиторы",
    "540": "Долгосрочная кредиторская задолженность",
    "550": "Обязательства по отложенным налогам",
    "560": "Прочие долгосрочные обязательства",
    "570": "Краткосрочные кредиты банков",
    "580": "Краткосрочные займы",
    "590": "Текущая часть долгосрочных обязательств",
    "600": "Кредиторская задолженность",
    "610": "Задолженность поставщикам",
    "630": "Авансы полученные",
    "640": "Задолженность по платежам в бюджет",
    "650": "Задолженность по страхованию",
    "660": "Задолженность по оплате труда",
    "670": "Задолженность учредителям",
    "680": "Задолженность дочерним предприятиям",
    "690": "Прочая кредиторская задолженность",
    "700": "Отложенные доходы",
    "710": "Резервы предстоящих расходов",
    "720": "Прочие текущие обязательства",
    "730": "Текущая часть долгосрочных обязательств",
    "740": "Прочие обязательства",
    "750": "Прочие краткосрочные обязательства",
    "760": "Прочие",
    "770": "ИТОГО по разделу II (пассив)",
    "780": "ВСЕГО по пассиву баланса",
}


def parse_number(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "").replace("\xa0", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def extract_line_code(val) -> Optional[str]:
    if val is None:
        return None
    s = re.sub(r"[^0-9]", "", str(val).strip())
    if len(s) == 3 and s.isdigit():
        return s
    if len(s) >= 3:
        return s[:3]
    return None


def parse_balance_xlsx(content: bytes, sheet_name: Optional[str] = None) -> ParsedBalance:
    """Parse NSBU balance sheet from Excel bytes."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    result = ParsedBalance()

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        result.errors.append(f"Cannot open file: {str(e)}")
        return result

    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([c for c in row])
    wb.close()

    if not all_rows:
        result.errors.append("File is empty")
        return result

    # Try to find company name in first 5 rows
    for row in all_rows[:5]:
        for cell in row:
            if cell and len(str(cell)) > 5:
                txt = str(cell).strip()
                if any(kw in txt.lower() for kw in ["ooo", "ao", "ооо", "ао", "чп", "гуп", "llc", "solutions"]):
                    result.company_name = txt
                    break
                if any(kw in txt.lower() for kw in ["баланс", "форма", "balance"]):
                    # Try to find date
                    for cell2 in row:
                        if cell2 and re.search(r"\d{2}[./]\d{2}[./]\d{4}", str(cell2)):
                            result.period_date = str(cell2).strip()

    # Find columns: code_col, begin_col, end_col
    code_col = None
    begin_col = None
    end_col = None

    # Strategy 1: find header row with keywords
    for row_idx, row in enumerate(all_rows):
        row_str = [str(c).lower().strip() if c else "" for c in row]
        for col_idx, cell_str in enumerate(row_str):
            if any(kw in cell_str for kw in ["код стр", "коды стр", "код\nстр"]):
                code_col = col_idx
            elif any(kw in cell_str for kw in ["на начало", "начало периода", "нач. пер", "begin"]):
                if begin_col is None:
                    begin_col = col_idx
            elif any(kw in cell_str for kw in ["на конец", "конец периода", "кон. пер", "end"]):
                if end_col is None:
                    end_col = col_idx
        if code_col is not None:
            break

    # Strategy 2: find column that contains 3-digit codes like 010, 020, etc.
    if code_col is None:
        for col_idx in range(min(5, len(all_rows[0]) if all_rows else 0)):
            code_count = 0
            for row in all_rows:
                if col_idx < len(row):
                    code = extract_line_code(row[col_idx])
                    if code and code in NSBU_LINES:
                        code_count += 1
            if code_count >= 5:
                code_col = col_idx
                break

    if code_col is None:
        result.errors.append("Cannot find line codes column")
        return result

    # Find begin/end columns by position relative to code column
    if begin_col is None or end_col is None:
        num_cols = max(len(r) for r in all_rows)
        for col_idx in range(code_col + 1, min(num_cols, code_col + 5)):
            has_numbers = 0
            for row in all_rows[3:]:
                if col_idx < len(row) and row[col_idx] is not None:
                    v = parse_number(row[col_idx])
                    if v != 0:
                        has_numbers += 1
            if has_numbers >= 5:
                if begin_col is None:
                    begin_col = col_idx
                elif end_col is None:
                    end_col = col_idx
                    break

    if begin_col is None:
        result.errors.append("Cannot find begin period column")
        return result
    if end_col is None:
        end_col = begin_col + 1
        result.warnings.append("End period column guessed")

    # Parse rows
    for row in all_rows:
        if code_col >= len(row):
            continue
        code = extract_line_code(row[code_col])
        if not code:
            continue
        if code not in NSBU_LINES:
            continue

        begin_val = parse_number(row[begin_col] if begin_col < len(row) else 0)
        end_val = parse_number(row[end_col] if end_col < len(row) else 0)

        name = NSBU_LINES.get(code, "")

        result.lines[code] = BalanceLineItem(
            code=code,
            name_ru=name,
            begin_period=begin_val,
            end_period=end_val,
        )

    if not result.lines:
        result.errors.append("No balance lines found in file")
        return result

    # Validation
    if "400" in result.lines and "780" in result.lines:
        diff = abs(result.total_assets_end - result.total_liabilities_end)
        if diff > 1:
            result.warnings.append(
                f"Balance mismatch: Assets={result.total_assets_end:,.0f}, "
                f"Liabilities={result.total_liabilities_end:,.0f}, Diff={diff:,.0f}"
            )

    logger.info(
        f"Parsed {len(result.lines)} lines, "
        f"balanced={result.is_balanced}, "
        f"company={result.company_name}"
    )
    return result


def balance_to_entries(parsed: ParsedBalance, period: str = "end") -> List[dict]:
    """Convert ParsedBalance to list of balance_entries dicts."""
    entries = []
    for code, item in sorted(parsed.lines.items()):
        val = item.end_period if period == "end" else item.begin_period
        entries.append({
            "account_code": code,
            "account_name": item.name_ru,
            "debit": val if val >= 0 else 0,
            "credit": abs(val) if val < 0 else 0,
            "balance": val,
        })
    return entries


def balance_to_wizard_data(parsed: ParsedBalance) -> dict:
    """Convert ParsedBalance to wizard step data for frontend."""
    def _get(code: str) -> dict:
        item = parsed.lines.get(code)
        if not item:
            return {"begin": 0, "end": 0}
        return {"begin": item.begin_period, "end": item.end_period}

    return {
        "company_name": parsed.company_name,
        "period_date": parsed.period_date,
        "is_balanced": parsed.is_balanced,
        "total_assets": _get("400"),
        "total_liabilities": _get("780"),
        "long_term_assets": {
            "fixed_assets": _get("012"),
            "intangible_assets": _get("022"),
            "long_term_investments": _get("030"),
            "equipment": _get("040"),
            "capital_investments": _get("050"),
            "long_term_receivables": _get("080"),
            "other_long_term": _get("100"),
            "total_section_1": _get("130"),
        },
        "current_assets": {
            "inventory": _get("140"),
            "wip": _get("150"),
            "finished_goods": _get("160"),
            "goods": _get("170"),
            "receivables": _get("210"),
            "cash": _get("220"),
            "short_term_investments": _get("240"),
            "total_section_2": _get("390"),
        },
        "equity": {
            "charter_capital": _get("410"),
            "additional_capital": _get("420"),
            "reserve_capital": _get("430"),
            "retained_earnings": _get("450"),
            "target_receipts": _get("460"),
            "total_equity": _get("480"),
        },
        "liabilities": {
            "long_term_obligations": _get("490"),
            "long_term_bank_credits": _get("500"),
            "short_term_bank_credits": _get("570"),
            "short_term_loans": _get("580"),
            "accounts_payable": _get("600"),
            "payable_suppliers": _get("610"),
            "payable_budget": _get("640"),
            "payable_salary": _get("660"),
            "other_current_liabilities": _get("720"),
            "total_liabilities_2": _get("770"),
        },
        "lines": {
            code: {
                "name": item.name_ru,
                "begin": item.begin_period,
                "end": item.end_period,
            }
            for code, item in sorted(parsed.lines.items())
        },
        "errors": parsed.errors,
        "warnings": parsed.warnings,
    }
