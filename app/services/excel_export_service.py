"""
Сервис Excel-экспорта — openpyxl с формулами и графиками.

Фаза 3, EXPORT-002:
  - Экспорт портфеля: Summary, Holdings, Analytics
  - Формулы: SUM, AVERAGE, процентные расчёты
  - PieChart для распределения активов
  - Условное форматирование (зелёный/красный P&L)
  - Фиксация строки заголовка (freeze panes)
  - Брендированные стили ячеек
"""

import io
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers,
    )
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl не установлен — Excel-экспорт недоступен")


# ── Цвета бренда ─────────────────────────────────────────────

BRAND_DARK_HEX = "1A2332"
BRAND_PRIMARY_HEX = "3B82F6"
BRAND_GREEN_HEX = "22C55E"
BRAND_RED_HEX = "EF4444"
BRAND_LIGHT_HEX = "F8FAFC"
WHITE_HEX = "FFFFFF"


def _create_styles(wb: "Workbook"):
    """Создать брендированные стили ячеек."""
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # Заголовок (тёмный фон, белый текст)
    header_style = NamedStyle(name="brand_header")
    header_style.font = Font(name="Calibri", size=11, bold=True, color=WHITE_HEX)
    header_style.fill = PatternFill(start_color=BRAND_DARK_HEX, end_color=BRAND_DARK_HEX, fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = thin_border
    wb.add_named_style(header_style)

    # Данные (обычные)
    data_style = NamedStyle(name="brand_data")
    data_style.font = Font(name="Calibri", size=10, color="1E293B")
    data_style.alignment = Alignment(vertical="center")
    data_style.border = thin_border
    wb.add_named_style(data_style)

    # Итого (полужирный)
    total_style = NamedStyle(name="brand_total")
    total_style.font = Font(name="Calibri", size=11, bold=True, color=BRAND_DARK_HEX)
    total_style.fill = PatternFill(start_color=BRAND_LIGHT_HEX, end_color=BRAND_LIGHT_HEX, fill_type="solid")
    total_style.alignment = Alignment(vertical="center")
    total_style.border = thin_border
    wb.add_named_style(total_style)

    # Числовые с разделителем
    num_style = NamedStyle(name="brand_number")
    num_style.font = Font(name="Calibri", size=10, color="1E293B")
    num_style.alignment = Alignment(horizontal="right", vertical="center")
    num_style.border = thin_border
    num_style.number_format = '#,##0.00'
    wb.add_named_style(num_style)

    # Процент
    pct_style = NamedStyle(name="brand_percent")
    pct_style.font = Font(name="Calibri", size=10, color="1E293B")
    pct_style.alignment = Alignment(horizontal="right", vertical="center")
    pct_style.border = thin_border
    pct_style.number_format = '0.00%'
    wb.add_named_style(pct_style)


def _add_brand_header_row(ws, title: str, cols: int):
    """Добавить брендированную строку заголовка в лист."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=1, value=f"AI Capital Management — {title}")
    cell.font = Font(name="Calibri", size=14, bold=True, color=BRAND_PRIMARY_HEX)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    cell2 = ws.cell(row=2, column=1, value=f"Сформировано: {now}")
    cell2.font = Font(name="Calibri", size=9, italic=True, color="64748B")

    # Пустая строка
    ws.row_dimensions[3].height = 6


class ExcelExportService:
    """Сервис экспорта в Excel."""

    @staticmethod
    def export_portfolio_summary(portfolio_data: dict) -> bytes:
        """
        Экспорт портфеля в xlsx.

        Листы:
          1. Summary — название, стоимость, ROI, дата
          2. Holdings — таблица активов с формулами
          3. Analytics — DCF/NPV/IRR (если доступны)
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl не установлен")

        wb = Workbook()
        _create_styles(wb)

        # ── Sheet 1: Summary ──
        ws1 = wb.active
        ws1.title = "Сводка"
        _add_brand_header_row(ws1, "Сводка по портфелю", 4)

        summary_headers = ["Параметр", "Значение"]
        for col, h in enumerate(summary_headers, 1):
            cell = ws1.cell(row=4, column=col, value=h)
            cell.style = "brand_header"

        summary_data = [
            ("Название портфеля", portfolio_data.get("name", "—")),
            ("Валюта", portfolio_data.get("currency", "USD")),
            ("Общая стоимость", portfolio_data.get("total_value", 0)),
            ("Количество активов", len(portfolio_data.get("assets", []))),
            ("ROI", portfolio_data.get("roi", "N/A")),
            ("Владелец", portfolio_data.get("owner", "—")),
        ]
        for i, (param, value) in enumerate(summary_data, 5):
            ws1.cell(row=i, column=1, value=param).style = "brand_data"
            c = ws1.cell(row=i, column=2, value=value)
            c.style = "brand_number" if isinstance(value, (int, float)) else "brand_data"

        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 25
        ws1.freeze_panes = "A5"

        # ── Sheet 2: Holdings ──
        ws2 = wb.create_sheet("Активы")
        _add_brand_header_row(ws2, "Состав портфеля", 7)

        headers = ["Компания", "Тикер", "Кол-во", "Цена", "Стоимость", "Доля %", "P&L"]
        for col, h in enumerate(headers, 1):
            ws2.cell(row=4, column=col, value=h).style = "brand_header"

        assets = portfolio_data.get("assets", [])
        for i, asset in enumerate(assets, 5):
            ws2.cell(row=i, column=1, value=asset.get("name", "")).style = "brand_data"
            ws2.cell(row=i, column=2, value=asset.get("symbol", "")).style = "brand_data"
            ws2.cell(row=i, column=3, value=asset.get("quantity", 0)).style = "brand_number"
            ws2.cell(row=i, column=4, value=asset.get("price", 0)).style = "brand_number"
            # Формула: Стоимость = Кол-во × Цена
            ws2.cell(row=i, column=5).value = f"=C{i}*D{i}"
            ws2.cell(row=i, column=5).style = "brand_number"
            ws2.cell(row=i, column=6, value=asset.get("weight", 0) / 100 if isinstance(asset.get("weight"), (int, float)) else 0).style = "brand_percent"
            pnl = asset.get("pnl", 0)
            ws2.cell(row=i, column=7, value=pnl).style = "brand_number"

        # Итого
        if assets:
            total_row = 5 + len(assets)
            ws2.cell(row=total_row, column=1, value="ИТОГО").style = "brand_total"
            for col in range(2, 5):
                ws2.cell(row=total_row, column=col).style = "brand_total"
            # Формула SUM для стоимости
            ws2.cell(row=total_row, column=5).value = f"=SUM(E5:E{total_row-1})"
            ws2.cell(row=total_row, column=5).style = "brand_total"
            # Средняя доля
            ws2.cell(row=total_row, column=6).value = f"=SUM(F5:F{total_row-1})"
            ws2.cell(row=total_row, column=6).style = "brand_total"
            # Сумма P&L
            ws2.cell(row=total_row, column=7).value = f"=SUM(G5:G{total_row-1})"
            ws2.cell(row=total_row, column=7).style = "brand_total"

            # Условное форматирование P&L: зелёный > 0, красный < 0
            green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            pnl_range = f"G5:G{total_row}"
            ws2.conditional_formatting.add(
                pnl_range,
                CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill)
            )
            ws2.conditional_formatting.add(
                pnl_range,
                CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
            )

        # Ширина колонок
        widths = [22, 10, 12, 14, 16, 10, 14]
        for i, w in enumerate(widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = "A5"

        # ── PieChart для распределения активов ──
        if assets and len(assets) >= 2:
            chart = PieChart()
            chart.title = "Распределение активов"
            chart.style = 26
            chart.width = 15
            chart.height = 10

            data_ref = Reference(ws2, min_col=5, min_row=4, max_row=4 + len(assets))
            cats_ref = Reference(ws2, min_col=1, min_row=5, max_row=4 + len(assets))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showVal = False

            ws2.add_chart(chart, f"I4")

        # ── Sheet 3: Analytics ──
        analytics = portfolio_data.get("analytics", {})
        if analytics:
            ws3 = wb.create_sheet("Аналитика")
            _add_brand_header_row(ws3, "Инвестиционная аналитика", 4)

            headers3 = ["Метрика", "Значение", "Единица", "Комментарий"]
            for col, h in enumerate(headers3, 1):
                ws3.cell(row=4, column=col, value=h).style = "brand_header"

            metrics = [
                ("NPV", analytics.get("npv", "N/A"), "USD", "Чистая приведённая стоимость"),
                ("IRR", analytics.get("irr", "N/A"), "%", "Внутренняя ставка доходности"),
                ("Payback", analytics.get("payback", "N/A"), "лет", "Срок окупаемости"),
                ("WACC", analytics.get("wacc", "N/A"), "%", "Средневзвешенная стоимость капитала"),
                ("DCF", analytics.get("dcf", "N/A"), "USD", "Дисконтированный денежный поток"),
            ]
            for i, (metric, value, unit, comment) in enumerate(metrics, 5):
                ws3.cell(row=i, column=1, value=metric).style = "brand_data"
                c = ws3.cell(row=i, column=2, value=value)
                c.style = "brand_number" if isinstance(value, (int, float)) else "brand_data"
                ws3.cell(row=i, column=3, value=unit).style = "brand_data"
                ws3.cell(row=i, column=4, value=comment).style = "brand_data"

            for col, w in enumerate([18, 16, 10, 35], 1):
                ws3.column_dimensions[get_column_letter(col)].width = w
            ws3.freeze_panes = "A5"

        # Сохранение
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def export_dd_report(dd_data: dict) -> bytes:
        """Экспорт DD-отчёта в xlsx."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl не установлен")

        wb = Workbook()
        _create_styles(wb)

        ws = wb.active
        ws.title = "DD Report"
        _add_brand_header_row(ws, "Due Diligence Report", 6)

        # Компания
        ws.cell(row=4, column=1, value="Компания").style = "brand_header"
        ws.cell(row=4, column=2, value=dd_data.get("company_name", "")).style = "brand_data"
        ws.cell(row=5, column=1, value="ИНН").style = "brand_header"
        ws.cell(row=5, column=2, value=dd_data.get("inn", "")).style = "brand_data"

        # Скоринг
        scores = dd_data.get("scores", {})
        if scores:
            ws.cell(row=7, column=1, value="Категория").style = "brand_header"
            ws.cell(row=7, column=2, value="Оценка").style = "brand_header"
            ws.cell(row=7, column=3, value="Статус").style = "brand_header"

            for i, (cat, score) in enumerate(scores.items(), 8):
                ws.cell(row=i, column=1, value=cat).style = "brand_data"
                ws.cell(row=i, column=2, value=score).style = "brand_number"
                s = int(score) if isinstance(score, (int, float)) else 0
                status = "Низкий риск" if s >= 7 else ("Средний риск" if s >= 4 else "Высокий риск")
                ws.cell(row=i, column=3, value=status).style = "brand_data"

        # Red flags
        flags = dd_data.get("red_flags", [])
        if flags:
            row_offset = 8 + len(scores) + 1
            ws.cell(row=row_offset, column=1, value="Red Flags").style = "brand_header"
            for i, flag in enumerate(flags, row_offset + 1):
                ws.cell(row=i, column=1, value=flag).style = "brand_data"

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def export_comparison(companies: list[dict]) -> bytes:
        """Экспорт сравнительной таблицы компаний."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl не установлен")

        wb = Workbook()
        _create_styles(wb)

        ws = wb.active
        ws.title = "Сравнение"
        _add_brand_header_row(ws, "Сравнительный анализ компаний", 8)

        headers = ["Компания", "ИНН", "Директор", "ОКЭД", "Уст. фонд", "Статус", "Адрес", "Телефон"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=h).style = "brand_header"

        for i, company in enumerate(companies, 5):
            ws.cell(row=i, column=1, value=company.get("name", "")).style = "brand_data"
            ws.cell(row=i, column=2, value=company.get("inn", "")).style = "brand_data"
            ws.cell(row=i, column=3, value=company.get("director", "")).style = "brand_data"
            ws.cell(row=i, column=4, value=company.get("oked", "")).style = "brand_data"
            ws.cell(row=i, column=5, value=company.get("charter_fund", "")).style = "brand_data"
            ws.cell(row=i, column=6, value=company.get("status", "")).style = "brand_data"
            ws.cell(row=i, column=7, value=company.get("address", "")).style = "brand_data"
            ws.cell(row=i, column=8, value=company.get("phone", "")).style = "brand_data"

        widths = [25, 12, 22, 15, 15, 15, 30, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A5"

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
