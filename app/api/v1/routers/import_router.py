import io
import csv
import re
from typing import List, Optional, Dict
from datetime import datetime, date

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from sqlalchemy import text
from pydantic import BaseModel
import httpx

from app.api.v1.deps import get_db

router = APIRouter(tags=["import"])


# ==================== Pydantic Models ====================

class ImportResult(BaseModel):
    status: str
    records_imported: int
    records_skipped: int
    errors: List[str]
    timestamp: str

class BalanceRow(BaseModel):
    account_code: str
    account_name: str
    debit: float = 0
    credit: float = 0
    balance: float = 0

class OneCConfig(BaseModel):
    base_url: str
    username: str
    password: str
    company_name: Optional[str] = None

class OneCTestResult(BaseModel):
    status: str
    message: str
    company_name: Optional[str] = None
    accounts_found: int = 0


# ==================== HELPERS ====================

def get_or_create_account(db: Session, code: str, name: str = "") -> Optional[int]:
    """Find or create chart_of_accounts entry, return its id."""
    row = db.execute(
        text("SELECT id FROM chart_of_accounts WHERE code = :code"),
        {"code": code}
    ).fetchone()
    if row:
        return row[0]
    # Determine category based on code
    code_num = int(re.sub(r'[^0-9]', '', code) or '0')
    if code_num < 1000:
        category = 'long_term_assets'
    elif code_num < 6000:
        category = 'current_assets'
    elif code_num < 8000:
        category = 'liabilities'
    elif code_num < 9000:
        category = 'equity'
    else:
        category = 'income'
    try:
        db.execute(text("""
            INSERT INTO chart_of_accounts (code, name_ru, category, level, is_active)
            VALUES (:code, :name, :cat, 1, true)
            ON CONFLICT (code) DO NOTHING
        """), {"code": code, "name": name or code, "cat": category})
        db.flush()
        row = db.execute(
            text("SELECT id FROM chart_of_accounts WHERE code = :code"),
            {"code": code}
        ).fetchone()
        return row[0] if row else None
    except Exception:
        return None


def save_balance_entries(
    db: Session, org_id: int, rows_data: List[BalanceRow],
    period_date: Optional[str] = None
) -> tuple:
    """Save BalanceRow list to balance_entries via chart_of_accounts lookup."""
    records_imported = 0
    records_skipped = 0
    errors = []
    p_date = period_date or date.today().isoformat()

    # Delete old entries for this org + period
    db.execute(
        text("DELETE FROM balance_entries WHERE organization_id = :oid AND period_date = :pd"),
        {"oid": org_id, "pd": p_date}
    )

    for row in rows_data:
        try:
            account_id = get_or_create_account(db, row.account_code, row.account_name)
            if not account_id:
                records_skipped += 1
                errors.append(f"Cannot create account {row.account_code}")
                continue
            db.execute(text("""
                INSERT INTO balance_entries
                    (organization_id, account_id, period_date, debit, credit, balance, source)
                VALUES (:oid, :aid, :pd, :d, :c, :b, 'import')
            """), {
                "oid": org_id,
                "aid": account_id,
                "pd": p_date,
                "d": row.debit,
                "c": row.credit,
                "b": row.balance,
            })
            records_imported += 1
        except Exception as e:
            records_skipped += 1
            errors.append(f"Account {row.account_code}: {str(e)}")

    db.commit()
    return records_imported, records_skipped, errors


# ==================== PARSE HELPERS ====================

def normalize_account_code(raw: str) -> Optional[str]:
    """Extract account code from string."""
    if not raw:
        return None
    clean = re.sub(r"[^0-9]", "", str(raw).strip())
    if len(clean) >= 3:
        return clean[:4] if len(clean) >= 4 else clean
    return None

def parse_number(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "").replace("\xa0", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0

def detect_columns(headers: List[str]) -> dict:
    mapping = {"code": None, "name": None, "debit": None, "credit": None, "balance": None}
    for i, h in enumerate(headers):
        hl = str(h).lower().strip() if h else ""
        if any(kw in hl for kw in ["счет", "счёт", "код", "code", "account"]):
            if mapping["code"] is None: mapping["code"] = i
        elif any(kw in hl for kw in ["наименование", "название", "name"]):
            if mapping["name"] is None: mapping["name"] = i
        elif any(kw in hl for kw in ["дебет", "debit", "дб", "dt"]):
            if mapping["debit"] is None: mapping["debit"] = i
        elif any(kw in hl for kw in ["кредит", "credit", "кр", "kt", "cr"]):
            if mapping["credit"] is None: mapping["credit"] = i
        elif any(kw in hl for kw in ["сальдо", "balance", "остаток", "итого"]):
            if mapping["balance"] is None: mapping["balance"] = i
    if mapping["code"] is None and len(headers) >= 1: mapping["code"] = 0
    if mapping["name"] is None and len(headers) >= 2: mapping["name"] = 1
    if mapping["debit"] is None and len(headers) >= 3: mapping["debit"] = 2
    if mapping["credit"] is None and len(headers) >= 4: mapping["credit"] = 3
    if mapping["balance"] is None: mapping["balance"] = mapping.get("debit")
    return mapping


# ==================== EXCEL / CSV IMPORT ====================

@router.post("/organizations/{org_id}/import/excel", response_model=ImportResult)
async def import_excel(
    org_id: int,
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Query(None),
    header_row: int = Query(1),
    period_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Import trial balance from XLSX/XLS/CSV."""
    org = db.execute(text("SELECT id FROM organizations WHERE id = :id"), {"id": org_id}).fetchone()
    if not org:
        raise HTTPException(404, "Organization not found")

    content = await file.read()
    filename = file.filename or ""
    errors = []
    rows_data: List[BalanceRow] = []

    if filename.lower().endswith((".xlsx", ".xls")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(500, "openpyxl not installed")
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        if len(all_rows) < 2:
            raise HTTPException(400, "File empty")
        headers = [str(h) if h else "" for h in all_rows[header_row - 1]]
        mapping = detect_columns(headers)
        for row_idx, row in enumerate(all_rows[header_row:], start=header_row + 1):
            try:
                code_raw = row[mapping["code"]] if mapping["code"] is not None and mapping["code"] < len(row) else None
                code = normalize_account_code(code_raw)
                if not code: continue
                name = row[mapping["name"]] if mapping["name"] is not None and mapping["name"] < len(row) else ""
                debit = parse_number(row[mapping["debit"]] if mapping["debit"] is not None and mapping["debit"] < len(row) else 0)
                credit = parse_number(row[mapping["credit"]] if mapping["credit"] is not None and mapping["credit"] < len(row) else 0)
                balance = debit - credit
                if mapping["balance"] is not None and mapping["balance"] < len(row) and mapping["balance"] != mapping["debit"]:
                    balance = parse_number(row[mapping["balance"]])
                rows_data.append(BalanceRow(account_code=code, account_name=str(name).strip() if name else "", debit=debit, credit=credit, balance=balance))
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
    elif filename.lower().endswith((".csv", ".tsv")):
        try:
            text_content = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text_content = content.decode("cp1251", errors="replace")
        delimiter = "\t" if filename.lower().endswith(".tsv") else ","
        if ";" in text_content[:500] and "," not in text_content[:500]: delimiter = ";"
        reader = csv.reader(io.StringIO(text_content), delimiter=delimiter)
        all_rows = list(reader)
        if len(all_rows) < 2:
            raise HTTPException(400, "File empty")
        headers = all_rows[header_row - 1]
        mapping = detect_columns(headers)
        for row_idx, row in enumerate(all_rows[header_row:], start=header_row + 1):
            try:
                code_raw = row[mapping["code"]] if mapping["code"] is not None and mapping["code"] < len(row) else None
                code = normalize_account_code(code_raw)
                if not code: continue
                name = row[mapping["name"]] if mapping["name"] is not None and mapping["name"] < len(row) else ""
                debit = parse_number(row[mapping["debit"]] if mapping["debit"] is not None and mapping["debit"] < len(row) else 0)
                credit = parse_number(row[mapping["credit"]] if mapping["credit"] is not None and mapping["credit"] < len(row) else 0)
                balance = debit - credit
                rows_data.append(BalanceRow(account_code=code, account_name=str(name).strip(), debit=debit, credit=credit, balance=balance))
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
    else:
        raise HTTPException(400, "Supported formats: .xlsx, .xls, .csv, .tsv")

    if not rows_data:
        raise HTTPException(400, f"No data extracted. Errors: {'; '.join(errors[:5])}")

    records_imported, records_skipped, save_errors = save_balance_entries(db, org_id, rows_data, period_date)
    errors.extend(save_errors)

    return ImportResult(
        status="success",
        records_imported=records_imported,
        records_skipped=records_skipped,
        errors=errors[:10],
        timestamp=datetime.now().isoformat()
    )


@router.get("/import/template")
async def download_import_template():
    """Download Excel import template."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance"
    headers = ["Code", "Account Name", "Debit", "Credit", "Balance"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    sample = [
        ["0100", "Fixed assets", 150000000, 0, 150000000],
        ["1000", "Materials", 28000000, 0, 28000000],
        ["5100", "Bank account", 89000000, 0, 89000000],
        ["6000", "Accounts payable", 0, 42000000, -42000000],
        ["8300", "Charter capital", 0, 100000000, -100000000],
    ]
    for ri, data in enumerate(sample, 2):
        for ci, val in enumerate(data, 1):
            ws.cell(row=ri, column=ci, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import Response
    return Response(content=buf.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=template_balance.xlsx"})


# ==================== 1C OData ====================

@router.post("/import/1c/test", response_model=OneCTestResult)
async def test_1c_connection(config: OneCConfig):
    try:
        async with httpx.AsyncClient(timeout=15.0, verify=False) as client:
            url = f"{config.base_url.rstrip('/')}/$metadata"
            resp = await client.get(url, auth=(config.username, config.password))
            if resp.status_code == 401:
                return OneCTestResult(status="error", message="Auth error")
            if resp.status_code != 200:
                return OneCTestResult(status="error", message=f"HTTP {resp.status_code}")
            return OneCTestResult(status="success", message="Connected", company_name=config.company_name)
    except Exception as e:
        return OneCTestResult(status="error", message=str(e))

@router.post("/organizations/{org_id}/import/1c", response_model=ImportResult)
async def import_from_1c(org_id: int, config: OneCConfig, db: Session = Depends(get_db)):
    raise HTTPException(501, "1C import not yet implemented")

@router.get("/import/1c/endpoints")
async def get_1c_endpoints():
    return {"description": "1C OData endpoints info", "status": "placeholder"}


# ==================== NSBU BALANCE SHEET IMPORT ====================

@router.post("/organizations/{org_id}/import/balance-nsbu")
async def import_balance_nsbu(
    org_id: int,
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Import NSBU balance sheet (lines 010-780) from Excel."""
    from app.services.balance_parser import (
        parse_balance_xlsx, balance_to_entries, balance_to_wizard_data
    )
    org = db.execute(
        text("SELECT id FROM organizations WHERE id = :id"),
        {"id": org_id}
    ).fetchone()
    if not org:
        raise HTTPException(404, "Organization not found")

    content = await file.read()
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls supported")

    parsed = parse_balance_xlsx(content, sheet_name)
    if parsed.errors and not parsed.lines:
        raise HTTPException(400, f"Parse failed: {'; '.join(parsed.errors)}")

    # Convert to BalanceRow list
    entries = balance_to_entries(parsed, period="end")
    rows_data = []
    for entry in entries:
        rows_data.append(BalanceRow(
            account_code=entry["account_code"],
            account_name=entry["account_name"],
            debit=entry["debit"],
            credit=entry["credit"],
            balance=entry["balance"],
        ))

    # Save using the shared helper
    records_imported, records_skipped, save_errors = save_balance_entries(
        db, org_id, rows_data
    )

    wizard_data = balance_to_wizard_data(parsed)
    wizard_data["import_result"] = {
        "status": "success",
        "records_imported": records_imported,
        "records_skipped": records_skipped,
        "errors": save_errors[:5],
        "timestamp": datetime.now().isoformat(),
    }
    return wizard_data


@router.post("/organizations/{org_id}/import/balance-nsbu/preview")
async def preview_balance_nsbu(
    org_id: int,
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Query(None),
):
    """Preview NSBU balance without saving."""
    from app.services.balance_parser import (
        parse_balance_xlsx, balance_to_wizard_data
    )
    content = await file.read()
    parsed = parse_balance_xlsx(content, sheet_name)
    return balance_to_wizard_data(parsed)
