"""
Analytics chapter API — KPI, DCF, multiples, stress-test, visualizations, decisions.
Uses real data from _portfolio_cache when available.
E2-08: Full NSBU+IFRS Excel export endpoint.
"""
import io
import logging
import math
import traceback
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

from app.api.v1.routers.portfolios import _portfolio_cache, _user_cache, _build_nsbu_rows, _build_ifrs_rows, _build_diff_rows
from app.api.v1.deps import get_db, get_current_user
from app.db.models.user import User
from sqlalchemy.orm import Session

router = APIRouter(tags=["analytics-chapter"])


# ---------------------------------------------------------------------------
# Helpers: extract balance aggregates from cache
# ---------------------------------------------------------------------------

def _get_balance_aggregates(user_id: Optional[int] = None) -> Optional[dict]:
    """Calculate balance aggregates from cached accounts + pnl data.

    When user_id is given, reads from the per-user cache partition.
    Falls back to scanning all users if user_id is None (backward-compat).
    """
    if user_id is not None:
        cache = _user_cache(user_id)
        accounts = cache.get("accounts")
        pnl = cache.get("pnl", {})
    else:
        # Backward-compat: scan all user caches for any with accounts
        cache = {}
        accounts = None
        pnl = {}
        for uid, ucache in _portfolio_cache.items():
            if ucache.get("accounts"):
                cache = ucache
                accounts = ucache["accounts"]
                pnl = ucache.get("pnl", {})
                break

    if not accounts:
        return None

    def _v(code: str, field: str = "current") -> float:
        acc = accounts.get(code)
        return acc[field] if acc else 0.0

    net_fa = _v("0100") - _v("0200")
    capex = _v("0800")
    non_current_assets = net_fa + capex

    inventories = _v("1000") + _v("1010")
    receivables = _v("2010") + _v("2300")
    cash = _v("5010") + _v("5110") + _v("5210")
    current_assets = inventories + receivables + cash
    total_assets = non_current_assets + current_assets

    equity_base = _v("8300") + _v("8500") + _v("8700")
    lt_liabilities = _v("7010") + _v("7800")
    st_liabilities = _v("6010") + _v("6110") + _v("6310") + _v("6710") + _v("6820") + _v("6610")
    total_liabilities = lt_liabilities + st_liabilities

    # Unclosed profit adjustment
    unclosed_profit = total_assets - (equity_base + total_liabilities)
    total_equity = equity_base + unclosed_profit

    revenue = pnl.get("total_revenue_end", 0.0)
    expenses = pnl.get("total_expenses_end", 0.0)

    # Detailed P&L breakdown — populated from income_expenses
    cost_of_goods = 0.0
    operating_expenses = 0.0
    other_expenses = 0.0
    tax = 0.0
    gross_profit = 0.0
    operating_profit = 0.0
    profit_before_tax = 0.0

    # Fallback 1: income_expenses list from parsed P&L (most reliable source)
    # Always try when revenue==0 because import may write 0 even when IE data exists
    ie_list = cache.get("income_expenses", [])
    if ie_list:
        ie_result = _revenue_costs_from_income_expenses(ie_list)
        if ie_result["revenue_cur"] > 0:
            revenue = ie_result["revenue_cur"]
        cost_of_goods = ie_result["costs_cur"]
        operating_expenses = ie_result["opex_cur"]
        other_expenses = ie_result["other_expenses_cur"]
        tax = ie_result["tax_cur"]
        if expenses == 0 and (cost_of_goods + operating_expenses) > 0:
            expenses = cost_of_goods + operating_expenses + other_expenses

    # Fallback 2: extract from ОСВ balance accounts (less reliable for P&L)
    if revenue == 0 and expenses == 0:
        for code, acc in accounts.items():
            code_str = str(code)
            cur = acc.get("credit_end", 0) or acc.get("current", 0)
            cur_d = acc.get("debit_end", 0) or acc.get("current", 0)
            if code_str.startswith("90"):
                revenue += abs(cur) if cur else abs(cur_d)
            elif code_str.startswith(("20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "44")):
                expenses += abs(cur_d) if cur_d else abs(cur)
            elif code_str.startswith(("91", "92", "93", "94", "95", "96", "97", "98", "99")):
                expenses += abs(cur_d) if cur_d else abs(cur)

    # Compute detailed P&L breakdown
    gross_profit = revenue - cost_of_goods
    operating_profit = gross_profit - operating_expenses
    profit_before_tax = operating_profit - other_expenses
    if tax == 0 and profit_before_tax > 0:
        tax = round(profit_before_tax * 0.15, 2)
    net_profit = profit_before_tax - tax

    # Previous period
    net_fa_p = _v("0100", "previous") - _v("0200", "previous")
    capex_p = _v("0800", "previous")
    inventories_p = _v("1000", "previous") + _v("1010", "previous")
    receivables_p = _v("2010", "previous") + _v("2300", "previous")
    cash_p = _v("5010", "previous") + _v("5110", "previous") + _v("5210", "previous")
    non_current_assets_p = net_fa_p + capex_p
    current_assets_p = inventories_p + receivables_p + cash_p
    total_assets_p = non_current_assets_p + current_assets_p

    return {
        "non_current_assets": non_current_assets,
        "current_assets": current_assets,
        "inventories": inventories,
        "receivables": receivables,
        "cash": cash,
        "total_assets": total_assets,
        "total_equity": total_equity,
        "equity_base": equity_base,
        "unclosed_profit": unclosed_profit,
        "lt_liabilities": lt_liabilities,
        "st_liabilities": st_liabilities,
        "total_liabilities": total_liabilities,
        "revenue": revenue,
        "expenses": expenses,
        "cost_of_goods": cost_of_goods,
        "gross_profit": gross_profit,
        "operating_expenses": operating_expenses,
        "operating_profit": operating_profit,
        "other_expenses": other_expenses,
        "profit_before_tax": profit_before_tax,
        "tax": tax,
        "net_profit": net_profit,
        "net_fa": net_fa,
        "capex": capex,
        # Previous period
        "total_assets_prev": total_assets_p,
        "non_current_assets_prev": non_current_assets_p,
        "current_assets_prev": current_assets_p,
        "inventories_prev": inventories_p,
        "receivables_prev": receivables_p,
        "cash_prev": cash_p,
    }


def _ifrs_adjusted_aggregates(agg: dict) -> dict:
    """Return IFRS-adjusted version of NSBU balance aggregates.

    Applies: IAS 16 revaluation (+15%), IFRS 9 ECL (-5% receivables),
    IFRS 16 ROU asset (= lease liability in lt_liabilities),
    IAS 12 deferred tax liability (15% of revaluation).
    """
    ifrs = dict(agg)
    net_fa = agg["net_fa"]
    ias16_reval = round(net_fa * 0.15, 2)
    ecl = round(agg["receivables"] * 0.05, 2)
    deferred_tax = round(ias16_reval * 0.15, 2)
    # ROU asset estimate (950,000 default if lease data not explicit)
    rou_asset = 950_000

    ifrs["non_current_assets"] = agg["non_current_assets"] + ias16_reval + rou_asset
    ifrs["receivables"] = agg["receivables"] - ecl
    ifrs["current_assets"] = agg["current_assets"] - ecl
    ifrs["total_assets"] = agg["total_assets"] + ias16_reval - ecl + rou_asset
    ifrs["total_equity"] = agg["total_equity"] + ias16_reval - ecl - deferred_tax
    ifrs["lt_liabilities"] = agg["lt_liabilities"] + rou_asset + deferred_tax
    ifrs["total_liabilities"] = agg["total_liabilities"] + rou_asset + deferred_tax
    ifrs["net_profit"] = agg["net_profit"] - ecl
    ifrs["net_fa"] = net_fa + ias16_reval
    return ifrs


def _revenue_costs_from_income_expenses(income_expenses: list) -> dict:
    """Extract revenue/costs/opex/other_expenses/tax from income_expenses by account codes.

    Account code prefixes (NSBU chart of accounts):
      90xx = revenue (sales + other income)
      20xx = cost of goods sold
      94xx = operating expenses (admin, selling, other)
      95xx, 96xx = other expenses (financial, extraordinary)
      9720 = income tax
    """
    import re as _re
    revenue_cur = 0.0
    revenue_prev = 0.0
    costs_cur = 0.0
    costs_prev = 0.0
    opex_cur = 0.0
    opex_prev = 0.0
    other_expenses_cur = 0.0
    other_expenses_prev = 0.0
    tax_cur = 0.0
    tax_prev = 0.0
    for item in (income_expenses or []):
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", ""))
        m = _re.match(r"(\d{4})", name)
        code = m.group(1) if m else ""
        cur = float(item.get("current_year") or item.get("current_period") or item.get("current") or 0)
        prev = float(item.get("previous_year") or item.get("previous_period") or item.get("previous") or 0)
        if code == "9720":
            tax_cur += cur
            tax_prev += prev
        elif code.startswith("90"):
            revenue_cur += cur
            revenue_prev += prev
        elif code.startswith("20"):
            costs_cur += cur
            costs_prev += prev
        elif code.startswith("94"):
            opex_cur += cur
            opex_prev += prev
        elif code.startswith("95") or code.startswith("96"):
            other_expenses_cur += cur
            other_expenses_prev += prev
    return {
        "revenue_cur": revenue_cur, "revenue_prev": revenue_prev,
        "costs_cur": costs_cur, "costs_prev": costs_prev,
        "opex_cur": opex_cur, "opex_prev": opex_prev,
        "other_expenses_cur": other_expenses_cur, "other_expenses_prev": other_expenses_prev,
        "tax_cur": tax_cur, "tax_prev": tax_prev,
    }


def _build_ifrs_income_rows(accounts: dict, pnl: Optional[dict] = None,
                             income_expenses: Optional[list] = None) -> list:
    """Build IFRS Comprehensive Income (IAS 1) rows from cache data.

    Starts with NSBU P&L data (from income_expenses or pnl dict), then applies
    IFRS adjustments (IFRS 16, IFRS 9, IAS 16).
    """
    def _v(code: str, field: str = "current") -> float:
        acc = accounts.get(code)
        return acc[field] if acc else 0.0

    pnl = pnl or {}
    revenue = pnl.get("total_revenue_end", 0)
    expenses = pnl.get("total_expenses_end", 0)
    revenue_prev = pnl.get("total_revenue_start", pnl.get("total_revenue_begin", 0))
    expenses_prev = pnl.get("total_expenses_start", pnl.get("total_expenses_begin", 0))

    # If pnl totals are zero, extract from income_expenses by account codes
    if revenue == 0 and expenses == 0 and income_expenses:
        ie = _revenue_costs_from_income_expenses(income_expenses)
        revenue = ie["revenue_cur"]
        expenses = ie["costs_cur"] + ie["opex_cur"]
        revenue_prev = ie["revenue_prev"]
        expenses_prev = ie["costs_prev"] + ie["opex_prev"]

    gross_profit = revenue - expenses
    gross_profit_prev = revenue_prev - expenses_prev

    # IFRS 16 lease adjustments (from account 6970 if exists)
    lease_payment = _v("6970")
    if lease_payment > 0:
        discount_rate = 0.18
        lease_term = 5
        pv_factor = (1 - (1 + discount_rate) ** (-lease_term)) / discount_rate
        rou_asset = lease_payment * pv_factor
        rou_depreciation = round(rou_asset / lease_term, 2)
        lease_interest = round(rou_asset * discount_rate, 2)
    else:
        rou_depreciation = 0
        lease_interest = 0

    # IFRS 9 ECL (5% of receivables)
    receivables = _v("4010") or (_v("2010") + _v("2300"))
    ecl_impairment = round(receivables * 0.05, 2)

    # IAS 16 revaluation (15% premium on PPE)
    net_fa = _v("0100") - _v("0200")
    oci_revaluation = round(net_fa * 0.15, 2)

    profit_before_tax = gross_profit - rou_depreciation - lease_interest - ecl_impairment
    income_tax = round(profit_before_tax * 0.15, 2) if profit_before_tax > 0 else 0
    net_profit = profit_before_tax - income_tax
    total_comprehensive = net_profit + oci_revaluation

    profit_before_tax_prev = gross_profit_prev
    income_tax_prev = round(profit_before_tax_prev * 0.15, 2) if profit_before_tax_prev > 0 else 0
    net_profit_prev = profit_before_tax_prev - income_tax_prev

    rows = [
        {"label": "I. ОТЧЁТ О ПРИБЫЛЯХ И УБЫТКАХ", "current": None, "previous": None, "isHeader": True},
        {"label": "Выручка", "current": revenue, "previous": revenue_prev, "note": "IAS 18"},
        {"label": "Себестоимость", "current": -expenses if expenses else None, "previous": -expenses_prev if expenses_prev else None, "note": "IAS 2"},
        {"label": "Валовая прибыль", "current": gross_profit, "previous": gross_profit_prev, "isTotal": True},
        {"label": "Амортизация ПП-актива (IFRS 16)", "current": -rou_depreciation if rou_depreciation else None, "previous": None, "note": "IFRS 16"},
        {"label": "Процентные расходы по аренде (IFRS 16)", "current": -lease_interest if lease_interest else None, "previous": None, "note": "IFRS 16"},
        {"label": "Обесценение дебиторки (IFRS 9 ECL)", "current": -ecl_impairment if ecl_impairment else None, "previous": None, "note": "IFRS 9"},
        {"label": "Прибыль до налога", "current": profit_before_tax, "previous": profit_before_tax_prev, "isTotal": True},
        {"label": "Налог на прибыль", "current": -income_tax if income_tax else None, "previous": -income_tax_prev if income_tax_prev else None, "note": "IAS 12"},
        {"label": "ЧИСТАЯ ПРИБЫЛЬ", "current": net_profit, "previous": net_profit_prev, "isTotal": True},
        {"label": "", "current": None, "previous": None},
        {"label": "II. ПРОЧИЙ СОВОКУПНЫЙ ДОХОД (OCI)", "current": None, "previous": None, "isHeader": True},
        {"label": "Переоценка ОС (IAS 16)", "current": oci_revaluation if oci_revaluation else None, "previous": None, "note": "IAS 16"},
        {"label": "ИТОГО СОВОКУПНЫЙ ДОХОД", "current": total_comprehensive, "previous": net_profit_prev, "isTotal": True},
    ]
    return rows


def _build_ifrs_cashflow_rows(accounts: dict, pnl: Optional[dict] = None,
                               income_expenses: Optional[list] = None,
                               cashflow_data: Optional[list] = None) -> list:
    """Build IFRS Cash Flow (IAS 7) rows from cache data.

    Starts with NSBU cash flow data, then applies IFRS adjustments (IFRS 16).
    """
    def _v(code: str, field: str = "current") -> float:
        acc = accounts.get(code)
        return acc[field] if acc else 0.0

    pnl = pnl or {}
    revenue = pnl.get("total_revenue_end", 0)
    expenses = pnl.get("total_expenses_end", 0)

    # If pnl totals are zero, extract from income_expenses by account codes
    if revenue == 0 and expenses == 0 and income_expenses:
        ie = _revenue_costs_from_income_expenses(income_expenses)
        revenue = ie["revenue_cur"]
        expenses = ie["costs_cur"] + ie["opex_cur"]

    # Cash positions
    cash_end = _v("5010") + _v("5110") + _v("5210")
    cash_begin = _v("5010", "previous") + _v("5110", "previous") + _v("5210", "previous")
    net_change = cash_end - cash_begin

    # IFRS 16 lease split
    lease_payment = _v("6970")
    if lease_payment > 0:
        discount_rate = 0.18
        lease_term = 5
        pv_factor = (1 - (1 + discount_rate) ** (-lease_term)) / discount_rate
        rou_asset = lease_payment * pv_factor
        lease_interest = round(rou_asset * discount_rate, 2)
        lease_principal = round(lease_payment - lease_interest, 2)
    else:
        lease_interest = 0
        lease_principal = 0

    # Estimate operating cash flows from P&L + working capital changes
    receivables_change = (_v("2010") + _v("2300")) - (_v("2010", "previous") + _v("2300", "previous"))
    payables_change = _v("6010") - _v("6010", "previous")
    tax_paid = _v("6310") + _v("6610")

    operating_cash = revenue - expenses - receivables_change + payables_change - tax_paid - lease_interest

    # Investing
    capex_change = _v("0100") - _v("0100", "previous")

    # Financing
    lt_loans_change = _v("7010") - _v("7010", "previous")
    st_loans_change = _v("6820") - _v("6820", "previous")

    operating_total = round(operating_cash, 2) if revenue else None
    investing_total = round(-capex_change, 2) if capex_change else None
    financing_total = round(lt_loans_change + st_loans_change - lease_principal, 2) if (lt_loans_change or st_loans_change or lease_principal) else None

    rows = [
        {"label": "I. ОПЕРАЦИОННАЯ ДЕЯТЕЛЬНОСТЬ", "current": None, "previous": None, "isHeader": True},
        {"label": "Поступления от покупателей", "current": revenue if revenue else None, "previous": None, "note": "IAS 7"},
        {"label": "Оплата поставщикам и персоналу", "current": -expenses if expenses else None, "previous": None, "note": "IAS 7"},
        {"label": "Процентные расходы по аренде (IFRS 16)", "current": -lease_interest if lease_interest else None, "previous": None, "note": "IFRS 16"},
        {"label": "Налоги уплаченные", "current": -tax_paid if tax_paid else None, "previous": None, "note": "IAS 12"},
        {"label": "Итого по операционной деятельности", "current": operating_total, "previous": None, "isTotal": True},
        {"label": "", "current": None, "previous": None},
        {"label": "II. ИНВЕСТИЦИОННАЯ ДЕЯТЕЛЬНОСТЬ", "current": None, "previous": None, "isHeader": True},
        {"label": "Приобретение основных средств", "current": -capex_change if capex_change > 0 else None, "previous": None, "note": "IAS 16"},
        {"label": "Выбытие основных средств", "current": -capex_change if capex_change < 0 else None, "previous": None, "note": "IAS 16"},
        {"label": "Итого по инвестиционной деятельности", "current": investing_total, "previous": None, "isTotal": True},
        {"label": "", "current": None, "previous": None},
        {"label": "III. ФИНАНСОВАЯ ДЕЯТЕЛЬНОСТЬ", "current": None, "previous": None, "isHeader": True},
        {"label": "Поступление/погашение кредитов", "current": round(lt_loans_change + st_loans_change, 2) if (lt_loans_change or st_loans_change) else None, "previous": None, "note": "IFRS 9"},
        {"label": "Погашение обязательства по аренде (IFRS 16)", "current": -lease_principal if lease_principal else None, "previous": None, "note": "IFRS 16"},
        {"label": "Итого по финансовой деятельности", "current": financing_total, "previous": None, "isTotal": True},
        {"label": "", "current": None, "previous": None},
        {"label": "Чистое изменение денежных средств", "current": round(net_change, 2), "previous": None, "isTotal": True},
        {"label": "Денежные средства на начало периода", "current": round(cash_begin, 2), "previous": None},
        {"label": "Денежные средства на конец периода", "current": round(cash_end, 2), "previous": None, "isTotal": True},
    ]
    return rows


def _build_ifrs_equity_rows(accounts: dict, pnl: Optional[dict] = None,
                             income_expenses: Optional[list] = None) -> list:
    """Build IFRS Changes in Equity (IAS 1) rows from cache data."""
    def _v(code: str, field: str = "current") -> float:
        acc = accounts.get(code)
        return acc[field] if acc else 0.0

    pnl = pnl or {}
    charter = _v("8300")
    charter_prev = _v("8300", "previous")
    reserve = _v("8500")
    reserve_prev = _v("8500", "previous")
    retained = _v("8700")
    retained_prev = _v("8700", "previous")

    revenue = pnl.get("total_revenue_end", 0)
    expenses = pnl.get("total_expenses_end", 0)

    # Fallback: if pnl totals are zero, extract from income_expenses by account codes
    if revenue == 0 and expenses == 0 and income_expenses:
        ie = _revenue_costs_from_income_expenses(income_expenses)
        revenue = ie["revenue_cur"]
        expenses = ie["costs_cur"] + ie["opex_cur"]

    net_profit = revenue - expenses

    # OCI from IAS 16 revaluation
    net_fa = _v("0100") - _v("0200")
    oci_revaluation = round(net_fa * 0.15, 2)

    # IFRS 9 ECL adjustment on trade receivables (5% of 2010)
    ecl_adjustment = round(_v("2010") * 0.05, 2)

    # IFRS-adjusted net profit
    net_profit_ifrs = net_profit - ecl_adjustment

    total_prev = charter_prev + reserve_prev + retained_prev
    total_cur = charter + reserve + retained + net_profit_ifrs + oci_revaluation

    rows = [
        {"label": "Начальное сальдо", "current": round(total_prev, 2), "previous": None, "isTotal": True},
        {"label": "  Уставный капитал", "current": charter_prev, "previous": None},
        {"label": "  Резервный капитал", "current": reserve_prev, "previous": None},
        {"label": "  Нераспределённая прибыль", "current": retained_prev, "previous": None},
        {"label": "", "current": None, "previous": None},
        {"label": "Изменения за период", "current": None, "previous": None, "isHeader": True},
        {"label": "Чистая прибыль за период (МСФО)", "current": round(net_profit_ifrs, 2) if net_profit_ifrs else None, "previous": None},
        {"label": "  в т.ч. ECL корректировка (IFRS 9)", "current": round(-ecl_adjustment, 2) if ecl_adjustment else None, "previous": None, "note": "IFRS 9"},
        {"label": "Переоценка ОС (OCI)", "current": oci_revaluation if oci_revaluation else None, "previous": None, "note": "IAS 16"},
        {"label": "Изменение уставного капитала", "current": round(charter - charter_prev, 2) if charter != charter_prev else None, "previous": None},
        {"label": "Изменение резервного капитала", "current": round(reserve - reserve_prev, 2) if reserve != reserve_prev else None, "previous": None},
        {"label": "", "current": None, "previous": None},
        {"label": "Конечное сальдо", "current": round(total_cur, 2), "previous": None, "isTotal": True},
    ]
    return rows


def _safe_div(a: float, b: float) -> float:
    return round(a / b, 4) if b != 0 else 0.0


def _kpi_status(value: float, norm_threshold: float, lower_is_better: bool = False) -> str:
    """Determine KPI status: ok / warn / bad."""
    if lower_is_better:
        if value <= norm_threshold:
            return "ok"
        elif value <= norm_threshold * 1.5:
            return "warn"
        return "bad"
    else:
        if value >= norm_threshold:
            return "ok"
        elif value >= norm_threshold * 0.5:
            return "warn"
        return "bad"


def _calc_kpis(agg: dict) -> list:
    """Calculate KPI groups from balance aggregates.

    Returns list of groups with 'title', 'icon', 'metrics' array.
    Each metric has: key, label, value, formula, norm, status, standard.
    """
    ca = agg["current_assets"]
    cl = agg["st_liabilities"]
    inv = agg["inventories"]
    cash = agg["cash"]
    ta = agg["total_assets"]
    te = agg["total_equity"]
    tl = agg["total_liabilities"]
    np_ = agg["net_profit"]
    rev = agg["revenue"]
    gp = agg.get("gross_profit", rev - agg.get("cost_of_goods", agg.get("expenses", 0)))
    recv = agg["receivables"]

    current_ratio = _safe_div(ca, cl)
    quick_ratio = _safe_div(ca - inv, cl)
    cash_ratio = _safe_div(cash, cl)
    roa = _safe_div(np_, ta)
    roe = _safe_div(np_, te)
    net_margin = _safe_div(np_, rev)
    gross_margin = _safe_div(gp, rev)
    debt_to_equity = _safe_div(tl, te)
    debt_ratio = _safe_div(tl, ta)
    equity_ratio = _safe_div(te, ta)
    asset_turnover = _safe_div(rev, ta)
    receivables_turnover = _safe_div(rev, recv) if recv else 0.0

    return [
        {
            "title": "Ликвидность",
            "icon": "💧",
            "metrics": [
                {"key": "current_ratio", "label": "Текущая ликвидность", "value": round(current_ratio, 2), "formula": "Оборотные активы / Краткосрочные обяз.", "norm": "> 2.0", "status": _kpi_status(current_ratio, 2.0), "standard": "both"},
                {"key": "quick_ratio", "label": "Быстрая ликвидность", "value": round(quick_ratio, 2), "formula": "(Оборотные - Запасы) / Краткосрочные обяз.", "norm": "> 1.0", "status": _kpi_status(quick_ratio, 1.0), "standard": "both"},
                {"key": "cash_ratio", "label": "Абсолютная ликвидность", "value": round(cash_ratio, 2), "formula": "Денежные средства / Краткосрочные обяз.", "norm": "> 0.2", "status": _kpi_status(cash_ratio, 0.2), "standard": "both"},
            ],
        },
        {
            "title": "Рентабельность",
            "icon": "📈",
            "metrics": [
                {"key": "roa", "label": "Рентабельность активов (ROA)", "value": round(roa, 4), "formula": "Чистая прибыль / Активы", "norm": "> 5%", "status": _kpi_status(roa, 0.05), "standard": "both"},
                {"key": "roe", "label": "Рентабельность капитала (ROE)", "value": round(roe, 4), "formula": "Чистая прибыль / Капитал", "norm": "> 15%", "status": _kpi_status(roe, 0.15), "standard": "both"},
                {"key": "net_margin", "label": "Чистая маржа", "value": round(net_margin, 4), "formula": "Чистая прибыль / Выручка", "norm": "> 10%", "status": _kpi_status(net_margin, 0.10), "standard": "both"},
                {"key": "gross_margin", "label": "Валовая маржа", "value": round(gross_margin, 4), "formula": "(Выручка - Себестоимость) / Выручка", "norm": "> 20%", "status": _kpi_status(gross_margin, 0.20), "standard": "both"},
            ],
        },
        {
            "title": "Левередж",
            "icon": "⚖️",
            "metrics": [
                {"key": "debt_to_equity", "label": "Долг / Капитал", "value": round(debt_to_equity, 2), "formula": "Обязательства / Капитал", "norm": "< 1.5", "status": _kpi_status(debt_to_equity, 1.5, lower_is_better=True), "standard": "both"},
                {"key": "debt_ratio", "label": "Коэффициент долга", "value": round(debt_ratio, 2), "formula": "Обязательства / Активы", "norm": "< 0.5", "status": _kpi_status(debt_ratio, 0.5, lower_is_better=True), "standard": "both"},
                {"key": "equity_ratio", "label": "Коэффициент автономии", "value": round(equity_ratio, 2), "formula": "Капитал / Активы", "norm": "> 0.5", "status": _kpi_status(equity_ratio, 0.5), "standard": "both"},
            ],
        },
        {
            "title": "Деловая активность",
            "icon": "🔄",
            "metrics": [
                {"key": "asset_turnover", "label": "Оборачиваемость активов", "value": round(asset_turnover, 2), "formula": "Выручка / Активы", "norm": "> 1.0", "status": _kpi_status(asset_turnover, 1.0), "standard": "both"},
                {"key": "receivables_turnover", "label": "Оборачиваемость дебиторки", "value": round(receivables_turnover, 2), "formula": "Выручка / Дебиторская задолж.", "norm": "> 4.0", "status": _kpi_status(receivables_turnover, 4.0), "standard": "both"},
            ],
        },
    ]


# ---------------------------------------------------------------------------
# GET /analytics/debug-cache  (temporary diagnostic endpoint)
# ---------------------------------------------------------------------------

@router.get("/analytics/debug-cache")
async def debug_cache(current_user: User = Depends(get_current_user)):
    cache = _user_cache(current_user.id)
    pnl = cache.get("pnl", {})
    ie = cache.get("income_expenses", [])
    accounts = cache.get("accounts", {})

    # Try to compute revenue the same way _get_balance_aggregates does
    revenue_from_pnl = pnl.get("total_revenue_end", 0.0)
    revenue_from_pnl2 = pnl.get("revenue", pnl.get("total_revenue", 0.0))

    # Compute from income_expenses
    revenue_from_ie = 0
    costs_from_ie = 0
    for item in (ie if isinstance(ie, list) else []):
        if isinstance(item, dict):
            code = str(item.get("name", ""))[:4]
            val = item.get("current_year", 0) or 0
            if code.startswith("90"):
                revenue_from_ie += val
            elif code.startswith("20") or code.startswith("94"):
                costs_from_ie += val

    # Compute from accounts (OSV)
    revenue_from_accounts = 0
    for code, acc in accounts.items():
        if code.startswith("90"):
            revenue_from_accounts += acc.get("credit_end", 0) or acc.get("current", 0) or 0

    return {
        "cache_keys": list(cache.keys()),
        "pnl_keys": list(pnl.keys()) if isinstance(pnl, dict) else str(type(pnl)),
        "pnl_raw": pnl,
        "income_expenses_count": len(ie) if isinstance(ie, list) else 0,
        "income_expenses_sample": ie[:3] if isinstance(ie, list) else [],
        "accounts_count": len(accounts),
        "accounts_90xx": {k: v for k, v in accounts.items() if k.startswith("90")},
        "revenue_from_pnl": revenue_from_pnl,
        "revenue_from_pnl2": revenue_from_pnl2,
        "revenue_from_ie": revenue_from_ie,
        "costs_from_ie": costs_from_ie,
        "revenue_from_accounts": revenue_from_accounts,
    }


# ---------------------------------------------------------------------------
# GET /analytics/kpi
# ---------------------------------------------------------------------------

@router.get("/analytics/kpi")
async def get_kpi(
    standard: Optional[str] = "nsbu",
    current_user: User = Depends(get_current_user),
):
    """KPI groups by accounting standard."""
    agg = _get_balance_aggregates(user_id=current_user.id)
    if agg is None:
        return JSONResponse({"groups": []})
    return JSONResponse({"groups": _calc_kpis(agg)})


@router.get("/analytics/ai-analysis")
async def get_ai_analysis(
    current_user: User = Depends(get_current_user),
):
    """AI-driven financial analysis conclusions based on KPI data."""
    agg = _get_balance_aggregates(user_id=current_user.id)
    if agg is None:
        return JSONResponse({"conclusions": [], "overall": "", "score": 0})
    conclusions, overall = _build_ai_analysis(agg)
    ok_count = sum(1 for _, s, _ in conclusions if s == "ok")
    score = round(ok_count / max(len(conclusions), 1) * 100, 1)
    return JSONResponse({
        "conclusions": [{"category": c, "status": s, "text": t} for c, s, t in conclusions],
        "overall": overall,
        "score": score,
    })


def _kpi_flat(agg: dict) -> dict:
    """Return a flat dict of {key: value} from KPI groups for stress-test use."""
    result = {}
    for group in _calc_kpis(agg):
        for m in group["metrics"]:
            result[m["key"]] = m["value"]
    return result


def _kpi_labels() -> dict:
    """Return a flat dict of {key: label} for human-readable metric names."""
    # Use a dummy agg to get the structure
    dummy = {k: 1.0 for k in [
        "current_assets", "st_liabilities", "inventories", "cash",
        "total_assets", "total_equity", "total_liabilities",
        "net_profit", "revenue", "expenses", "receivables",
    ]}
    result = {}
    for group in _calc_kpis(dummy):
        for m in group["metrics"]:
            result[m["key"]] = m["label"]
    return result


# ---------------------------------------------------------------------------
# POST /analytics/dcf
# ---------------------------------------------------------------------------

class DcfInput(BaseModel):
    revenue: float = 0
    growth_rate: float = 0
    wacc: float = 0
    terminal_growth: float = 0
    years: int = 5


@router.post("/analytics/dcf")
async def calculate_dcf(
    data: DcfInput,
    current_user: User = Depends(get_current_user),
):
    """DCF valuation using cached P&L data (revenue, net profit)."""
    agg = _get_balance_aggregates(user_id=current_user.id)

    # Use cached data if available, otherwise fall back to input
    revenue = data.revenue if data.revenue else (agg["revenue"] if agg else 0)
    net_profit = agg["net_profit"] if agg else 0
    total_assets = agg["total_assets"] if agg else 0
    total_liabilities = agg["total_liabilities"] if agg else 0
    total_equity = agg["total_equity"] if agg else 0
    cash = agg["cash"] if agg else 0

    wacc = data.wacc if data.wacc else 0.15
    growth_rate = data.growth_rate if data.growth_rate else 0.05
    terminal_growth = data.terminal_growth if data.terminal_growth else 0.03
    years = data.years if data.years else 5

    if revenue == 0 and net_profit == 0:
        return JSONResponse({
            "wacc": wacc,
            "enterprise_value": 0,
            "equity_value": 0,
            "intrinsic_value_per_share": 0,
            "pv_fcff": 0,
            "terminal_value": 0,
            "fcff_year1": 0,
            "fcff_year2": 0,
            "fcff_year3": 0,
        })

    # Estimate FCFF from net profit (simplified: FCFF ≈ Net Profit * 0.8 as proxy)
    base_fcff = net_profit * 0.8 if net_profit > 0 else revenue * 0.10

    # Project FCFF for each year
    fcff_projections = []
    for yr in range(1, years + 1):
        fcff_projections.append(base_fcff * ((1 + growth_rate) ** yr))

    # PV of projected FCFFs
    pv_fcff = sum(fcff / ((1 + wacc) ** yr) for yr, fcff in enumerate(fcff_projections, 1))

    # Terminal value (Gordon Growth Model)
    last_fcff = fcff_projections[-1] if fcff_projections else base_fcff
    if wacc > terminal_growth:
        terminal_value = last_fcff * (1 + terminal_growth) / (wacc - terminal_growth)
    else:
        terminal_value = last_fcff * 20  # fallback multiplier

    pv_terminal = terminal_value / ((1 + wacc) ** years)

    enterprise_value = pv_fcff + pv_terminal
    # Equity = EV - Net Debt, where Net Debt = Total Liabilities - Cash
    net_debt = total_liabilities - cash
    equity_value = enterprise_value - net_debt

    return JSONResponse({
        "wacc": round(wacc, 4),
        "enterprise_value": round(enterprise_value, 0),
        "equity_value": round(equity_value, 0),
        "intrinsic_value_per_share": round(equity_value, 0),
        "pv_fcff": round(pv_fcff, 0),
        "terminal_value": round(pv_terminal, 0),
        "net_debt": round(net_debt, 0),
        "fcff_year1": round(fcff_projections[0], 0) if len(fcff_projections) > 0 else 0,
        "fcff_year2": round(fcff_projections[1], 0) if len(fcff_projections) > 1 else 0,
        "fcff_year3": round(fcff_projections[2], 0) if len(fcff_projections) > 2 else 0,
    })


# ---------------------------------------------------------------------------
# GET /analytics/multiples
# ---------------------------------------------------------------------------

@router.get("/analytics/multiples")
async def get_multiples():
    """Market multiples stub."""
    return JSONResponse({})


# ---------------------------------------------------------------------------
# POST /analytics/stress-test
# ---------------------------------------------------------------------------

class StressTestInput(BaseModel):
    scenario: str = "crisis_2008"
    severity: str = "moderate"
    standard: str = "both"


_DEFAULT_SCENARIOS = {
    "crisis_2008": {"name": "Кризис 2008", "revenue_shock": -0.30, "cost_shock": 0.10, "interest_shock": 0, "fx_shock": 0},
    "covid_2020": {"name": "COVID-2020", "revenue_shock": -0.20, "cost_shock": 0.05, "interest_shock": 0, "fx_shock": 0},
    "rate_hike": {"name": "Рост ставок", "revenue_shock": 0, "cost_shock": 0, "interest_shock": 0.50, "fx_shock": 0},
    "currency_shock": {"name": "Валютный шок", "revenue_shock": 0, "cost_shock": 0, "interest_shock": 0, "fx_shock": 2.0},
    "commodity_drop": {"name": "Рост цен на сырьё", "revenue_shock": 0, "cost_shock": 0.25, "interest_shock": 0, "fx_shock": 0},
}

_SEVERITY_MULTIPLIERS = {"mild": 0.5, "moderate": 1.0, "severe": 1.5, "extreme": 2.0}


def _stress_status(delta_pct: float) -> str:
    """Determine stress-test status from delta percentage."""
    if abs(delta_pct) < 10:
        return "ok"
    elif abs(delta_pct) < 25:
        return "warn"
    return "bad"


@router.post("/analytics/stress-test")
async def run_stress_test(
    data: StressTestInput = StressTestInput(),
    current_user: User = Depends(get_current_user),
):
    """Stress test: apply single scenario shocks to baseline KPIs.

    Accepts: scenario (string id), severity (mild/moderate/severe/extreme), standard (nsbu/ifrs/both).
    Returns: flat results array with baseline/stressed values per standard.
    """
    agg = _get_balance_aggregates(user_id=current_user.id)
    if agg is None:
        return JSONResponse({"results": [], "ai_summary": []})

    sc = _DEFAULT_SCENARIOS.get(data.scenario, _DEFAULT_SCENARIOS["crisis_2008"])
    severity_mult = _SEVERITY_MULTIPLIERS.get(data.severity, 1.0)

    rev_shock = sc["revenue_shock"] * severity_mult
    cost_shock = sc["cost_shock"] * severity_mult
    interest_shock = sc["interest_shock"] * severity_mult
    fx_shock = sc["fx_shock"] * severity_mult

    stressed_rev = agg["revenue"] * (1 + rev_shock)
    stressed_exp = agg["expenses"] * (1 + cost_shock)
    interest_extra = agg["lt_liabilities"] * interest_shock * 0.10
    fx_extra = agg["total_assets"] * 0.05 * fx_shock
    stressed_np = stressed_rev - stressed_exp - interest_extra - fx_extra

    # NSBU stressed
    stressed_agg = dict(agg)
    stressed_agg["revenue"] = stressed_rev
    stressed_agg["expenses"] = stressed_exp
    stressed_agg["net_profit"] = stressed_np

    # IFRS baseline & stressed (with IAS 16, IFRS 9, IFRS 16 adjustments)
    agg_ifrs = _ifrs_adjusted_aggregates(agg)
    stressed_rev_ifrs = agg_ifrs["revenue"] * (1 + rev_shock)
    stressed_exp_ifrs = agg_ifrs["expenses"] * (1 + cost_shock)
    interest_extra_ifrs = agg_ifrs["lt_liabilities"] * interest_shock * 0.10
    fx_extra_ifrs = agg_ifrs["total_assets"] * 0.05 * fx_shock
    stressed_np_ifrs = stressed_rev_ifrs - stressed_exp_ifrs - interest_extra_ifrs - fx_extra_ifrs

    stressed_agg_ifrs = dict(agg_ifrs)
    stressed_agg_ifrs["revenue"] = stressed_rev_ifrs
    stressed_agg_ifrs["expenses"] = stressed_exp_ifrs
    stressed_agg_ifrs["net_profit"] = stressed_np_ifrs

    baseline_flat = _kpi_flat(agg)
    stressed_flat = _kpi_flat(stressed_agg)
    baseline_flat_ifrs = _kpi_flat(agg_ifrs)
    stressed_flat_ifrs = _kpi_flat(stressed_agg_ifrs)
    labels = _kpi_labels()

    results = []
    for key, base_val in baseline_flat.items():
        stressed_val = stressed_flat.get(key, 0)
        base_val_ifrs = baseline_flat_ifrs.get(key, 0)
        stressed_val_ifrs = stressed_flat_ifrs.get(key, 0)
        delta_pct_nsbu = round((stressed_val - base_val) / base_val * 100, 1) if base_val != 0 else 0.0
        delta_pct_ifrs = round((stressed_val_ifrs - base_val_ifrs) / base_val_ifrs * 100, 1) if base_val_ifrs != 0 else 0.0
        results.append({
            "metric": labels.get(key, key),
            "baseline_nsbu": base_val,
            "baseline_ifrs": base_val_ifrs,
            "stressed_nsbu": stressed_val,
            "stressed_ifrs": stressed_val_ifrs,
            "delta_pct_nsbu": delta_pct_nsbu,
            "delta_pct_ifrs": delta_pct_ifrs,
            "status_nsbu": _stress_status(delta_pct_nsbu),
            "status_ifrs": _stress_status(delta_pct_ifrs),
        })

    ai_summary = [
        "Базовые KPI рассчитаны из загруженной ОСВ.",
        f"Чистая прибыль базового сценария (НСБУ): {agg['net_profit']:,.0f}",
        f"Чистая прибыль базового сценария (МСФО): {agg_ifrs['net_profit']:,.0f}",
        f"Сценарий: {sc['name']} (severity: {data.severity}).",
    ]
    if agg["revenue"] == 0 and agg["expenses"] == 0:
        ai_summary.append("⚠️ Данные P&L (выручка, расходы) отсутствуют — рентабельность = 0.")

    return JSONResponse({"results": results, "ai_summary": ai_summary})


# ---------------------------------------------------------------------------
# GET /analytics/visualizations
# ---------------------------------------------------------------------------

@router.get("/analytics/visualizations")
async def get_visualizations(
    current_user: User = Depends(get_current_user),
):
    """Chart/visualization data from real balance data.

    Returns arrays ready for Recharts consumption:
    - waterfall: WaterfallItem[] with cumulative values
    - tornado: TornadoItem[] with impact percentages
    - bubble: BubbleItem[] with x/y/size
    - heatmap: { data: (string|number)[][], months: string[] }
    """
    agg = _get_balance_aggregates(user_id=current_user.id)
    if agg is None:
        return JSONResponse({})

    # --- Waterfall: balance movement (UZS) ---
    start_balance = agg["total_assets_prev"]
    revenue = agg["revenue"]
    costs = -agg["expenses"]
    inv_delta = -(agg["non_current_assets"] - agg.get("non_current_assets_prev", 0))
    end_balance = agg["total_assets"]

    cum = start_balance
    waterfall = [
        {"name": "Начальный баланс", "value": start_balance, "cumulative": cum, "type": "start"},
    ]
    cum += revenue
    waterfall.append({"name": "Выручка", "value": revenue, "cumulative": cum, "type": "positive"})
    cum += costs
    waterfall.append({"name": "Себестоимость", "value": costs, "cumulative": cum, "type": "negative"})
    # Operating expenses = difference not explained by other items
    op_exp = end_balance - cum - inv_delta
    if op_exp < 0:
        cum += op_exp
        waterfall.append({"name": "Операционные расходы", "value": op_exp, "cumulative": cum, "type": "negative"})
    if abs(inv_delta) > 0:
        cum += inv_delta
        waterfall.append({"name": "Инвестиции", "value": inv_delta, "cumulative": cum, "type": "negative" if inv_delta < 0 else "positive"})
    waterfall.append({"name": "Итоговый баланс", "value": end_balance, "cumulative": end_balance, "type": "total"})

    # --- Tornado: sensitivity analysis (impact on net profit ±10%) ---
    base_np = agg["net_profit"]
    rev_impact = agg["revenue"] * 0.10
    cost_impact = agg["expenses"] * 0.10
    interest_impact = agg["lt_liabilities"] * 0.01
    fx_impact = agg["total_assets"] * 0.005
    tornado = []
    if rev_impact:
        tornado.append({"factor": "Выручка ±10%", "impact": round(rev_impact / max(abs(base_np), 1) * 100, 1)})
    if cost_impact:
        tornado.append({"factor": "Себестоимость ±10%", "impact": round(-cost_impact / max(abs(base_np), 1) * 100, 1)})
    if interest_impact:
        tornado.append({"factor": "Процентные расходы ±10%", "impact": round(-interest_impact / max(abs(base_np), 1) * 100, 1)})
    if fx_impact:
        tornado.append({"factor": "Курсовые разницы", "impact": round(-fx_impact / max(abs(base_np), 1) * 100, 1)})
    # Add operating expense sensitivity
    if agg["expenses"]:
        overhead = agg["expenses"] * 0.05  # 5% overhead change
        tornado.append({"factor": "Накладные расходы ±5%", "impact": round(-overhead / max(abs(base_np), 1) * 100, 1)})

    # --- Bubble: revenue vs profitability vs total assets ---
    ta = agg["total_assets"]
    profitability = round(base_np / max(agg["revenue"], 1) * 100, 1) if agg["revenue"] else 0
    bubble = [
        {"name": "Основные средства", "x": round(agg["net_fa"]), "y": 5, "size": round(agg["net_fa"])},
        {"name": "Запасы", "x": round(agg["inventories"]), "y": 12, "size": round(agg["inventories"])},
        {"name": "Дебиторская задолженность", "x": round(agg["receivables"]), "y": 8, "size": round(agg["receivables"])},
        {"name": "Денежные средства", "x": round(agg["cash"]), "y": 2, "size": round(agg["cash"])},
        {"name": "Капитальные вложения", "x": round(agg["capex"]), "y": 15, "size": round(agg["capex"])},
    ]
    # Filter out zero-value bubbles
    bubble = [b for b in bubble if b["size"] > 0]

    # --- Heatmap: correlation matrix of financial metrics ---

    metrics = {
        "Выручка": agg["revenue"],
        "Себестоимость": agg["expenses"],
        "Чист. прибыль": base_np,
        "Активы": ta,
        "Капитал": agg["total_equity"],
        "Обязательства": agg["total_liabilities"],
    }
    metric_names = list(metrics.keys())
    metric_vals = list(metrics.values())
    n = len(metric_names)

    # Build correlation-like heatmap based on normalized ratios
    heatmap_data = []
    for i in range(n):
        row = [metric_names[i]]
        for j in range(n):
            if i == j:
                row.append(100.0)
            else:
                vi, vj = metric_vals[i], metric_vals[j]
                if vi != 0 and vj != 0:
                    ratio = min(abs(vi), abs(vj)) / max(abs(vi), abs(vj)) * 100
                    sign = 1 if (vi > 0) == (vj > 0) else -1
                    row.append(round(ratio * sign, 1))
                else:
                    row.append(0.0)
        heatmap_data.append(row)

    return JSONResponse({
        "waterfall": waterfall,
        "tornado": tornado,
        "bubble": bubble,
        "heatmap": heatmap_data,
        "heatmap_months": metric_names,
    })


# ---------------------------------------------------------------------------
# POST /decisions/impact
# ---------------------------------------------------------------------------

class ImpactInput(BaseModel):
    decision_type: str = ""
    params: Dict[str, Any] = {}


@router.post("/decisions/impact")
async def calculate_impact(data: ImpactInput = ImpactInput()):
    """Decision impact calculator stub."""
    return JSONResponse({"rows": []})


# ---------------------------------------------------------------------------
# POST /analytics/export/full-report — E2-08: Full NSBU+IFRS Excel export
# ---------------------------------------------------------------------------

def _build_ai_analysis(agg: dict) -> tuple:
    """Generate AI-driven financial analysis conclusions based on KPI data.

    Returns (conclusions, overall) where conclusions is a list of
    (category, status, text) tuples and overall is a summary string.
    """
    conclusions = []

    # Liquidity
    ca = agg.get("current_assets", 0)
    cl = agg.get("st_liabilities", 0)
    cr = _safe_div(ca, cl) if cl else 0
    if cr > 2:
        conclusions.append(("Ликвидность", "ok",
            f"Текущая ликвидность {cr:.2f}x — в норме. Компания способна покрывать краткосрочные обязательства."))
    elif cr > 1:
        conclusions.append(("Ликвидность", "warn",
            f"Текущая ликвидность {cr:.2f}x — на нижней границе нормы. Рекомендуется увеличить оборотные активы."))
    else:
        conclusions.append(("Ликвидность", "bad",
            f"Текущая ликвидность {cr:.2f}x — КРИТИЧЕСКИ НИЗКАЯ. Высокий риск неплатёжеспособности."))

    # Cash ratio
    cash = agg.get("cash", 0)
    cash_r = _safe_div(cash, cl) if cl else 0
    if cash_r > 0.2:
        conclusions.append(("Абсолютная ликвидность", "ok",
            f"Абсолютная ликвидность {cash_r:.2f}x — достаточный запас денежных средств."))
    elif cash_r > 0.1:
        conclusions.append(("Абсолютная ликвидность", "warn",
            f"Абсолютная ликвидность {cash_r:.2f}x — минимальный запас. Рекомендуется увеличить резерв."))
    else:
        conclusions.append(("Абсолютная ликвидность", "bad",
            f"Абсолютная ликвидность {cash_r:.2f}x — КРИТИЧНО. Недостаточно денежных средств для покрытия обязательств."))

    # Profitability — ROA
    ta = agg.get("total_assets", 0)
    te = agg.get("total_equity", 0)
    np_ = agg.get("net_profit", 0)
    rev = agg.get("revenue", 0)
    exp = agg.get("expenses", 0)

    roa = _safe_div(np_, ta) if ta else 0
    if roa > 0.05:
        conclusions.append(("Рентабельность активов (ROA)", "ok",
            f"ROA = {roa*100:.1f}% — активы генерируют достаточную прибыль."))
    elif roa > 0.02:
        conclusions.append(("Рентабельность активов (ROA)", "warn",
            f"ROA = {roa*100:.1f}% — рентабельность ниже нормы (5%). Рекомендуется повысить эффективность."))
    else:
        conclusions.append(("Рентабельность активов (ROA)", "bad",
            f"ROA = {roa*100:.1f}% — КРИТИЧЕСКИ НИЗКАЯ рентабельность активов."))

    # Profitability — ROE
    roe = _safe_div(np_, te) if te else 0
    if roe > 0.15:
        conclusions.append(("Рентабельность капитала (ROE)", "ok",
            f"ROE = {roe*100:.1f}% — высокая отдача на капитал."))
    elif roe > 0.075:
        conclusions.append(("Рентабельность капитала (ROE)", "warn",
            f"ROE = {roe*100:.1f}% — ниже нормы (15%). Капитал используется недостаточно эффективно."))
    else:
        conclusions.append(("Рентабельность капитала (ROE)", "bad",
            f"ROE = {roe*100:.1f}% — НИЗКАЯ отдача на капитал. Рассмотреть реструктуризацию."))

    # Net margin
    margin = _safe_div(np_, rev) if rev else 0
    if margin > 0.10:
        conclusions.append(("Чистая маржа", "ok",
            f"Чистая маржа {margin*100:.1f}% — высокая рентабельность продаж."))
    elif margin > 0.05:
        conclusions.append(("Чистая маржа", "warn",
            f"Чистая маржа {margin*100:.1f}% — ниже нормы (10%). Рекомендуется оптимизировать затраты."))
    else:
        conclusions.append(("Чистая маржа", "bad",
            f"Чистая маржа {margin*100:.1f}% — КРИТИЧНО. Бизнес убыточен или на грани."))

    # Gross margin
    gross_margin = _safe_div(rev - exp, rev) if rev else 0
    if gross_margin > 0.20:
        conclusions.append(("Валовая маржа", "ok",
            f"Валовая маржа {gross_margin*100:.1f}% — хорошая наценка на продукцию."))
    elif gross_margin > 0.10:
        conclusions.append(("Валовая маржа", "warn",
            f"Валовая маржа {gross_margin*100:.1f}% — ниже нормы (20%). Пересмотреть ценообразование."))
    else:
        conclusions.append(("Валовая маржа", "bad",
            f"Валовая маржа {gross_margin*100:.1f}% — КРИТИЧНО. Себестоимость слишком высока."))

    # Debt
    tl = agg.get("total_liabilities", 0)
    de = _safe_div(tl, te) if te else 0
    if de < 1.5:
        conclusions.append(("Долговая нагрузка", "ok",
            f"Долг/Капитал = {de:.2f}x — финансовая устойчивость в норме."))
    elif de < 2.25:
        conclusions.append(("Долговая нагрузка", "warn",
            f"Долг/Капитал = {de:.2f}x — повышенная долговая нагрузка. Рекомендуется снизить заимствования."))
    else:
        conclusions.append(("Долговая нагрузка", "bad",
            f"Долг/Капитал = {de:.2f}x — ВЫСОКИЙ РИСК. Компания чрезмерно закредитована."))

    # Autonomy
    equity_ratio = _safe_div(te, ta) if ta else 0
    if equity_ratio > 0.5:
        conclusions.append(("Автономия", "ok",
            f"Коэффициент автономии {equity_ratio:.2f} — компания финансово независима."))
    elif equity_ratio > 0.25:
        conclusions.append(("Автономия", "warn",
            f"Коэффициент автономии {equity_ratio:.2f} — зависимость от заёмных средств. Рекомендуется увеличить капитал."))
    else:
        conclusions.append(("Автономия", "bad",
            f"Коэффициент автономии {equity_ratio:.2f} — КРИТИЧНО. Высокая зависимость от кредиторов."))

    # Asset turnover
    at = _safe_div(rev, ta) if ta else 0
    if at > 1.0:
        conclusions.append(("Деловая активность", "ok",
            f"Оборачиваемость активов {at:.2f}x — активы генерируют выручку эффективно."))
    elif at > 0.5:
        conclusions.append(("Деловая активность", "warn",
            f"Оборачиваемость активов {at:.2f}x — ниже нормы. Часть активов может быть неэффективна."))
    else:
        conclusions.append(("Деловая активность", "bad",
            f"Оборачиваемость активов {at:.2f}x — НИЗКАЯ. Активы используются неэффективно."))

    # Overall score
    ok_count = sum(1 for _, s, _ in conclusions if s == "ok")
    score = ok_count / max(len(conclusions), 1) * 100
    if score >= 70:
        overall = "ПОЛОЖИТЕЛЬНАЯ ОЦЕНКА — компания финансово устойчива и показывает хорошие результаты"
    elif score >= 40:
        overall = "УМЕРЕННАЯ ОЦЕНКА — есть области для улучшения, общее состояние удовлетворительное"
    else:
        overall = "НЕГАТИВНАЯ ОЦЕНКА — требуются срочные меры по улучшению финансового состояния"

    return conclusions, overall


class ExportRequest(BaseModel):
    portfolio_id: int = 0


@router.post("/analytics/export/full-report")
def export_full_report(
    req: ExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Generate a comprehensive 18-sheet Excel report covering NSBU + IFRS
    balance, P&L, cash flow, equity, fixed assets, adjustments, KPIs,
    stress tests, investment decisions, visualizations, and reconciliation.
    """
    try:
        result = _do_export_full_report(req, db, current_user)
        logger.info("Export full report succeeded for user %s", current_user.id)
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Export full report failed: %s\n%s", e, traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Ошибка генерации отчёта: {str(e)}")


def _do_export_full_report(req: ExportRequest, db: Session, current_user: User):
    from datetime import date as _date
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink

    # --- Styling constants ---
    BLUE_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    PURPLE_FILL = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")
    AMBER_FILL = PatternFill(start_color="B7950B", end_color="B7950B", fill_type="solid")
    RED_FILL = PatternFill(start_color="922B21", end_color="922B21", fill_type="solid")
    GRAY_FILL = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    LIGHT_BLUE = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    LIGHT_PURPLE = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
    WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14)
    BOLD = Font(bold=True, size=10)
    NORMAL = Font(size=10)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )
    NUM_FMT = '#,##0'

    def style_header_row(ws, row_num, fill):
        for cell in ws[row_num]:
            cell.font = WHITE_FONT
            cell.fill = fill
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

    def style_data_cell(cell, is_number=False, is_bold=False):
        cell.font = BOLD if is_bold else NORMAL
        cell.border = THIN_BORDER
        if is_number and cell.value is not None:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right")

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    def write_no_data(ws, _fill=None):
        ws.append(["Нет данных. Импортируйте файл 1С и запустите конвертацию МСФО."])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
        ws[ws.max_row][0].font = Font(italic=True, size=11, color="7F8C8D")
        ws[ws.max_row][0].alignment = Alignment(horizontal="center")

    def write_sheet_header(ws, title, fill, company_info_d):
        """Write standard rows 1-3 (title, org+period, blank) then row 4 handled by caller."""
        ws.append([title])
        ws[ws.max_row][0].font = TITLE_FONT
        org = company_info_d.get("name", "") if company_info_d else ""
        period = company_info_d.get("period", "") if company_info_d else ""
        ws.append([f"Организация: {org}   |   Период: {period}"])
        ws[ws.max_row][0].font = BOLD
        ws.append([])  # blank row 3

    def _write_financial_stmt(ws, fill, light_fill, stmt_type, standard, headers, row_builder_fn, row_builder_args):
        """Generic helper to write a financial statement sheet from DB or cache."""
        from app.db.models.ifrs import FinancialStatement as FS
        ws.append(headers)
        style_header_row(ws, ws.max_row, fill)
        try:
            stmts = db.query(FS).filter(FS.statement_type == stmt_type, FS.standard == standard).order_by(FS.created_at.desc()).limit(2).all()
        except Exception:
            stmts = []
        if stmts and stmts[0].data:
            data = stmts[0].data
            rows_list = data if isinstance(data, list) else [{k: v} for k, v in data.items()] if isinstance(data, dict) else []
            for item in rows_list:
                if isinstance(item, dict):
                    ws.append([item.get("label", ""), item.get("current", item.get("amount")), item.get("previous")])
                    r = ws.max_row
                    is_hdr = item.get("isHeader") or item.get("isTotal")
                    style_data_cell(ws.cell(r, 1), is_bold=is_hdr)
                    style_data_cell(ws.cell(r, 2), is_number=True, is_bold=is_hdr)
                    style_data_cell(ws.cell(r, 3), is_number=True, is_bold=is_hdr)
                    if item.get("isHeader") and light_fill:
                        for ci in range(len(headers)):
                            ws.cell(r, ci + 1).fill = light_fill
        elif row_builder_fn:
            row_builder_fn(ws, *row_builder_args)
        else:
            write_no_data(ws, fill)

    # --- Get cached data ---
    cache = _user_cache(current_user.id)
    accounts = cache.get("accounts")
    pnl = cache.get("pnl")
    company_info = cache.get("company_info", {})

    wb = Workbook()

    # Sheet names (18 total) — order matters for hyperlinks in TOC
    SHEET_NAMES = [
        "Содержание",
        "Баланс НСБУ",
        "ОПиУ НСБУ",
        "ДДС НСБУ",
        "Капитал НСБУ",
        "Движение ОС",
        "Баланс МСФО",
        "Совокупный доход",
        "ДДС МСФО",
        "Капитал МСФО",
        "Корректировки",
        "Коэффициенты",
        "Стресс-тест НСБУ",
        "Стресс-тест МСФО",
        "Решения",
        "AI-анализ",
        "Визуализации",
        "Reconciliation НСБУ→МСФО",
    ]

    # ===================================================================
    # SHEET 1: Содержание (Table of Contents)
    # ===================================================================
    ws_toc = wb.active
    ws_toc.title = "Содержание"
    ws_toc.sheet_properties.tabColor = "374151"

    ws_toc.append(["Полный финансовый отчёт"])
    ws_toc[1][0].font = Font(bold=True, size=16)
    org_name_display = company_info.get("name", "—") if company_info else "—"
    inn_display = company_info.get("inn", "—") if company_info else "—"
    period_display = company_info.get("period", "—") if company_info else "—"
    ws_toc.append([f"Организация: {org_name_display}"])
    ws_toc[2][0].font = BOLD
    ws_toc.append([f"ИНН: {inn_display}"])
    ws_toc[3][0].font = BOLD
    ws_toc.append([f"Период: {period_display}"])
    ws_toc[4][0].font = BOLD
    ws_toc.append([f"Дата генерации: {_date.today().strftime('%d.%m.%Y')}"])
    ws_toc[5][0].font = BOLD
    ws_toc.append([])
    ws_toc.append(["Содержание:"])
    ws_toc[7][0].font = Font(bold=True, size=12)

    for idx, sname in enumerate(SHEET_NAMES):
        row_num = 8 + idx
        ws_toc.append([f"{idx + 1}. {sname}"])
        cell = ws_toc.cell(row=row_num, column=1)
        cell.font = Font(color="1D4ED8", underline="single", size=11)
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sname}'!A1", display=f"{idx + 1}. {sname}")

    ws_toc.column_dimensions["A"].width = 50
    auto_width(ws_toc)

    # ===================================================================
    # SHEET 2: Баланс НСБУ (Form 1)
    # ===================================================================
    ws_nsbu = wb.create_sheet("Баланс НСБУ")
    ws_nsbu.sheet_properties.tabColor = "1D4ED8"
    write_sheet_header(ws_nsbu, "Баланс (Форма 1) — НСБУ", BLUE_FILL, company_info)

    ws_nsbu.append(["Показатель", "Код", "Текущий период", "Предыдущий период"])
    style_header_row(ws_nsbu, ws_nsbu.max_row, BLUE_FILL)

    if accounts:
        for row in _build_nsbu_rows(accounts, pnl):
            ws_nsbu.append([row.get("label", ""), row.get("code", ""), row.get("current"), row.get("previous")])
            r = ws_nsbu.max_row
            is_hdr = row.get("isHeader") or row.get("isTotalAsset") or row.get("isTotalLiability")
            style_data_cell(ws_nsbu.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_nsbu.cell(r, 2))
            style_data_cell(ws_nsbu.cell(r, 3), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_nsbu.cell(r, 4), is_number=True, is_bold=is_hdr)
            if row.get("isHeader"):
                for ci in range(4):
                    ws_nsbu.cell(r, ci + 1).fill = LIGHT_BLUE
    else:
        write_no_data(ws_nsbu, BLUE_FILL)
    auto_width(ws_nsbu)

    # ===================================================================
    # SHEET 3: ОПиУ НСБУ (Form 2)
    # ===================================================================
    ws_pnl_nsbu = wb.create_sheet("ОПиУ НСБУ")
    ws_pnl_nsbu.sheet_properties.tabColor = "1D4ED8"
    write_sheet_header(ws_pnl_nsbu, "Отчёт о прибылях и убытках (Форма 2) — НСБУ", BLUE_FILL, company_info)

    ws_pnl_nsbu.append(["Показатель", "Текущий период", "Предыдущий период"])
    style_header_row(ws_pnl_nsbu, ws_pnl_nsbu.max_row, BLUE_FILL)

    from app.db.models.ifrs import FinancialStatement

    def _safe_query(query_fn):
        """Safely execute a DB query, return empty list on failure."""
        try:
            return query_fn()
        except Exception as exc:
            logger.warning("DB query failed in export: %s", exc)
            try:
                db.rollback()
            except Exception:
                pass
            return []

    pnl_nsbu_stmts = _safe_query(lambda: db.query(FinancialStatement).filter(
        FinancialStatement.statement_type == "P&L",
        FinancialStatement.standard == "nsbu",
    ).order_by(FinancialStatement.created_at.desc()).limit(2).all())

    pnl_nsbu_written = False
    if pnl_nsbu_stmts and pnl_nsbu_stmts[0].data:
        stmt_data = pnl_nsbu_stmts[0].data
        if isinstance(stmt_data, list):
            for item in stmt_data:
                ws_pnl_nsbu.append([item.get("label", ""), item.get("current", item.get("amount")), item.get("previous")])
                r = ws_pnl_nsbu.max_row
                is_hdr = item.get("isHeader") or item.get("isTotal")
                style_data_cell(ws_pnl_nsbu.cell(r, 1), is_bold=is_hdr)
                style_data_cell(ws_pnl_nsbu.cell(r, 2), is_number=True, is_bold=is_hdr)
                style_data_cell(ws_pnl_nsbu.cell(r, 3), is_number=True, is_bold=is_hdr)
            pnl_nsbu_written = True

    if not pnl_nsbu_written:
        # Also try generic P&L statements or cache
        pnl_generic = _safe_query(lambda: db.query(FinancialStatement).filter(
            FinancialStatement.statement_type == "P&L"
        ).order_by(FinancialStatement.created_at.desc()).limit(2).all())
        if pnl_generic and pnl_generic[0].data:
            stmt_data = pnl_generic[0].data
            if isinstance(stmt_data, list):
                for item in stmt_data:
                    ws_pnl_nsbu.append([item.get("label", ""), item.get("current", item.get("amount")), item.get("previous")])
                    r = ws_pnl_nsbu.max_row
                    is_hdr = item.get("isHeader") or item.get("isTotal")
                    style_data_cell(ws_pnl_nsbu.cell(r, 1), is_bold=is_hdr)
                    style_data_cell(ws_pnl_nsbu.cell(r, 2), is_number=True, is_bold=is_hdr)
                    style_data_cell(ws_pnl_nsbu.cell(r, 3), is_number=True, is_bold=is_hdr)
                pnl_nsbu_written = True

    # Try detailed breakdown from income_expenses cache
    if not pnl_nsbu_written:
        import re as _re_pnl
        from collections import OrderedDict
        ie_list = cache.get("income_expenses", [])
        if ie_list:
            # Classify items by section based on account code prefix
            section_map = OrderedDict([
                ("I. ВЫРУЧКА", []),
                ("II. СЕБЕСТОИМОСТЬ", []),
                ("III. ОПЕРАЦИОННЫЕ РАСХОДЫ", []),
                ("IV. ПРОЧИЕ ДОХОДЫ", []),
                ("V. ПРОЧИЕ РАСХОДЫ", []),
                ("VI. НАЛОГ НА ПРИБЫЛЬ", []),
            ])
            for item in ie_list:
                if not isinstance(item, dict):
                    continue
                name = str(item.get("name", ""))
                m = _re_pnl.match(r"(\d{4})", name)
                code = m.group(1) if m else ""
                cur = float(item.get("current_year") or item.get("current_period") or item.get("current") or 0)
                prev = float(item.get("previous_year") or item.get("previous_period") or item.get("previous") or 0)
                if code == "9720":
                    section_map["VI. НАЛОГ НА ПРИБЫЛЬ"].append((name, cur, prev))
                elif code.startswith("90"):
                    section_map["I. ВЫРУЧКА"].append((name, cur, prev))
                elif code.startswith("91"):
                    section_map["IV. ПРОЧИЕ ДОХОДЫ"].append((name, cur, prev))
                elif code.startswith("20"):
                    section_map["II. СЕБЕСТОИМОСТЬ"].append((name, cur, prev))
                elif code.startswith("94"):
                    section_map["III. ОПЕРАЦИОННЫЕ РАСХОДЫ"].append((name, cur, prev))
                elif code.startswith(("95", "96")):
                    section_map["V. ПРОЧИЕ РАСХОДЫ"].append((name, cur, prev))

            if any(items for items in section_map.values()):
                total_revenue_cur = 0.0
                total_revenue_prev = 0.0
                total_costs_cur = 0.0
                total_costs_prev = 0.0
                total_opex_cur = 0.0
                total_opex_prev = 0.0
                total_other_income_cur = 0.0
                total_other_income_prev = 0.0
                total_other_expenses_cur = 0.0
                total_other_expenses_prev = 0.0
                total_tax_cur = 0.0
                total_tax_prev = 0.0

                for sec_name, items_list in section_map.items():
                    if not items_list:
                        continue
                    # Section header
                    ws_pnl_nsbu.append([sec_name, None, None])
                    r = ws_pnl_nsbu.max_row
                    style_data_cell(ws_pnl_nsbu.cell(r, 1), is_bold=True)
                    for ci in range(3):
                        ws_pnl_nsbu.cell(r, ci + 1).fill = LIGHT_BLUE
                    sec_cur = 0.0
                    sec_prev = 0.0
                    for name, cur, prev in items_list:
                        ws_pnl_nsbu.append([f"  {name}", cur, prev])
                        r = ws_pnl_nsbu.max_row
                        style_data_cell(ws_pnl_nsbu.cell(r, 1))
                        style_data_cell(ws_pnl_nsbu.cell(r, 2), is_number=True)
                        style_data_cell(ws_pnl_nsbu.cell(r, 3), is_number=True)
                        sec_cur += cur
                        sec_prev += prev
                    # Section subtotal
                    ws_pnl_nsbu.append([f"Итого {sec_name}", sec_cur, sec_prev])
                    r = ws_pnl_nsbu.max_row
                    style_data_cell(ws_pnl_nsbu.cell(r, 1), is_bold=True)
                    style_data_cell(ws_pnl_nsbu.cell(r, 2), is_number=True, is_bold=True)
                    style_data_cell(ws_pnl_nsbu.cell(r, 3), is_number=True, is_bold=True)
                    if "ВЫРУЧКА" in sec_name and "ПРОЧИЕ" not in sec_name:
                        total_revenue_cur = sec_cur
                        total_revenue_prev = sec_prev
                    elif "СЕБЕСТОИМОСТЬ" in sec_name:
                        total_costs_cur = sec_cur
                        total_costs_prev = sec_prev
                    elif "ОПЕРАЦИОННЫЕ" in sec_name:
                        total_opex_cur = sec_cur
                        total_opex_prev = sec_prev
                    elif "ПРОЧИЕ ДОХОДЫ" in sec_name:
                        total_other_income_cur = sec_cur
                        total_other_income_prev = sec_prev
                    elif "ПРОЧИЕ РАСХОДЫ" in sec_name:
                        total_other_expenses_cur = sec_cur
                        total_other_expenses_prev = sec_prev
                    elif "НАЛОГ" in sec_name:
                        total_tax_cur = sec_cur
                        total_tax_prev = sec_prev

                # Detailed summary rows
                ws_pnl_nsbu.append([])  # blank row

                def _pnl_summary_row(label, cur_val, prev_val, is_highlight=False):
                    ws_pnl_nsbu.append([label, cur_val, prev_val])
                    r = ws_pnl_nsbu.max_row
                    style_data_cell(ws_pnl_nsbu.cell(r, 1), is_bold=True)
                    style_data_cell(ws_pnl_nsbu.cell(r, 2), is_number=True, is_bold=True)
                    style_data_cell(ws_pnl_nsbu.cell(r, 3), is_number=True, is_bold=True)
                    if is_highlight:
                        for ci in range(3):
                            ws_pnl_nsbu.cell(r, ci + 1).fill = LIGHT_BLUE

                gross_cur = total_revenue_cur - total_costs_cur
                gross_prev = total_revenue_prev - total_costs_prev
                _pnl_summary_row("ВАЛОВАЯ ПРИБЫЛЬ", gross_cur, gross_prev, is_highlight=True)

                op_profit_cur = gross_cur - total_opex_cur
                op_profit_prev = gross_prev - total_opex_prev
                _pnl_summary_row("ОПЕРАЦИОННАЯ ПРИБЫЛЬ", op_profit_cur, op_profit_prev)

                pbt_cur = op_profit_cur + total_other_income_cur - total_other_expenses_cur
                pbt_prev = op_profit_prev + total_other_income_prev - total_other_expenses_prev
                _pnl_summary_row("ПРИБЫЛЬ ДО НАЛОГООБЛОЖЕНИЯ", pbt_cur, pbt_prev)

                # Tax: use parsed tax if available, else estimate 15%
                if total_tax_cur == 0 and pbt_cur > 0:
                    total_tax_cur = round(pbt_cur * 0.15, 2)
                if total_tax_prev == 0 and pbt_prev > 0:
                    total_tax_prev = round(pbt_prev * 0.15, 2)
                _pnl_summary_row("НАЛОГ НА ПРИБЫЛЬ", total_tax_cur, total_tax_prev)

                net_cur = pbt_cur - total_tax_cur
                net_prev = pbt_prev - total_tax_prev
                _pnl_summary_row("ЧИСТАЯ ПРИБЫЛЬ", net_cur, net_prev, is_highlight=True)

                pnl_nsbu_written = True

    if not pnl_nsbu_written and pnl and (pnl.get("total_revenue_end", 0) or pnl.get("total_expenses_end", 0)):
        pnl_rows = [
            ("Выручка (доходы)", pnl.get("total_revenue_end", 0), pnl.get("total_revenue_begin", 0), True),
            ("Себестоимость (расходы)", pnl.get("total_expenses_end", 0), pnl.get("total_expenses_begin", 0), False),
            ("Валовая прибыль",
             pnl.get("total_revenue_end", 0) - pnl.get("total_expenses_end", 0),
             pnl.get("total_revenue_begin", 0) - pnl.get("total_expenses_begin", 0),
             True),
        ]
        for label, cur, prev, is_hdr in pnl_rows:
            ws_pnl_nsbu.append([label, cur, prev])
            r = ws_pnl_nsbu.max_row
            style_data_cell(ws_pnl_nsbu.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_pnl_nsbu.cell(r, 2), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_pnl_nsbu.cell(r, 3), is_number=True, is_bold=is_hdr)
        pnl_nsbu_written = True

    if not pnl_nsbu_written:
        write_no_data(ws_pnl_nsbu, BLUE_FILL)
    auto_width(ws_pnl_nsbu)

    # ===================================================================
    # SHEET 4: ДДС НСБУ (Form 4 — Cash Flow Statement)
    # ===================================================================
    ws_cf_nsbu = wb.create_sheet("ДДС НСБУ")
    ws_cf_nsbu.sheet_properties.tabColor = "1D4ED8"
    write_sheet_header(ws_cf_nsbu, "Отчёт о движении денежных средств (Форма 4) — НСБУ", BLUE_FILL, company_info)

    ws_cf_nsbu.append(["Показатель", "Текущий период", "Предыдущий период"])
    style_header_row(ws_cf_nsbu, ws_cf_nsbu.max_row, BLUE_FILL)

    cf_nsbu_stmts = _safe_query(lambda: db.query(FinancialStatement).filter(
        FinancialStatement.statement_type == "CF",
        FinancialStatement.standard == "nsbu",
    ).order_by(FinancialStatement.created_at.desc()).limit(2).all())

    cf_nsbu_written = False
    if cf_nsbu_stmts and cf_nsbu_stmts[0].data and isinstance(cf_nsbu_stmts[0].data, list):
        for item in cf_nsbu_stmts[0].data:
            ws_cf_nsbu.append([item.get("label", ""), item.get("current", item.get("amount")), item.get("previous")])
            r = ws_cf_nsbu.max_row
            is_hdr = item.get("isHeader") or item.get("isTotal")
            style_data_cell(ws_cf_nsbu.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_cf_nsbu.cell(r, 2), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_cf_nsbu.cell(r, 3), is_number=True, is_bold=is_hdr)
        cf_nsbu_written = True

    if not cf_nsbu_written:
        cached_cf = cache.get("cashflow", [])
        if cached_cf:
            # Group cashflow items by section (ОПЕРАЦИОННАЯ, ИНВЕСТИЦИОННАЯ, ФИНАНСОВАЯ)
            from collections import OrderedDict
            section_order = ["ОПЕРАЦИОННАЯ", "ИНВЕСТИЦИОННАЯ", "ФИНАНСОВАЯ"]
            sections: dict = OrderedDict()
            balance_items = []
            for item in cached_cf:
                sec = item.get("section", "")
                grp = (item.get("group") or "").upper().strip()
                if sec == "balances":
                    balance_items.append(item)
                    continue
                # Normalize group name to a known section
                matched = None
                for s in section_order:
                    if s in grp:
                        matched = s
                        break
                if matched is None:
                    matched = grp or "ПРОЧЕЕ"
                sections.setdefault(matched, []).append(item)

            grand_total = 0.0
            grand_total_prev = 0.0
            for sec_name in section_order:
                items_in_sec = sections.get(sec_name, [])
                if not items_in_sec:
                    continue
                header_label = {
                    "ОПЕРАЦИОННАЯ": "I. ОПЕРАЦИОННАЯ ДЕЯТЕЛЬНОСТЬ",
                    "ИНВЕСТИЦИОННАЯ": "II. ИНВЕСТИЦИОННАЯ ДЕЯТЕЛЬНОСТЬ",
                    "ФИНАНСОВАЯ": "III. ФИНАНСОВАЯ ДЕЯТЕЛЬНОСТЬ",
                }.get(sec_name, sec_name)
                ws_cf_nsbu.append([header_label, None, None])
                r = ws_cf_nsbu.max_row
                style_data_cell(ws_cf_nsbu.cell(r, 1), is_bold=True)
                for ci in range(3):
                    ws_cf_nsbu.cell(r, ci + 1).fill = LIGHT_BLUE
                subtotal = 0.0
                subtotal_prev = 0.0
                for it in items_in_sec:
                    name = it.get("name") or it.get("label") or ""
                    net_val = float(it.get("net") or 0)
                    inflow = float(it.get("inflow") or 0)
                    outflow = float(it.get("outflow") or 0)
                    current_val = net_val if net_val else (inflow - outflow)
                    prev_net = float(it.get("previous_net") or it.get("previous") or it.get("net_previous") or 0)
                    prev_in = float(it.get("previous_inflow") or 0)
                    prev_out = float(it.get("previous_outflow") or 0)
                    prev_val = prev_net if prev_net else (prev_in - prev_out)
                    subtotal += current_val
                    subtotal_prev += prev_val
                    ws_cf_nsbu.append([f"  {name}", current_val, prev_val if prev_val else None])
                    r = ws_cf_nsbu.max_row
                    style_data_cell(ws_cf_nsbu.cell(r, 1))
                    style_data_cell(ws_cf_nsbu.cell(r, 2), is_number=True)
                    style_data_cell(ws_cf_nsbu.cell(r, 3), is_number=True)
                ws_cf_nsbu.append([f"Итого {header_label.split('. ', 1)[-1]}", subtotal, subtotal_prev if subtotal_prev else None])
                r = ws_cf_nsbu.max_row
                style_data_cell(ws_cf_nsbu.cell(r, 1), is_bold=True)
                style_data_cell(ws_cf_nsbu.cell(r, 2), is_number=True, is_bold=True)
                style_data_cell(ws_cf_nsbu.cell(r, 3), is_number=True, is_bold=True)
                grand_total += subtotal
                grand_total_prev += subtotal_prev

            # Write remaining sections (e.g. ПРОЧЕЕ)
            for sec_name, items_in_sec in sections.items():
                if sec_name in section_order:
                    continue
                for it in items_in_sec:
                    name = it.get("name") or it.get("label") or ""
                    net_val = float(it.get("net") or 0)
                    inflow = float(it.get("inflow") or 0)
                    outflow = float(it.get("outflow") or 0)
                    current_val = net_val if net_val else (inflow - outflow)
                    prev_net = float(it.get("previous_net") or it.get("previous") or it.get("net_previous") or 0)
                    prev_in = float(it.get("previous_inflow") or 0)
                    prev_out = float(it.get("previous_outflow") or 0)
                    prev_val = prev_net if prev_net else (prev_in - prev_out)
                    grand_total += current_val
                    grand_total_prev += prev_val
                    ws_cf_nsbu.append([name, current_val, prev_val if prev_val else None])
                    r = ws_cf_nsbu.max_row
                    style_data_cell(ws_cf_nsbu.cell(r, 1))
                    style_data_cell(ws_cf_nsbu.cell(r, 2), is_number=True)
                    style_data_cell(ws_cf_nsbu.cell(r, 3), is_number=True)

            ws_cf_nsbu.append(["ИТОГО ЧИСТЫЙ ДЕНЕЖНЫЙ ПОТОК", grand_total, grand_total_prev if grand_total_prev else None])
            r = ws_cf_nsbu.max_row
            style_data_cell(ws_cf_nsbu.cell(r, 1), is_bold=True)
            style_data_cell(ws_cf_nsbu.cell(r, 2), is_number=True, is_bold=True)
            style_data_cell(ws_cf_nsbu.cell(r, 3), is_number=True, is_bold=True)

            # Cash balances (beginning/end)
            if balance_items:
                ws_cf_nsbu.append([])
                for bi in balance_items:
                    name = bi.get("name") or bi.get("label") or ""
                    val = float(bi.get("amount") or bi.get("net") or 0)
                    prev_val = float(bi.get("previous") or 0)
                    ws_cf_nsbu.append([name, val, prev_val or None])
                    r = ws_cf_nsbu.max_row
                    style_data_cell(ws_cf_nsbu.cell(r, 1), is_bold=True)
                    style_data_cell(ws_cf_nsbu.cell(r, 2), is_number=True, is_bold=True)
                    style_data_cell(ws_cf_nsbu.cell(r, 3), is_number=True, is_bold=True)
            cf_nsbu_written = True

    if not cf_nsbu_written:
        write_no_data(ws_cf_nsbu, BLUE_FILL)
    auto_width(ws_cf_nsbu)

    # ===================================================================
    # SHEET 5: Капитал НСБУ (Form 5 — Changes in Equity)
    # ===================================================================
    ws_eq_nsbu = wb.create_sheet("Капитал НСБУ")
    ws_eq_nsbu.sheet_properties.tabColor = "1D4ED8"
    write_sheet_header(ws_eq_nsbu, "Отчёт об изменениях в капитале (Форма 5) — НСБУ", BLUE_FILL, company_info)

    ws_eq_nsbu.append(["Показатель", "Остаток на начало", "Движение", "Остаток на конец"])
    style_header_row(ws_eq_nsbu, ws_eq_nsbu.max_row, BLUE_FILL)

    eq_nsbu_stmts = _safe_query(lambda: db.query(FinancialStatement).filter(
        FinancialStatement.statement_type == "Equity",
        FinancialStatement.standard == "nsbu",
    ).order_by(FinancialStatement.created_at.desc()).limit(2).all())

    eq_nsbu_written = False
    if eq_nsbu_stmts and eq_nsbu_stmts[0].data and isinstance(eq_nsbu_stmts[0].data, list):
        for item in eq_nsbu_stmts[0].data:
            ws_eq_nsbu.append([
                item.get("label", item.get("name", "")),
                item.get("balance_start", item.get("previous")),
                item.get("movement"),
                item.get("balance_end", item.get("current", item.get("amount"))),
            ])
            r = ws_eq_nsbu.max_row
            is_hdr = item.get("isHeader") or item.get("isTotal")
            style_data_cell(ws_eq_nsbu.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_eq_nsbu.cell(r, 2), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_eq_nsbu.cell(r, 3), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_eq_nsbu.cell(r, 4), is_number=True, is_bold=is_hdr)
        eq_nsbu_written = True

    if not eq_nsbu_written:
        cached_eq = cache.get("capital_rows", [])
        if cached_eq:
            total_start = 0.0
            total_movement = 0.0
            total_end = 0.0
            for item in cached_eq:
                name = item.get("name") or item.get("label") or ""
                b_start = float(item.get("balance_start") or item.get("previous") or 0)
                movement = float(item.get("movement") or 0)
                b_end = float(item.get("balance_end") or item.get("current") or 0)
                total_start += b_start
                total_movement += movement
                total_end += b_end
                ws_eq_nsbu.append([name, b_start, movement, b_end])
                r = ws_eq_nsbu.max_row
                style_data_cell(ws_eq_nsbu.cell(r, 1))
                style_data_cell(ws_eq_nsbu.cell(r, 2), is_number=True)
                style_data_cell(ws_eq_nsbu.cell(r, 3), is_number=True)
                style_data_cell(ws_eq_nsbu.cell(r, 4), is_number=True)
            # Total row
            ws_eq_nsbu.append(["ИТОГО КАПИТАЛ", total_start, total_movement, total_end])
            r = ws_eq_nsbu.max_row
            style_data_cell(ws_eq_nsbu.cell(r, 1), is_bold=True)
            style_data_cell(ws_eq_nsbu.cell(r, 2), is_number=True, is_bold=True)
            style_data_cell(ws_eq_nsbu.cell(r, 3), is_number=True, is_bold=True)
            style_data_cell(ws_eq_nsbu.cell(r, 4), is_number=True, is_bold=True)
            eq_nsbu_written = True

    if not eq_nsbu_written:
        write_no_data(ws_eq_nsbu, BLUE_FILL)
    auto_width(ws_eq_nsbu)

    # ===================================================================
    # SHEET 6: Движение ОС (Form 3 — Fixed Assets Movement)
    # ===================================================================
    ws_fa = wb.create_sheet("Движение ОС")
    ws_fa.sheet_properties.tabColor = "1D4ED8"
    write_sheet_header(ws_fa, "Движение основных средств (Форма 3) — НСБУ", BLUE_FILL, company_info)

    ws_fa.append(["Группа ОС", "На начало", "Поступило", "Выбыло", "На конец", "Амортизация"])
    style_header_row(ws_fa, ws_fa.max_row, BLUE_FILL)

    if accounts:
        gross_fa = accounts.get("0100", {})
        depr = accounts.get("0200", {})
        capex_acc = accounts.get("0800", {})
        fa_begin = (gross_fa.get("previous") or 0)
        fa_end = (gross_fa.get("current") or 0)
        depr_begin = (depr.get("previous") or 0)
        depr_end = (depr.get("current") or 0)
        # Real additions from debit turnover, disposals from credit turnover
        additions = float(gross_fa.get("debit_turnover") or gross_fa.get("debit_end") or 0)
        disposals = float(gross_fa.get("credit_turnover") or gross_fa.get("credit_end") or 0)
        # Fallback: compute from balances if turnover not available
        if additions == 0 and disposals == 0:
            capex_val = (capex_acc.get("current") or 0) - (capex_acc.get("previous") or 0)
            additions = max(capex_val, 0) if capex_val > 0 else max(fa_end - fa_begin, 0)
            disposals = fa_begin + additions - fa_end if fa_begin + additions > fa_end else 0
        depr_period = depr_end - depr_begin
        net_begin = fa_begin - depr_begin
        net_end = fa_end - depr_end

        ws_fa.append(["Первоначальная стоимость ОС", fa_begin, additions, disposals, fa_end, None])
        r = ws_fa.max_row
        for ci in range(6):
            style_data_cell(ws_fa.cell(r, ci + 1), is_number=(ci > 0))
        ws_fa.append(["Накопленная амортизация", depr_begin, depr_period, None, depr_end, depr_end])
        r = ws_fa.max_row
        for ci in range(6):
            style_data_cell(ws_fa.cell(r, ci + 1), is_number=(ci > 0))
        ws_fa.append(["Остаточная стоимость ОС", net_begin, None, None, net_end, None])
        r = ws_fa.max_row
        style_data_cell(ws_fa.cell(r, 1), is_bold=True)
        for ci in range(1, 6):
            style_data_cell(ws_fa.cell(r, ci + 1), is_number=True, is_bold=True)
        for ci in range(6):
            ws_fa.cell(r, ci + 1).fill = LIGHT_BLUE
        ws_fa.append([])
        ws_fa.append(["Начислено амортизации за период", None, None, None, None, depr_period])
        r = ws_fa.max_row
        style_data_cell(ws_fa.cell(r, 1), is_bold=True)
        style_data_cell(ws_fa.cell(r, 6), is_number=True, is_bold=True)
        # Capex line
        capex_begin = (capex_acc.get("previous") or 0)
        capex_end = (capex_acc.get("current") or 0)
        if capex_begin or capex_end:
            ws_fa.append(["Капитальные вложения (0800)", capex_begin, None, None, capex_end, None])
            r = ws_fa.max_row
            for ci in range(6):
                style_data_cell(ws_fa.cell(r, ci + 1), is_number=(ci > 0))
    else:
        write_no_data(ws_fa, BLUE_FILL)
    auto_width(ws_fa)

    # ===================================================================
    # SHEET 7: Баланс МСФО (IAS 1 — Financial Position)
    # ===================================================================
    ws_ifrs = wb.create_sheet("Баланс МСФО")
    ws_ifrs.sheet_properties.tabColor = "8B5CF6"
    write_sheet_header(ws_ifrs, "Отчёт о финансовом положении (IAS 1) — МСФО", PURPLE_FILL, company_info)

    ws_ifrs.append(["Показатель", "Примечание", "Текущий период", "Предыдущий период"])
    style_header_row(ws_ifrs, ws_ifrs.max_row, PURPLE_FILL)

    if accounts:
        for row in _build_ifrs_rows(accounts, pnl):
            ws_ifrs.append([row.get("label", ""), row.get("note", ""), row.get("current"), row.get("previous")])
            r = ws_ifrs.max_row
            is_hdr = row.get("isHeader") or row.get("isTotal")
            style_data_cell(ws_ifrs.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_ifrs.cell(r, 2))
            style_data_cell(ws_ifrs.cell(r, 3), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_ifrs.cell(r, 4), is_number=True, is_bold=is_hdr)
            if row.get("isHeader"):
                for ci in range(4):
                    ws_ifrs.cell(r, ci + 1).fill = LIGHT_PURPLE
    else:
        write_no_data(ws_ifrs, PURPLE_FILL)
    auto_width(ws_ifrs)

    # ===================================================================
    # SHEET 8: Совокупный доход (IAS 1 — P&L + OCI)
    # ===================================================================
    ws_oci = wb.create_sheet("Совокупный доход")
    ws_oci.sheet_properties.tabColor = "8B5CF6"
    write_sheet_header(ws_oci, "Отчёт о совокупном доходе (IAS 1) — МСФО", PURPLE_FILL, company_info)

    ws_oci.append(["Показатель", "Текущий период", "Предыдущий период"])
    style_header_row(ws_oci, ws_oci.max_row, PURPLE_FILL)

    oci_stmts = _safe_query(lambda: db.query(FinancialStatement).filter(
        FinancialStatement.statement_type == "P&L",
        FinancialStatement.standard == "ifrs",
    ).order_by(FinancialStatement.created_at.desc()).limit(2).all())

    if oci_stmts and oci_stmts[0].data and isinstance(oci_stmts[0].data, list):
        for item in oci_stmts[0].data:
            ws_oci.append([item.get("label", ""), item.get("current", item.get("amount")), item.get("previous")])
            r = ws_oci.max_row
            is_hdr = item.get("isHeader") or item.get("isTotal")
            style_data_cell(ws_oci.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_oci.cell(r, 2), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_oci.cell(r, 3), is_number=True, is_bold=is_hdr)
            if item.get("isHeader"):
                for ci in range(3):
                    ws_oci.cell(r, ci + 1).fill = LIGHT_PURPLE
    elif accounts:
        # Use the IFRS income builder which pulls from pnl + income_expenses
        _ie = cache.get("income_expenses", [])
        for item in _build_ifrs_income_rows(accounts, pnl, income_expenses=_ie):
            ws_oci.append([item.get("label", ""), item.get("current", item.get("amount")), item.get("previous")])
            r = ws_oci.max_row
            is_hdr = item.get("isHeader") or item.get("isTotal")
            style_data_cell(ws_oci.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_oci.cell(r, 2), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_oci.cell(r, 3), is_number=True, is_bold=is_hdr)
            if item.get("isHeader"):
                for ci in range(3):
                    ws_oci.cell(r, ci + 1).fill = LIGHT_PURPLE
    elif pnl and (pnl.get("total_revenue_end", 0) or pnl.get("total_expenses_end", 0)):
        rev_e = pnl.get("total_revenue_end", 0)
        exp_e = pnl.get("total_expenses_end", 0)
        rev_b = pnl.get("total_revenue_begin", 0)
        exp_b = pnl.get("total_expenses_begin", 0)
        for label, cur, prev, hdr in [
            ("Выручка", rev_e, rev_b, True),
            ("Себестоимость", exp_e, exp_b, False),
            ("Операционная прибыль", rev_e - exp_e, rev_b - exp_b, True),
            ("Прочий совокупный доход (OCI)", 0, 0, False),
            ("Итого совокупный доход", rev_e - exp_e, rev_b - exp_b, True),
        ]:
            ws_oci.append([label, cur, prev])
            r = ws_oci.max_row
            style_data_cell(ws_oci.cell(r, 1), is_bold=hdr)
            style_data_cell(ws_oci.cell(r, 2), is_number=True, is_bold=hdr)
            style_data_cell(ws_oci.cell(r, 3), is_number=True, is_bold=hdr)
    else:
        write_no_data(ws_oci, PURPLE_FILL)
    auto_width(ws_oci)

    # ===================================================================
    # SHEET 9: ДДС МСФО (IAS 7)
    # ===================================================================
    ws_cf_ifrs = wb.create_sheet("ДДС МСФО")
    ws_cf_ifrs.sheet_properties.tabColor = "8B5CF6"
    write_sheet_header(ws_cf_ifrs, "Отчёт о движении денежных средств (IAS 7) — МСФО", PURPLE_FILL, company_info)

    ws_cf_ifrs.append(["Показатель", "Текущий период", "Предыдущий период"])
    style_header_row(ws_cf_ifrs, ws_cf_ifrs.max_row, PURPLE_FILL)

    cf_ifrs_stmts = _safe_query(lambda: db.query(FinancialStatement).filter(
        FinancialStatement.statement_type == "CF",
        FinancialStatement.standard == "ifrs",
    ).order_by(FinancialStatement.created_at.desc()).limit(2).all())

    cf_ifrs_written = False
    if cf_ifrs_stmts and cf_ifrs_stmts[0].data and isinstance(cf_ifrs_stmts[0].data, list):
        for item in cf_ifrs_stmts[0].data:
            ws_cf_ifrs.append([item.get("label", ""), item.get("current", item.get("amount")), item.get("previous")])
            r = ws_cf_ifrs.max_row
            is_hdr = item.get("isHeader") or item.get("isTotal")
            style_data_cell(ws_cf_ifrs.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_cf_ifrs.cell(r, 2), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_cf_ifrs.cell(r, 3), is_number=True, is_bold=is_hdr)
        cf_ifrs_written = True

    if not cf_ifrs_written and accounts:
        # Build IFRS CashFlow from cached balance data
        _ie = cache.get("income_expenses", [])
        _cf = cache.get("cashflow", [])
        _ifrs_cf_rows = _build_ifrs_cashflow_rows(accounts, pnl, income_expenses=_ie, cashflow_data=_cf)
        for item in _ifrs_cf_rows:
            ws_cf_ifrs.append([item.get("label", ""), item.get("current", item.get("amount")), item.get("previous")])
            r = ws_cf_ifrs.max_row
            is_hdr = item.get("isHeader") or item.get("isTotal")
            style_data_cell(ws_cf_ifrs.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_cf_ifrs.cell(r, 2), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_cf_ifrs.cell(r, 3), is_number=True, is_bold=is_hdr)
        cf_ifrs_written = True

    if not cf_ifrs_written:
        write_no_data(ws_cf_ifrs, PURPLE_FILL)
    auto_width(ws_cf_ifrs)

    # ===================================================================
    # SHEET 10: Капитал МСФО (IAS 1 — Changes in Equity)
    # ===================================================================
    ws_eq_ifrs = wb.create_sheet("Капитал МСФО")
    ws_eq_ifrs.sheet_properties.tabColor = "8B5CF6"
    write_sheet_header(ws_eq_ifrs, "Отчёт об изменениях в капитале (IAS 1) — МСФО", PURPLE_FILL, company_info)

    ws_eq_ifrs.append(["Показатель", "Текущий период", "Предыдущий период"])
    style_header_row(ws_eq_ifrs, ws_eq_ifrs.max_row, PURPLE_FILL)

    eq_ifrs_stmts = _safe_query(lambda: db.query(FinancialStatement).filter(
        FinancialStatement.statement_type == "Equity",
        FinancialStatement.standard == "ifrs",
    ).order_by(FinancialStatement.created_at.desc()).limit(2).all())

    eq_ifrs_written = False
    if eq_ifrs_stmts and eq_ifrs_stmts[0].data and isinstance(eq_ifrs_stmts[0].data, list):
        for item in eq_ifrs_stmts[0].data:
            ws_eq_ifrs.append([item.get("label", ""), item.get("current", item.get("amount")), item.get("previous")])
            r = ws_eq_ifrs.max_row
            is_hdr = item.get("isHeader") or item.get("isTotal")
            style_data_cell(ws_eq_ifrs.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_eq_ifrs.cell(r, 2), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_eq_ifrs.cell(r, 3), is_number=True, is_bold=is_hdr)
        eq_ifrs_written = True

    if not eq_ifrs_written and accounts:
        _ifrs_eq_rows = _build_ifrs_equity_rows(accounts, pnl, income_expenses=cache.get("income_expenses", []))
        for item in _ifrs_eq_rows:
            ws_eq_ifrs.append([item.get("label", ""), item.get("current", item.get("amount")), item.get("previous")])
            r = ws_eq_ifrs.max_row
            is_hdr = item.get("isHeader") or item.get("isTotal")
            style_data_cell(ws_eq_ifrs.cell(r, 1), is_bold=is_hdr)
            style_data_cell(ws_eq_ifrs.cell(r, 2), is_number=True, is_bold=is_hdr)
            style_data_cell(ws_eq_ifrs.cell(r, 3), is_number=True, is_bold=is_hdr)
        eq_ifrs_written = True

    if not eq_ifrs_written:
        write_no_data(ws_eq_ifrs, PURPLE_FILL)
    auto_width(ws_eq_ifrs)

    # ===================================================================
    # SHEET 11: Корректировки НСБУ → МСФО
    # ===================================================================
    ws_adj = wb.create_sheet("Корректировки")
    ws_adj.sheet_properties.tabColor = "B7950B"
    write_sheet_header(ws_adj, "Корректировки НСБУ → МСФО", AMBER_FILL, company_info)

    ws_adj.append(["Тип", "Счёт", "Сумма НСБУ", "Сумма МСФО", "Разница", "Описание"])
    style_header_row(ws_adj, ws_adj.max_row, AMBER_FILL)

    try:
        from app.db.models.ifrs import IFRSAdjustment
        adjustments = db.query(IFRSAdjustment).order_by(IFRSAdjustment.created_at.desc()).limit(200).all()
    except Exception:
        adjustments = []

    adj_type_labels = {
        "ifrs16_lease": "МСФО 16 Аренда",
        "ias16_revaluation": "МСФО 16 Переоценка",
        "ias36_impairment": "МСФО 36 Обесценение",
        "oci": "Прочий совокупный доход",
    }

    if adjustments:
        for adj in adjustments:
            ws_adj.append([
                adj_type_labels.get(adj.adjustment_type, adj.adjustment_type or ""),
                adj.account_code or "",
                float(adj.nsbu_amount) if adj.nsbu_amount is not None else None,
                float(adj.ifrs_amount) if adj.ifrs_amount is not None else None,
                float(adj.difference) if adj.difference is not None else None,
                adj.description or "",
            ])
            r = ws_adj.max_row
            for ci in range(6):
                style_data_cell(ws_adj.cell(r, ci + 1), is_number=(ci in (2, 3, 4)))
    elif accounts:
        # Compute actual IFRS adjustments from accounts data
        def _adj_v(code: str) -> float:
            acc = accounts.get(code)
            return acc["current"] if acc else 0.0

        net_fa = _adj_v("0100") - _adj_v("0200")
        receivables_adj = _adj_v("2010") + _adj_v("2300")
        lease_val = _adj_v("7800")
        lease_payment = _adj_v("6970")

        # IAS 16: PPE revaluation — 15% premium on net fixed assets
        ias16_reval = round(net_fa * 0.15, 2)
        # IFRS 9: ECL impairment — 5% of receivables
        ecl_impairment = round(receivables_adj * 0.05, 2)
        # IFRS 16: Lease ROU asset computation
        ifrs16_rou = 0.0
        if lease_payment > 0:
            discount_rate = 0.18
            lease_term = 5
            pv_factor = (1 - (1 + discount_rate) ** (-lease_term)) / discount_rate
            ifrs16_rou = round(lease_payment * pv_factor, 2)
        elif lease_val > 0:
            ifrs16_rou = round(lease_val * 0.10, 2)  # Approximate ROU adjustment

        # IAS 12: Deferred tax on revaluation (15% of IAS 16 revaluation)
        deferred_tax = round(ias16_reval * 0.15, 2)

        computed_adjustments = [
            ("IAS 16 Переоценка ОС", "0100", net_fa, net_fa + ias16_reval, ias16_reval,
             f"Переоценка основных средств +15% = +{ias16_reval:,.0f}. Увеличивает прочий совокупный доход (OCI)."),
            ("IFRS 9 ECL-резерв", "2010", receivables_adj, receivables_adj - ecl_impairment, -ecl_impairment,
             f"Резерв ожидаемых кредитных убытков 5% = -{ecl_impairment:,.0f}. Уменьшает дебиторскую задолженность."),
        ]
        if ifrs16_rou > 0:
            computed_adjustments.append(
                ("IFRS 16 Аренда (ROU)", "7800", lease_val or lease_payment, (lease_val or lease_payment) + ifrs16_rou, ifrs16_rou,
                 f"Актив права пользования (ROU) = +{ifrs16_rou:,.0f}. Капитализация операционной аренды."),
            )
        computed_adjustments.append(
            ("IAS 12 Отложенный налог", "OCI", 0, deferred_tax, -deferred_tax,
             f"ОНО на переоценку 15% = -{deferred_tax:,.0f}. Уменьшает OCI, увеличивает отложенные обязательства."),
        )

        # Summary
        total_diff = sum(a[4] for a in computed_adjustments)
        computed_adjustments.append(
            ("ИТОГО КОРРЕКТИРОВКИ", "", None, None, total_diff,
             f"Чистый эффект МСФО-корректировок: {'+' if total_diff > 0 else ''}{total_diff:,.0f}"),
        )

        for adj_type, acc_code, nsbu_val, ifrs_val, diff, desc in computed_adjustments:
            is_total = adj_type.startswith("ИТОГО")
            ws_adj.append([adj_type, acc_code, nsbu_val, ifrs_val, diff, desc])
            r = ws_adj.max_row
            for ci in range(6):
                style_data_cell(ws_adj.cell(r, ci + 1), is_number=(ci in (2, 3, 4)), is_bold=is_total)
            # Color the diff column
            diff_cell = ws_adj.cell(r, 5)
            if diff and diff > 0:
                diff_cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            elif diff and diff < 0:
                diff_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    else:
        write_no_data(ws_adj, AMBER_FILL)
    auto_width(ws_adj)

    # ===================================================================
    # SHEET 12: Коэффициенты (KPIs — НСБУ и МСФО рядом)
    # ===================================================================
    ws_kpi = wb.create_sheet("Коэффициенты")
    ws_kpi.sheet_properties.tabColor = "1E8449"
    write_sheet_header(ws_kpi, "Финансовые коэффициенты (НСБУ и МСФО)", GREEN_FILL, company_info)

    ws_kpi.append(["Показатель", "Группа", "НСБУ", "МСФО", "Норма", "Статус НСБУ", "Статус МСФО"])
    style_header_row(ws_kpi, ws_kpi.max_row, GREEN_FILL)

    agg = _get_balance_aggregates(user_id=current_user.id)
    if agg:
        agg_ifrs_kpi = _ifrs_adjusted_aggregates(agg)
        nsbu_kpis = _calc_kpis(agg)
        ifrs_kpis = _calc_kpis(agg_ifrs_kpi)
        for g_idx, group in enumerate(nsbu_kpis):
            ifrs_group = ifrs_kpis[g_idx] if g_idx < len(ifrs_kpis) else group
            for m_idx, m in enumerate(group["metrics"]):
                m_ifrs = ifrs_group["metrics"][m_idx] if m_idx < len(ifrs_group["metrics"]) else m
                ws_kpi.append([
                    m["label"], group["title"], m["value"], m_ifrs["value"],
                    m["norm"], m["status"], m_ifrs["status"],
                ])
                r = ws_kpi.max_row
                for ci in range(7):
                    style_data_cell(ws_kpi.cell(r, ci + 1), is_number=(ci in (2, 3)))
                for ci, st in [(5, m["status"]), (6, m_ifrs["status"])]:
                    st_cell = ws_kpi.cell(r, ci + 1)
                    if st == "ok":
                        st_cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                    elif st == "warn":
                        st_cell.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                    elif st == "bad":
                        st_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    else:
        write_no_data(ws_kpi, GREEN_FILL)
    auto_width(ws_kpi)

    # ===================================================================
    # SHEET 13: Стресс-тест НСБУ
    # ===================================================================
    ws_stress_nsbu = wb.create_sheet("Стресс-тест НСБУ")
    ws_stress_nsbu.sheet_properties.tabColor = "922B21"
    write_sheet_header(ws_stress_nsbu, "Стресс-тест — НСБУ", RED_FILL, company_info)

    ws_stress_nsbu.append(["Показатель", "Базовое значение", "Стресс-значение", "Изменение %", "Статус"])
    style_header_row(ws_stress_nsbu, ws_stress_nsbu.max_row, RED_FILL)

    if agg:
        for sc_key, sc in _DEFAULT_SCENARIOS.items():
            ws_stress_nsbu.append([f"--- {sc['name']} ---", "", "", "", ""])
            r = ws_stress_nsbu.max_row
            for ci in range(5):
                ws_stress_nsbu.cell(r, ci + 1).font = BOLD
                ws_stress_nsbu.cell(r, ci + 1).fill = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")

            stressed_rev = agg["revenue"] * (1 + sc["revenue_shock"])
            stressed_exp = agg["expenses"] * (1 + sc["cost_shock"])
            interest_extra = agg["lt_liabilities"] * sc["interest_shock"] * 0.10
            fx_extra = agg["total_assets"] * 0.05 * sc["fx_shock"]
            stressed_np = stressed_rev - stressed_exp - interest_extra - fx_extra
            stressed_agg = dict(agg)
            stressed_agg["revenue"] = stressed_rev
            stressed_agg["expenses"] = stressed_exp
            stressed_agg["net_profit"] = stressed_np

            baseline_flat = _kpi_flat(agg)
            stressed_flat = _kpi_flat(stressed_agg)
            labels = _kpi_labels()

            for key, base_val in baseline_flat.items():
                stressed_val = stressed_flat.get(key, 0)
                delta_pct = round((stressed_val - base_val) / base_val * 100, 1) if base_val != 0 else 0.0
                ws_stress_nsbu.append([
                    labels.get(key, key), round(base_val, 4), round(stressed_val, 4),
                    delta_pct, _stress_status(delta_pct),
                ])
                r = ws_stress_nsbu.max_row
                for ci in range(5):
                    style_data_cell(ws_stress_nsbu.cell(r, ci + 1), is_number=(ci in (1, 2, 3)))
                st_cell = ws_stress_nsbu.cell(r, 5)
                status_val = _stress_status(delta_pct)
                if status_val == "ok":
                    st_cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                elif status_val == "warn":
                    st_cell.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                elif status_val == "bad":
                    st_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    else:
        # Fallback: DB stress tests
        try:
            from app.db.models.stress_retrospective import StressTest as StressTestModel
            stress_tests = db.query(StressTestModel).order_by(StressTestModel.created_at.desc()).limit(20).all()
        except Exception:
            stress_tests = []
        if stress_tests:
            # Rewrite headers for DB data
            ws_stress_nsbu.delete_rows(ws_stress_nsbu.max_row)
            ws_stress_nsbu.append(["Сценарий", "Стоимость до", "Стоимость после", "Потеря %", "Макс. потеря актива %"])
            style_header_row(ws_stress_nsbu, ws_stress_nsbu.max_row, RED_FILL)
            for st in stress_tests:
                ws_stress_nsbu.append([
                    st.scenario_name or "", st.portfolio_value_before, st.portfolio_value_after,
                    round(st.total_loss_pct, 2) if st.total_loss_pct is not None else None,
                    round(st.max_single_asset_loss_pct, 2) if st.max_single_asset_loss_pct is not None else None,
                ])
                r = ws_stress_nsbu.max_row
                for ci in range(5):
                    style_data_cell(ws_stress_nsbu.cell(r, ci + 1), is_number=(ci > 0))
        else:
            write_no_data(ws_stress_nsbu, RED_FILL)
    auto_width(ws_stress_nsbu)

    # ===================================================================
    # SHEET 14: Стресс-тест МСФО
    # ===================================================================
    ws_stress_ifrs = wb.create_sheet("Стресс-тест МСФО")
    ws_stress_ifrs.sheet_properties.tabColor = "922B21"
    write_sheet_header(ws_stress_ifrs, "Стресс-тест — МСФО", RED_FILL, company_info)

    ws_stress_ifrs.append(["Показатель", "Базовое значение", "Стресс-значение", "Изменение %", "Статус"])
    style_header_row(ws_stress_ifrs, ws_stress_ifrs.max_row, RED_FILL)

    agg_ifrs = _ifrs_adjusted_aggregates(agg) if agg else None
    if agg_ifrs:
        for sc_key, sc in _DEFAULT_SCENARIOS.items():
            ws_stress_ifrs.append([f"--- {sc['name']} ---", "", "", "", ""])
            r = ws_stress_ifrs.max_row
            for ci in range(5):
                ws_stress_ifrs.cell(r, ci + 1).font = BOLD
                ws_stress_ifrs.cell(r, ci + 1).fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")

            stressed_rev = agg_ifrs["revenue"] * (1 + sc["revenue_shock"])
            stressed_exp = agg_ifrs["expenses"] * (1 + sc["cost_shock"])
            interest_extra = agg_ifrs["lt_liabilities"] * sc["interest_shock"] * 0.10
            fx_extra = agg_ifrs["total_assets"] * 0.05 * sc["fx_shock"]
            stressed_np = stressed_rev - stressed_exp - interest_extra - fx_extra
            stressed_agg = dict(agg_ifrs)
            stressed_agg["revenue"] = stressed_rev
            stressed_agg["expenses"] = stressed_exp
            stressed_agg["net_profit"] = stressed_np

            baseline_flat = _kpi_flat(agg_ifrs)
            stressed_flat = _kpi_flat(stressed_agg)
            labels = _kpi_labels()

            for key, base_val in baseline_flat.items():
                stressed_val = stressed_flat.get(key, 0)
                delta_pct = round((stressed_val - base_val) / base_val * 100, 1) if base_val != 0 else 0.0
                ws_stress_ifrs.append([
                    labels.get(key, key), round(base_val, 4), round(stressed_val, 4),
                    delta_pct, _stress_status(delta_pct),
                ])
                r = ws_stress_ifrs.max_row
                for ci in range(5):
                    style_data_cell(ws_stress_ifrs.cell(r, ci + 1), is_number=(ci in (1, 2, 3)))
                st_cell = ws_stress_ifrs.cell(r, 5)
                status_val = _stress_status(delta_pct)
                if status_val == "ok":
                    st_cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                elif status_val == "warn":
                    st_cell.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                elif status_val == "bad":
                    st_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    else:
        write_no_data(ws_stress_ifrs, RED_FILL)
    auto_width(ws_stress_ifrs)

    # ===================================================================
    # SHEET 15: Решения (Investment Decisions)
    # ===================================================================
    ws_dec = wb.create_sheet("Решения")
    ws_dec.sheet_properties.tabColor = "374151"
    write_sheet_header(ws_dec, "Инвестиционные решения", GRAY_FILL, company_info)

    ws_dec.append(["Актив", "Тип", "Статус", "Приоритет", "Категория", "Сумма", "Цена", "Итого", "Дата"])
    style_header_row(ws_dec, ws_dec.max_row, GRAY_FILL)

    try:
        from app.db.models.investment_decision import InvestmentDecision
        decisions = db.query(InvestmentDecision).filter(
            InvestmentDecision.created_by == current_user.id,
        ).order_by(InvestmentDecision.created_at.desc()).limit(100).all()
    except Exception:
        decisions = []

    if decisions:
        for d in decisions:
            total_val = (d.amount or 0) * (d.price or 0)
            ws_dec.append([
                d.asset_name or "",
                d.decision_type.value if d.decision_type else "",
                d.status.value if d.status else "",
                d.priority.value if d.priority else "",
                d.category.value if d.category else "",
                d.amount,
                d.price,
                round(total_val, 2) if total_val else None,
                d.created_at.strftime("%d.%m.%Y") if d.created_at else "",
            ])
            r = ws_dec.max_row
            for ci in range(9):
                style_data_cell(ws_dec.cell(r, ci + 1), is_number=(ci in (5, 6, 7)))
    else:
        write_no_data(ws_dec, GRAY_FILL)
    auto_width(ws_dec)

    # ===================================================================
    # SHEET 16: AI-анализ (AI-driven financial analysis)
    # ===================================================================
    TEAL_FILL = PatternFill(start_color="0E7C7B", end_color="0E7C7B", fill_type="solid")
    OK_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    WARN_FILL = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    BAD_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

    ws_ai = wb.create_sheet("AI-анализ")
    ws_ai.sheet_properties.tabColor = "0E7C7B"

    ws_ai.append(["AI-анализ финансового состояния"])
    ws_ai[ws_ai.max_row][0].font = TITLE_FONT
    org = company_info.get("name", "") if company_info else ""
    period = company_info.get("period", "") if company_info else ""
    ws_ai.append([f"Организация: {org}   |   Период: {period}"])
    ws_ai[ws_ai.max_row][0].font = BOLD
    ws_ai.append([])  # blank row 3

    if agg:
        ws_ai.append(["Показатель", "Статус", "Вывод и рекомендация"])
        style_header_row(ws_ai, ws_ai.max_row, TEAL_FILL)

        ai_conclusions, ai_overall = _build_ai_analysis(agg)
        status_labels = {"ok": "OK", "warn": "ВНИМАНИЕ", "bad": "КРИТИЧНО"}
        for category, status, text in ai_conclusions:
            ws_ai.append([category, status_labels.get(status, status), text])
            r = ws_ai.max_row
            style_data_cell(ws_ai.cell(r, 1), is_bold=True)
            style_data_cell(ws_ai.cell(r, 2), is_bold=True)
            style_data_cell(ws_ai.cell(r, 3))
            ws_ai.cell(r, 3).alignment = Alignment(wrap_text=True)
            # Color entire row based on status
            row_fill = {"ok": OK_FILL, "warn": WARN_FILL, "bad": BAD_FILL}.get(status, None)
            if row_fill:
                for ci in range(3):
                    ws_ai.cell(r, ci + 1).fill = row_fill

        # Blank row before overall
        ws_ai.append([])
        ws_ai.append([f"ОБЩАЯ ОЦЕНКА: {ai_overall}"])
        r = ws_ai.max_row
        ws_ai.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws_ai.cell(r, 1).font = Font(bold=True, size=12)
        ws_ai.cell(r, 1).alignment = Alignment(horizontal="center", wrap_text=True)
        # Color overall row
        ok_count = sum(1 for _, s, _ in ai_conclusions if s == "ok")
        score = ok_count / max(len(ai_conclusions), 1) * 100
        if score >= 70:
            ws_ai.cell(r, 1).fill = OK_FILL
        elif score >= 40:
            ws_ai.cell(r, 1).fill = WARN_FILL
        else:
            ws_ai.cell(r, 1).fill = BAD_FILL

        # Score row
        ws_ai.append([f"Показатели в норме: {ok_count} из {len(ai_conclusions)} ({score:.0f}%)"])
        r = ws_ai.max_row
        ws_ai.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws_ai.cell(r, 1).font = BOLD
        ws_ai.cell(r, 1).alignment = Alignment(horizontal="center")
    else:
        write_no_data(ws_ai, TEAL_FILL)

    ws_ai.column_dimensions["A"].width = 30
    ws_ai.column_dimensions["B"].width = 15
    ws_ai.column_dimensions["C"].width = 70
    auto_width(ws_ai)

    # ===================================================================
    # SHEET 17: Визуализации (Visualization data tables + AI explanations)
    # ===================================================================
    VIZ_BLUE = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    VIZ_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SECTION_FONT = Font(bold=True, size=13, color="FFFFFF")
    SECTION_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    AI_FONT = Font(italic=True, size=10, color="808080")
    PCT_FMT = '0.0%'

    ws_viz = wb.create_sheet("Визуализации")
    ws_viz.sheet_properties.tabColor = "2563EB"
    write_sheet_header(ws_viz, "Визуализации — Данные диаграмм и AI-пояснения", VIZ_BLUE, company_info)

    if agg:
        # Re-compute visualization data (same logic as GET /analytics/visualizations)
        viz_start = agg["total_assets_prev"]
        viz_revenue = agg["revenue"]
        viz_costs = -agg["expenses"]
        viz_inv_delta = -(agg["non_current_assets"] - agg.get("non_current_assets_prev", 0))
        viz_end = agg["total_assets"]
        viz_op_exp = viz_end - (viz_start + viz_revenue + viz_costs) - viz_inv_delta

        # ---- Section 1: Waterfall ----
        ws_viz.append(["Раздел 1: Каскадная диаграмма (Waterfall) — Движение баланса"])
        sec1_row = ws_viz.max_row
        ws_viz.merge_cells(start_row=sec1_row, start_column=1, end_row=sec1_row, end_column=3)
        ws_viz.cell(sec1_row, 1).font = SECTION_FONT
        ws_viz.cell(sec1_row, 1).fill = SECTION_FILL
        for ci in range(1, 4):
            ws_viz.cell(sec1_row, ci).fill = SECTION_FILL

        ws_viz.append(["Показатель", "Сумма (UZS)", "Тип"])
        style_header_row(ws_viz, ws_viz.max_row, VIZ_BLUE)

        wf_rows = [
            ("Начальный баланс", viz_start, "base"),
            ("+ Выручка", viz_revenue, "positive"),
            ("- Себестоимость", viz_costs, "negative"),
        ]
        if viz_op_exp < 0:
            wf_rows.append(("- Операционные расходы", viz_op_exp, "negative"))
        if abs(viz_inv_delta) > 0:
            sign = "+/-" if viz_inv_delta < 0 else "+"
            wf_rows.append((f"{sign} Инвестиции", viz_inv_delta, "negative" if viz_inv_delta < 0 else "positive"))
        wf_rows.append(("= Итоговый баланс", viz_end, "total"))

        for label, val, typ in wf_rows:
            is_total = typ in ("base", "total")
            ws_viz.append([label, val, typ])
            r = ws_viz.max_row
            style_data_cell(ws_viz.cell(r, 1), is_bold=is_total)
            style_data_cell(ws_viz.cell(r, 2), is_number=True, is_bold=is_total)
            style_data_cell(ws_viz.cell(r, 3))

        # AI explanation for waterfall
        bal_change_pct = round((viz_end - viz_start) / max(abs(viz_start), 1) * 100, 1)
        cost_rev_pct = round(agg["expenses"] / max(agg["revenue"], 1) * 100, 1)
        ws_viz.append([])
        ai_text_wf = (
            f"AI пояснение: Баланс {'вырос' if bal_change_pct >= 0 else 'снизился'} на {abs(bal_change_pct)}% за период. "
            f"Основной драйвер {'роста' if viz_revenue > 0 else 'изменения'} — выручка ({viz_revenue:,.0f} сум). "
            f"Себестоимость составила {cost_rev_pct}% от выручки."
        )
        ws_viz.append([ai_text_wf])
        r = ws_viz.max_row
        ws_viz.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws_viz.cell(r, 1).font = AI_FONT
        ws_viz.cell(r, 1).alignment = Alignment(wrap_text=True)

        # Blank separator
        ws_viz.append([])
        ws_viz.append([])

        # ---- Section 2: Tornado ----
        ws_viz.append(["Раздел 2: Диаграмма Торнадо — Анализ чувствительности"])
        sec2_row = ws_viz.max_row
        ws_viz.merge_cells(start_row=sec2_row, start_column=1, end_row=sec2_row, end_column=2)
        ws_viz.cell(sec2_row, 1).font = SECTION_FONT
        ws_viz.cell(sec2_row, 1).fill = SECTION_FILL
        for ci in range(1, 3):
            ws_viz.cell(sec2_row, ci).fill = SECTION_FILL

        ws_viz.append(["Фактор", "Влияние на прибыль (%)"])
        style_header_row(ws_viz, ws_viz.max_row, VIZ_BLUE)

        base_np = agg["net_profit"]
        tornado_factors = []
        cost_impact_val = agg["expenses"] * 0.10
        cost_impact_pct = round(cost_impact_val / max(abs(base_np), 1) * 100, 1)
        tornado_factors.append(("Себестоимость ±10%", cost_impact_pct))

        inv_impact_val = abs(viz_inv_delta) * 0.10
        inv_impact_pct = round(inv_impact_val / max(abs(base_np), 1) * 100, 1)
        tornado_factors.append(("Инвестиции ±10%", inv_impact_pct))

        fx_impact_val = agg["total_assets"] * 0.005
        fx_impact_pct = round(fx_impact_val / max(abs(base_np), 1) * 100, 1)
        tornado_factors.append(("Курсовые разницы ±10%", fx_impact_pct))

        interest_impact_val = agg["lt_liabilities"] * 0.01
        interest_impact_pct = round(interest_impact_val / max(abs(base_np), 1) * 100, 1)
        tornado_factors.append(("Процентные расходы ±10%", interest_impact_pct))

        # Sort by absolute impact descending
        tornado_factors.sort(key=lambda x: abs(x[1]), reverse=True)

        for factor_name, impact_pct in tornado_factors:
            ws_viz.append([factor_name, impact_pct / 100])
            r = ws_viz.max_row
            style_data_cell(ws_viz.cell(r, 1))
            style_data_cell(ws_viz.cell(r, 2), is_number=True)
            ws_viz.cell(r, 2).number_format = '±0.0%'

        # AI explanation for tornado
        top_factor = tornado_factors[0] if tornado_factors else ("—", 0)
        ws_viz.append([])
        ai_text_tn = (
            f"AI пояснение: Наибольшее влияние на прибыль оказывает {top_factor[0].split(' ±')[0]}. "
            f"При изменении на 10% прибыль изменится на ±{abs(top_factor[1]):.1f}%. "
            f"Рекомендуется хеджировать риск {top_factor[0].split(' ±')[0].lower()}."
        )
        ws_viz.append([ai_text_tn])
        r = ws_viz.max_row
        ws_viz.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws_viz.cell(r, 1).font = AI_FONT
        ws_viz.cell(r, 1).alignment = Alignment(wrap_text=True)

        ws_viz.append([])
        ws_viz.append([])

        # ---- Section 3: Bubble ----
        ws_viz.append(["Раздел 3: Пузырьковая диаграмма — Структура активов"])
        sec3_row = ws_viz.max_row
        ws_viz.merge_cells(start_row=sec3_row, start_column=1, end_row=sec3_row, end_column=4)
        ws_viz.cell(sec3_row, 1).font = SECTION_FONT
        ws_viz.cell(sec3_row, 1).fill = SECTION_FILL
        for ci in range(1, 5):
            ws_viz.cell(sec3_row, ci).fill = SECTION_FILL

        ws_viz.append(["Категория", "Стоимость (UZS)", "Доходность (%)", "Доля в активах (%)"])
        style_header_row(ws_viz, ws_viz.max_row, VIZ_BLUE)

        ta = agg["total_assets"]
        bubble_items = [
            ("Основные средства", agg["net_fa"], 5.0),
            ("Запасы", agg["inventories"], 12.0),
            ("Дебиторская задолженность", agg["receivables"], 8.0),
            ("Денежные средства", agg["cash"], 2.0),
        ]
        largest_name = ""
        largest_share = 0.0
        inv_share = 0.0

        for name, value, yield_pct in bubble_items:
            share = round(value / max(ta, 1) * 100, 1)
            if share > largest_share:
                largest_share = share
                largest_name = name
            if name == "Запасы":
                inv_share = share
            ws_viz.append([name, value, yield_pct / 100, share / 100])
            r = ws_viz.max_row
            style_data_cell(ws_viz.cell(r, 1))
            style_data_cell(ws_viz.cell(r, 2), is_number=True)
            ws_viz.cell(r, 3).number_format = '0.0%'
            ws_viz.cell(r, 3).border = THIN_BORDER
            ws_viz.cell(r, 4).number_format = '0.0%'
            ws_viz.cell(r, 4).border = THIN_BORDER

        ws_viz.append([])
        ai_text_bb = (
            f"AI пояснение: Наибольшую долю активов составляют {largest_name.lower()} ({largest_share:.1f}%). "
            f"Запасы занимают {inv_share:.1f}% — рассмотреть оптимизацию складских запасов."
        )
        ws_viz.append([ai_text_bb])
        r = ws_viz.max_row
        ws_viz.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws_viz.cell(r, 1).font = AI_FONT
        ws_viz.cell(r, 1).alignment = Alignment(wrap_text=True)

        ws_viz.append([])
        ws_viz.append([])

        # ---- Section 4: Heatmap ----
        ws_viz.append(["Раздел 4: Тепловая карта — Корреляция показателей"])
        sec4_row = ws_viz.max_row
        ws_viz.merge_cells(start_row=sec4_row, start_column=1, end_row=sec4_row, end_column=7)
        ws_viz.cell(sec4_row, 1).font = SECTION_FONT
        ws_viz.cell(sec4_row, 1).fill = SECTION_FILL
        for ci in range(1, 8):
            ws_viz.cell(sec4_row, ci).fill = SECTION_FILL

        hm_metrics = {
            "Выручка": agg["revenue"],
            "Себестоимость": agg["expenses"],
            "Чист.прибыль": base_np,
            "Активы": ta,
            "Капитал": agg["total_equity"],
            "Обязательства": agg["total_liabilities"],
        }
        hm_names = list(hm_metrics.keys())
        hm_vals = list(hm_metrics.values())
        n = len(hm_names)

        # Header row
        ws_viz.append([""] + hm_names)
        style_header_row(ws_viz, ws_viz.max_row, VIZ_BLUE)

        # Correlation matrix
        rev_np_corr = 0.0
        liab_eq_corr = 0.0
        for i in range(n):
            row_data = [hm_names[i]]
            for j in range(n):
                if i == j:
                    row_data.append(100.0)
                else:
                    vi, vj = hm_vals[i], hm_vals[j]
                    if vi != 0 and vj != 0:
                        ratio = min(abs(vi), abs(vj)) / max(abs(vi), abs(vj)) * 100
                        sign = 1 if (vi > 0) == (vj > 0) else -1
                        corr_val = round(ratio * sign, 1)
                    else:
                        corr_val = 0.0
                    row_data.append(corr_val)
                    # Track specific correlations for AI text
                    if hm_names[i] == "Выручка" and hm_names[j] == "Чист.прибыль":
                        rev_np_corr = corr_val
                    if hm_names[i] == "Обязательства" and hm_names[j] == "Капитал":
                        liab_eq_corr = corr_val
            ws_viz.append(row_data)
            r = ws_viz.max_row
            style_data_cell(ws_viz.cell(r, 1), is_bold=True)
            for ci in range(2, n + 2):
                cell = ws_viz.cell(r, ci)
                cell.number_format = '0.0'
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                # Color-code: green for strong positive, red for negative
                val = cell.value
                if isinstance(val, (int, float)):
                    if val >= 80:
                        cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                    elif val >= 50:
                        cell.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                    elif val < 0:
                        cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

        ws_viz.append([])
        ai_text_hm = (
            f"AI пояснение: Сильная положительная корреляция между выручкой и чистой прибылью ({rev_np_corr:.1f}%). "
            f"{'Отрицательная' if liab_eq_corr < 0 else 'Положительная'} корреляция между обязательствами и капиталом ({liab_eq_corr:.1f}%)."
        )
        ws_viz.append([ai_text_hm])
        r = ws_viz.max_row
        ws_viz.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws_viz.cell(r, 1).font = AI_FONT
        ws_viz.cell(r, 1).alignment = Alignment(wrap_text=True)

    else:
        write_no_data(ws_viz, VIZ_BLUE)

    ws_viz.column_dimensions["A"].width = 35
    ws_viz.column_dimensions["B"].width = 25
    ws_viz.column_dimensions["C"].width = 20
    ws_viz.column_dimensions["D"].width = 22
    ws_viz.column_dimensions["E"].width = 18
    ws_viz.column_dimensions["F"].width = 18
    ws_viz.column_dimensions["G"].width = 18
    auto_width(ws_viz)

    # ===================================================================
    # SHEET 18: Reconciliation НСБУ→МСФО
    # ===================================================================
    RECON_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    ws_recon = wb.create_sheet("Reconciliation НСБУ→МСФО")
    ws_recon.sheet_properties.tabColor = "1D4ED8"
    write_sheet_header(ws_recon, "Reconciliation НСБУ → МСФО", RECON_FILL, company_info)

    if agg:
        agg_ifrs_recon = _ifrs_adjusted_aggregates(agg)
        net_fa_recon = agg["net_fa"]
        ias16_reval_recon = round(net_fa_recon * 0.15, 2)
        ecl_recon = round(agg["receivables"] * 0.05, 2)
        deferred_tax_recon = round(ias16_reval_recon * 0.15, 2)
        rou_asset_recon = 950_000
        nsbu_net_profit = agg["net_profit"]
        # IFRS 16 effect: ROU depreciation vs lease expense
        ifrs16_effect = round(rou_asset_recon / 5 * -1 + rou_asset_recon * 0.18 * -1, 2)  # net negative (depr + interest)
        ifrs_net_profit = nsbu_net_profit - ecl_recon + ifrs16_effect

        # ---- Table 1: Reconciliation прибыли ----
        ws_recon.append(["Таблица 1: Reconciliation прибыли"])
        r = ws_recon.max_row
        ws_recon.cell(r, 1).font = Font(bold=True, size=12)
        ws_recon.append([])

        ws_recon.append(["Показатель", "Сумма", "Примечание"])
        style_header_row(ws_recon, ws_recon.max_row, RECON_FILL)

        recon_profit_rows = [
            ("Чистая прибыль НСБУ", nsbu_net_profit, "из ОПиУ"),
            ("Переоценка ОС (в OCI)", ias16_reval_recon, "IAS 16"),
            ("ECL резерв", -ecl_recon, "IFRS 9"),
            ("IFRS 16 эффект", ifrs16_effect, "Амортизация ROU vs аренда"),
            ("Отложенный налог", -deferred_tax_recon, "15% переоценки"),
            ("Чистая прибыль МСФО", ifrs_net_profit + ias16_reval_recon - deferred_tax_recon, ""),
        ]
        for label, val, note in recon_profit_rows:
            is_total = label.startswith("Чистая прибыль МСФО")
            ws_recon.append([label, val, note])
            r = ws_recon.max_row
            style_data_cell(ws_recon.cell(r, 1), is_bold=is_total)
            style_data_cell(ws_recon.cell(r, 2), is_number=True, is_bold=is_total)
            style_data_cell(ws_recon.cell(r, 3))
            if is_total:
                for ci in range(3):
                    ws_recon.cell(r, ci + 1).fill = LIGHT_BLUE

        ws_recon.append([])
        ws_recon.append([])

        # ---- Table 2: Reconciliation баланса ----
        ws_recon.append(["Таблица 2: Reconciliation баланса"])
        r = ws_recon.max_row
        ws_recon.cell(r, 1).font = Font(bold=True, size=12)
        ws_recon.append([])

        ws_recon.append(["Статья", "НСБУ", "Корректировка", "МСФО"])
        style_header_row(ws_recon, ws_recon.max_row, RECON_FILL)

        adj_assets = agg_ifrs_recon["total_assets"] - agg["total_assets"]
        adj_equity = agg_ifrs_recon["total_equity"] - agg["total_equity"]
        adj_liab = agg_ifrs_recon["total_liabilities"] - agg["total_liabilities"]

        recon_balance_rows = [
            ("Итого активов", agg["total_assets"], adj_assets, agg_ifrs_recon["total_assets"]),
            ("Итого капитал", agg["total_equity"], adj_equity, agg_ifrs_recon["total_equity"]),
            ("Итого обязательств", agg["total_liabilities"], adj_liab, agg_ifrs_recon["total_liabilities"]),
        ]
        for label, nsbu_val, adj_val, ifrs_val in recon_balance_rows:
            ws_recon.append([label, nsbu_val, adj_val, ifrs_val])
            r = ws_recon.max_row
            style_data_cell(ws_recon.cell(r, 1), is_bold=True)
            style_data_cell(ws_recon.cell(r, 2), is_number=True, is_bold=True)
            style_data_cell(ws_recon.cell(r, 3), is_number=True)
            style_data_cell(ws_recon.cell(r, 4), is_number=True, is_bold=True)
            # Color adjustment column
            adj_cell = ws_recon.cell(r, 3)
            if adj_val > 0:
                adj_cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            elif adj_val < 0:
                adj_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

        ws_recon.append([])
        ws_recon.append([])

        # ---- Table 3: Детализация корректировок с проводками ----
        ws_recon.append(["Таблица 3: Детализация корректировок с проводками"])
        r = ws_recon.max_row
        ws_recon.cell(r, 1).font = Font(bold=True, size=12)
        ws_recon.append([])

        ws_recon.append(["Корректировка", "Дт", "Кт", "Сумма", "Стандарт"])
        style_header_row(ws_recon, ws_recon.max_row, RECON_FILL)

        journal_entries = [
            ("Переоценка ОС", "Основные средства", "OCI / Добавочный капитал", ias16_reval_recon, "IAS 16"),
            ("ECL резерв", "Расход по ECL", "Резерв сомнит. долгов", ecl_recon, "IFRS 9"),
            ("Признание ROU", "ROU актив", "Обязательство по аренде", rou_asset_recon, "IFRS 16"),
            ("ОНО переоценка", "OCI", "Отложенное налог. обяз.", deferred_tax_recon, "IAS 12"),
        ]
        for adj_name, dt, kt, amount, standard in journal_entries:
            ws_recon.append([adj_name, dt, kt, amount, standard])
            r = ws_recon.max_row
            style_data_cell(ws_recon.cell(r, 1), is_bold=True)
            style_data_cell(ws_recon.cell(r, 2))
            style_data_cell(ws_recon.cell(r, 3))
            style_data_cell(ws_recon.cell(r, 4), is_number=True)
            style_data_cell(ws_recon.cell(r, 5))

    else:
        write_no_data(ws_recon, RECON_FILL)

    ws_recon.column_dimensions["A"].width = 35
    ws_recon.column_dimensions["B"].width = 30
    ws_recon.column_dimensions["C"].width = 30
    ws_recon.column_dimensions["D"].width = 18
    ws_recon.column_dimensions["E"].width = 15
    auto_width(ws_recon)

    # --- Save and return ---
    try:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
    except Exception as e:
        logger.error("Failed to save workbook: %s\n%s", e, traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Ошибка сохранения Excel: {str(e)}")

    org_name = company_info.get("name", "portfolio") if company_info else "portfolio"
    safe_name = "".join(c for c in org_name if c.isascii() and (c.isalnum() or c in " _-")).strip()[:30] or "report"
    filename = f"report_{safe_name}_{req.portfolio_id}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# GET /analytics/reconciliation — НСБУ → МСФО reconciliation
# ---------------------------------------------------------------------------

@router.get("/analytics/reconciliation")
async def get_reconciliation(
    current_user: User = Depends(get_current_user),
):
    """Reconciliation of NSBU to IFRS: profit bridge + balance bridge."""
    agg = _get_balance_aggregates(user_id=current_user.id)
    if not agg:
        return JSONResponse({"profit_rows": [], "balance_rows": []})

    cache = _user_cache(current_user.id)

    net_fa = agg["net_fa"]
    receivables = agg["receivables"]
    nsbu_net_profit = agg["net_profit"]

    # IFRS adjustments (same logic as _ifrs_adjusted_aggregates)
    ias16_reval = round(net_fa * 0.15, 2)         # IAS 16 revaluation surplus
    ecl = round(receivables * 0.05, 2)            # IFRS 9 expected credit loss
    deferred_tax = round(ias16_reval * 0.15, 2)   # Deferred tax on revaluation (15%)

    # IFRS 16 lease effect: ROU depreciation vs operating lease payments
    # Approximate from account 7800 (lease liability)
    accounts = cache.get("accounts", {})
    lease_7800 = 0.0
    acc_7800 = accounts.get("7800")
    if acc_7800:
        lease_7800 = float(acc_7800.get("current", 0))

    # IFRS 16 effect on P&L: depreciation of ROU + interest - operating lease payment
    # ROU depreciation ~= lease_liability / remaining_term (approx 3.5 yrs)
    # Interest ~= lease_liability * discount_rate (approx 10%)
    # Operating lease payment under NSBU ~= lease_liability / 3
    ifrs16_depr = round(lease_7800 / 3.5, 2) if lease_7800 else 0.0
    ifrs16_interest = round(lease_7800 * 0.10, 2) if lease_7800 else 0.0
    ifrs16_nsbu_expense = round(lease_7800 / 3.0, 2) if lease_7800 else 0.0
    ifrs16_pnl_effect = round(ifrs16_nsbu_expense - ifrs16_depr - ifrs16_interest, 2)

    ifrs_net_profit = round(nsbu_net_profit + ias16_reval - ecl + ifrs16_pnl_effect - deferred_tax, 2)

    # ROU asset for balance
    rou_asset = round(lease_7800 * 3.5, 2) if lease_7800 else 0.0

    # --- Profit reconciliation ---
    profit_rows = [
        {
            "label": "Чистая прибыль НСБУ",
            "nsbu": nsbu_net_profit,
            "adjustment": None,
            "ifrs": None,
            "note": "из ОПиУ НСБУ",
        },
        {
            "label": "+ Переоценка ОС (OCI)",
            "nsbu": None,
            "adjustment": ias16_reval,
            "ifrs": None,
            "note": "IAS 16, в OCI",
        },
        {
            "label": "- ECL резерв (IFRS 9)",
            "nsbu": None,
            "adjustment": -ecl,
            "ifrs": None,
            "note": "уменьшает прибыль",
        },
        {
            "label": "+/- IFRS 16 эффект",
            "nsbu": None,
            "adjustment": ifrs16_pnl_effect,
            "ifrs": None,
            "note": "амортизация ROU vs арендные платежи",
        },
        {
            "label": "- Отложенный налог",
            "nsbu": None,
            "adjustment": -deferred_tax,
            "ifrs": None,
            "note": "15% от переоценки",
        },
        {
            "label": "= Чистая прибыль МСФО",
            "nsbu": None,
            "adjustment": None,
            "ifrs": ifrs_net_profit,
            "note": "",
            "isTotal": True,
        },
    ]

    # --- Balance reconciliation ---
    nsbu_total_assets = agg["total_assets"]
    nsbu_total_equity = agg["total_equity"]
    nsbu_total_liabilities = agg["total_liabilities"]

    asset_adj = round(ias16_reval - ecl + rou_asset, 2)
    equity_adj = round(ias16_reval - ecl - deferred_tax, 2)
    liabilities_adj = round(rou_asset + deferred_tax, 2)

    ifrs_total_assets = round(nsbu_total_assets + asset_adj, 2)
    ifrs_total_equity = round(nsbu_total_equity + equity_adj, 2)
    ifrs_total_liabilities = round(nsbu_total_liabilities + liabilities_adj, 2)

    balance_rows = [
        {
            "label": "Итого активов",
            "nsbu": nsbu_total_assets,
            "adjustment": asset_adj,
            "ifrs": ifrs_total_assets,
            "detail": f"+{ias16_reval:,.0f} (IAS 16) -{ecl:,.0f} (ECL) +{rou_asset:,.0f} (ROU)",
        },
        {
            "label": "Итого капитал",
            "nsbu": nsbu_total_equity,
            "adjustment": equity_adj,
            "ifrs": ifrs_total_equity,
            "detail": f"+{ias16_reval:,.0f} (OCI) -{ecl:,.0f} (ECL) -{deferred_tax:,.0f} (DT)",
        },
        {
            "label": "Итого обязательств",
            "nsbu": nsbu_total_liabilities,
            "adjustment": liabilities_adj,
            "ifrs": ifrs_total_liabilities,
            "detail": f"+{rou_asset:,.0f} (IFRS 16) +{deferred_tax:,.0f} (DT)",
        },
    ]

    return JSONResponse({
        "profit_rows": profit_rows,
        "balance_rows": balance_rows,
        "adjustments_summary": {
            "ias16_revaluation": ias16_reval,
            "ifrs9_ecl": ecl,
            "ifrs16_rou_asset": rou_asset,
            "ifrs16_pnl_effect": ifrs16_pnl_effect,
            "deferred_tax": deferred_tax,
        },
    })
