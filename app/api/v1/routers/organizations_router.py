"""
API Router: Organizations, Balance Entries, Import (1C/Excel/Manual)
TZ#2 P0-P1 Implementation
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
import io
from app.api.v1.deps import get_db
from sqlalchemy.orm import Session
from sqlalchemy import select, func, and_
from pydantic import BaseModel, Field
from typing import Optional, List
from datetime import date, datetime
from decimal import Decimal
import json

router = APIRouter(tags=["organizations"])


# ── Pydantic Schemas ──────────────────────────────────────

class OrganizationCreate(BaseModel):
    name: str = Field(..., max_length=500)
    inn: Optional[str] = Field(None, max_length=9)
    ownership_form: Optional[str] = None
    oked: Optional[str] = None
    registration_date: Optional[date] = None
    director: Optional[str] = None
    charter_capital: Optional[float] = None
    charter_currency: str = "UZS"
    address: Optional[str] = None
    mode: str = "solo"
    accounting_currency: str = "UZS"
    parent_id: Optional[int] = None
    ownership_share: float = 100.0


class OrganizationResponse(BaseModel):
    id: int
    name: str
    inn: Optional[str]
    ownership_form: Optional[str]
    oked: Optional[str]
    mode: str
    accounting_currency: str
    parent_id: Optional[int]
    is_active: bool
    created_at: datetime

    class Config:
        from_attributes = True


class BalanceEntryCreate(BaseModel):
    account_code: str
    period_date: date
    debit: float = 0
    credit: float = 0
    balance: float = 0
    currency: str = "UZS"
    description: Optional[str] = None


class BalanceEntryResponse(BaseModel):
    id: int
    organization_id: int
    account_code: str
    account_name: str
    category: str
    period_date: date
    debit: float
    credit: float
    balance: float
    currency: str
    source: str

    class Config:
        from_attributes = True


class BalanceBatchCreate(BaseModel):
    organization_id: int
    period_date: date
    entries: List[BalanceEntryCreate]


class BalanceSummary(BaseModel):
    total_assets: float
    long_term_assets: float
    current_assets: float
    total_liabilities: float
    total_equity: float
    balance_check: bool
    period_date: date


class ChartAccountResponse(BaseModel):
    id: int
    code: str
    name_ru: str
    name_uz: Optional[str]
    category: str
    level: int
    parent_code: Optional[str]

    class Config:
        from_attributes = True


# ── Organizations CRUD ────────────────────────────────────

@router.post("/organizations", response_model=OrganizationResponse)
async def create_organization(data: OrganizationCreate, db: Session = Depends(get_db)):
    """Create organization (Solo/Branch/Holding)"""
    from app.db.models.organization_models import Organization

    org = Organization(
        user_id=1,  # TODO: get from JWT
        name=data.name,
        inn=data.inn,
        ownership_form=data.ownership_form,
        oked=data.oked,
        registration_date=data.registration_date,
        director=data.director,
        charter_capital=data.charter_capital,
        charter_currency=data.charter_currency,
        address=data.address,
        mode=data.mode,
        accounting_currency=data.accounting_currency,
        parent_id=data.parent_id,
        ownership_share=data.ownership_share,
    )
    db.add(org)
    db.commit()
    db.refresh(org)
    return org


@router.get("/organizations", response_model=List[OrganizationResponse])
async def list_organizations(
    mode: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """List all organizations for current user"""
    from app.db.models.organization_models import Organization

    query = select(Organization).where(Organization.user_id == 1, Organization.is_active == True)
    if mode:
        query = query.where(Organization.mode == mode)
    query = query.order_by(Organization.created_at.desc())

    result = db.execute(query)
    return result.scalars().all()


@router.get("/organizations/{org_id}", response_model=OrganizationResponse)
async def get_organization(org_id: int, db: Session = Depends(get_db)):
    """Get organization by ID with hierarchy"""
    from app.db.models.organization_models import Organization

    org = db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    return org


@router.get("/organizations/{org_id}/children", response_model=List[OrganizationResponse])
async def get_org_children(org_id: int, db: Session = Depends(get_db)):
    """Get child organizations (branches/subsidiaries)"""
    from app.db.models.organization_models import Organization

    result = db.execute(
        select(Organization).where(Organization.parent_id == org_id, Organization.is_active == True)
    )
    return result.scalars().all()


# ── Chart of Accounts ─────────────────────────────────────

@router.get("/chart-of-accounts", response_model=List[ChartAccountResponse])
async def list_accounts(
    category: Optional[str] = None,
    level: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Get NSBU chart of accounts"""
    from app.db.models.organization_models import ChartOfAccounts

    query = select(ChartOfAccounts).where(ChartOfAccounts.is_active == True)
    if category:
        query = query.where(ChartOfAccounts.category == category)
    if level:
        query = query.where(ChartOfAccounts.level == level)
    query = query.order_by(ChartOfAccounts.code)

    result = db.execute(query)
    return result.scalars().all()


# ── Balance Entries ───────────────────────────────────────

@router.post("/organizations/{org_id}/balance")
async def create_balance_entries(
    org_id: int,
    data: BalanceBatchCreate,
    db: Session = Depends(get_db)
):
    """Batch create balance entries for organization"""
    from app.db.models.organization_models import Organization, ChartOfAccounts, BalanceEntry

    org = db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    created = []
    errors = []

    for entry in data.entries:
        account = db.execute(
            select(ChartOfAccounts).where(ChartOfAccounts.code == entry.account_code)
        ).scalar()

        if not account:
            errors.append(f"Account {entry.account_code} not found")
            continue

        be = BalanceEntry(
            organization_id=org_id,
            account_id=account.id,
            period_date=data.period_date,
            debit=entry.debit,
            credit=entry.credit,
            balance=entry.balance,
            currency=entry.currency,
            description=entry.description,
            source="manual",
        )
        db.add(be)
        created.append(entry.account_code)

    db.commit()

    return {
        "status": "ok",
        "created": len(created),
        "errors": errors,
        "period_date": data.period_date.isoformat()
    }


@router.get("/organizations/{org_id}/balance", response_model=List[BalanceEntryResponse])
async def get_balance(
    org_id: int,
    period_date: Optional[date] = None,
    category: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get balance entries for organization"""
    from app.db.models.organization_models import BalanceEntry, ChartOfAccounts

    query = (
        select(
            BalanceEntry.id,
            BalanceEntry.organization_id,
            ChartOfAccounts.code.label("account_code"),
            ChartOfAccounts.name_ru.label("account_name"),
            ChartOfAccounts.category,
            BalanceEntry.period_date,
            BalanceEntry.debit,
            BalanceEntry.credit,
            BalanceEntry.balance,
            BalanceEntry.currency,
            BalanceEntry.source,
        )
        .join(ChartOfAccounts, BalanceEntry.account_id == ChartOfAccounts.id)
        .where(BalanceEntry.organization_id == org_id)
    )

    if period_date:
        query = query.where(BalanceEntry.period_date == period_date)
    if category:
        query = query.where(ChartOfAccounts.category == category)

    query = query.order_by(ChartOfAccounts.code)
    result = db.execute(query)
    return result.all()


@router.get("/organizations/{org_id}/balance/summary", response_model=BalanceSummary)
async def get_balance_summary(
    org_id: int,
    period_date: date = Query(...),
    db: Session = Depends(get_db)
):
    """Get balance summary (total assets, liabilities, equity)"""
    from app.db.models.organization_models import BalanceEntry, ChartOfAccounts

    def sum_category(cat):
        result = db.execute(
            select(func.coalesce(func.sum(BalanceEntry.balance), 0))
            .join(ChartOfAccounts, BalanceEntry.account_id == ChartOfAccounts.id)
            .where(
                BalanceEntry.organization_id == org_id,
                BalanceEntry.period_date == period_date,
                ChartOfAccounts.category == cat
            )
        )
        return float(result.scalar() or 0)

    lta = sum_category("long_term_assets")
    ca = sum_category("current_assets")
    liab = sum_category("liabilities")
    eq = sum_category("equity")
    total_assets = lta + ca
    total_liabilities = liab + eq

    return BalanceSummary(
        total_assets=total_assets,
        long_term_assets=lta,
        current_assets=ca,
        total_liabilities=liab,
        total_equity=eq,
        balance_check=abs(total_assets - total_liabilities) < 0.01,
        period_date=period_date,
    )


# ── Import: Excel ─────────────────────────────────────────

@router.post("/organizations/{org_id}/import/excel")
async def import_excel(
    org_id: int,
    file: UploadFile = File(...),
    period_date: date = Query(...),
    db: Session = Depends(get_db)
):
    """Import balance from Excel/CSV (ОСВ format)"""
    from app.db.models.organization_models import Organization, ChartOfAccounts, BalanceEntry, ImportSession
    import io

    org = db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    session = ImportSession(
        organization_id=org_id,
        user_id=1,
        source_type="excel",
        filename=file.filename,
        status="processing",
    )
    db.add(session)
    db.commit()
    db.refresh(session)

    try:
        content = await file.read()

        if file.filename.endswith('.csv'):
            import csv
            reader = csv.DictReader(io.StringIO(content.decode('utf-8')))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

        imported = 0
        failed = 0
        errors = []

        for row in rows:
            code = str(row.get("code", row.get("Счёт", row.get("account", "")))).strip()
            if not code:
                continue

            account = db.execute(
                select(ChartOfAccounts).where(ChartOfAccounts.code == code)
            ).scalar()

            if not account:
                failed += 1
                errors.append(f"Account {code} not found in chart")
                continue

            debit = float(row.get("debit", row.get("Дебет", row.get("дебет", 0))) or 0)
            credit = float(row.get("credit", row.get("Кредит", row.get("кредит", 0))) or 0)
            balance = float(row.get("balance", row.get("Сальдо", row.get("сальдо", 0))) or 0)

            if balance == 0:
                balance = debit - credit

            be = BalanceEntry(
                organization_id=org_id,
                account_id=account.id,
                period_date=period_date,
                debit=debit,
                credit=credit,
                balance=balance,
                currency=org.accounting_currency,
                source="excel",
                import_session_id=session.id,
            )
            db.add(be)
            imported += 1

        session.status = "completed"
        session.records_total = len(rows)
        session.records_imported = imported
        session.records_failed = failed
        session.error_log = errors if errors else None
        session.completed_at = datetime.utcnow()
        db.commit()

        return {
            "status": "completed",
            "session_id": session.id,
            "total": len(rows),
            "imported": imported,
            "failed": failed,
            "errors": errors[:10],
        }

    except Exception as e:
        session.status = "failed"
        session.error_log = [str(e)]
        db.commit()
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")


@router.post("/organizations/{org_id}/import/manual")
async def import_manual(
    org_id: int,
    data: BalanceBatchCreate,
    db: Session = Depends(get_db)
):
    """Manual balance entry (Solo wizard)"""
    return await create_balance_entries(org_id, data, db)


# ── Import: 1C OData ─────────────────────────────────────

class ODataConnectionConfig(BaseModel):
    base_url: str = Field(..., description="URL 1C OData, e.g. http://server/base/odata/standard.odata")
    username: str
    password: str
    period_date: date


@router.post("/organizations/{org_id}/import/1c-odata")
async def import_1c_odata(
    org_id: int,
    config: ODataConnectionConfig,
    db: Session = Depends(get_db)
):
    """Import balance from 1C via OData REST API"""
    from app.db.models.organization_models import Organization, ChartOfAccounts, BalanceEntry, ImportSession
    import httpx

    org = db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    session = ImportSession(
        organization_id=org_id,
        user_id=1,
        source_type="1c_odata",
        filename=config.base_url,
        status="processing",
    )
    db.add(session)
    db.commit()
    db.refresh(session)

    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(
                f"{config.base_url}/AccountingRegister_Хозрасчетный/Balance"
                f"?$format=json"
                f"&Период={config.period_date.isoformat()}T00:00:00",
                auth=(config.username, config.password),
            )

            if response.status_code != 200:
                raise HTTPException(status_code=502, detail=f"1C returned {response.status_code}")

            data = response.json()
            records = data.get("value", [])

        imported = 0
        failed = 0
        errors = []

        for record in records:
            code = str(record.get("Account_Code", "")).strip()[:4]
            if not code:
                continue

            account = db.execute(
                select(ChartOfAccounts).where(ChartOfAccounts.code == code)
            ).scalar()

            if not account:
                failed += 1
                errors.append(f"1C account {code} not mapped")
                continue

            be = BalanceEntry(
                organization_id=org_id,
                account_id=account.id,
                period_date=config.period_date,
                debit=float(record.get("ДебетовоеСальдо", record.get("AmountDr", 0)) or 0),
                credit=float(record.get("КредитовоеСальдо", record.get("AmountCr", 0)) or 0),
                balance=float(record.get("СальдоКонечное", record.get("Balance", 0)) or 0),
                currency=org.accounting_currency,
                source="1c_odata",
                import_session_id=session.id,
            )
            db.add(be)
            imported += 1

        session.status = "completed"
        session.records_total = len(records)
        session.records_imported = imported
        session.records_failed = failed
        session.error_log = errors if errors else None
        session.completed_at = datetime.utcnow()
        db.commit()

        return {
            "status": "completed",
            "session_id": session.id,
            "total": len(records),
            "imported": imported,
            "failed": failed,
            "errors": errors[:10],
        }

    except httpx.RequestError as e:
        session.status = "failed"
        session.error_log = [f"Connection error: {str(e)}"]
        db.commit()
        raise HTTPException(status_code=502, detail=f"Cannot connect to 1C: {str(e)}")


@router.get("/organizations/{org_id}/import/history")
async def import_history(org_id: int, db: Session = Depends(get_db)):
    """Get import history for organization"""
    from app.db.models.organization_models import ImportSession

    result = db.execute(
        select(ImportSession)
        .where(ImportSession.organization_id == org_id)
        .order_by(ImportSession.created_at.desc())
        .limit(50)
    )
    return result.scalars().all()


# -- Export: Excel balance -------
@router.get("/organizations/{org_id}/export/excel")
async def export_balance_excel(
    org_id: int,
    period_date: date = Query(...),
    db: Session = Depends(get_db)
):
    """Export balance to Excel (.xlsx)"""
    from app.db.models.organization_models import (
        Organization, BalanceEntry, ChartOfAccounts
    )
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    
    org = db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    entries = db.execute(
        select(
            ChartOfAccounts.code,
            ChartOfAccounts.name_ru,
            ChartOfAccounts.category,
            BalanceEntry.debit,
            BalanceEntry.credit,
            BalanceEntry.balance,
        )
        .join(ChartOfAccounts, BalanceEntry.account_id == ChartOfAccounts.id)
        .where(BalanceEntry.organization_id == org_id,
               BalanceEntry.period_date == period_date)
        .order_by(ChartOfAccounts.code)
    ).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance"
    bold = Font(bold=True)
    thin = Side(style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Header
    ws.append([f"Organization: {org.name}"])
    ws.append([f"Period: {period_date.isoformat()}"])
    ws.append([])
    headers = ["Code", "Account", "Category", "Debit", "Credit", "Balance"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = bold
        cell.border = border

    for e in entries:
        ws.append([e.code, e.name_ru, e.category, e.debit, e.credit, e.balance])
        
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_org = "".join(c for c in org.name if c.isascii() and (c.isalnum() or c in " _-")).strip()[:30] or "org"
    fname = f"balance_{safe_org}_{period_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
    

# -- Export: PDF balance -------
@router.get("/organizations/{org_id}/export/pdf")
async def export_balance_pdf(
    org_id: int,
    period_date: date = Query(...),
    db: Session = Depends(get_db)
):
    """Export balance to PDF"""
    from app.db.models.organization_models import (
        Organization, BalanceEntry, ChartOfAccounts
    )

    org = db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
        
    entries = db.execute(
        select(
            ChartOfAccounts.code,
            ChartOfAccounts.name_ru,
            ChartOfAccounts.category,
            BalanceEntry.debit,
            BalanceEntry.credit,
            BalanceEntry.balance,
        )
        .join(ChartOfAccounts, BalanceEntry.account_id == ChartOfAccounts.id)
        .where(BalanceEntry.organization_id == org_id,
               BalanceEntry.period_date == period_date)
        .order_by(ChartOfAccounts.code)
    ).all()
    
    # Build HTML table for PDF
    rows_html = ""
    for e in entries:
        rows_html += f"<tr><td>{e.code}</td><td>{e.name_ru}</td>"
        rows_html += f"<td>{e.debit:,.2f}</td><td>{e.credit:,.2f}</td>"
        rows_html += f"<td>{e.balance:,.2f}</td></tr>"

    html = f"""
    <html><head><meta charset="utf-8">
    <style>
        body {{ font-family: Arial; font-size: 12px; }}
        h1 {{ font-size: 16px; }}
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #333; padding: 4px 8px; }}
        th {{ background: #eee; }}
        .right {{ text-align: right; }}
    </style></head><body>
        <h1>Balance: {org.name}</h1>
    <p>Period: {period_date}</p>
    <table>
    <tr><th>Code</th><th>Account</th><th>Debit</th><th>Credit</th><th>Balance</th></tr>
    {rows_html}
    </table></body></html>
    """

    # Return as HTML (browser can print to PDF)
    safe_org = "".join(c for c in org.name if c.isascii() and (c.isalnum() or c in " _-")).strip()[:30] or "org"
    return StreamingResponse(
        io.BytesIO(html.encode("utf-8")),
        media_type="text/html",
        headers={"Content-Disposition":
                 f'attachment; filename="balance_{safe_org}_{period_date}.html"'},
    )
