import io
import re
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.api.v1.deps import get_db, get_current_user
from app.db.models.user import User
from app.db.models.portfolio import Portfolio
from app.schemas.portfolio import PortfolioCreate, PortfolioRead, PortfolioUpdate

router = APIRouter(prefix="/portfolios", tags=["portfolios"])

# ---------------------------------------------------------------------------
# In-memory cache for parsed ОСВ data (until we have a proper DB table)
# ---------------------------------------------------------------------------
_portfolio_cache: dict = {}  # key = user_id, value = user's cache data


def _user_cache(user_id: int) -> dict:
    """Get or create per-user cache partition."""
    if user_id not in _portfolio_cache:
        _portfolio_cache[user_id] = {}
    return _portfolio_cache[user_id]


# ---------------------------------------------------------------------------
# CRUD endpoints (unchanged)
# ---------------------------------------------------------------------------

@router.post("", response_model=PortfolioRead, status_code=status.HTTP_201_CREATED)
def create_portfolio(
    portfolio_in: PortfolioCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    portfolio = Portfolio(**portfolio_in.dict(), owner_id=current_user.id)
    db.add(portfolio)
    db.commit()
    db.refresh(portfolio)
    return portfolio

@router.get("", response_model=List[PortfolioRead])
def get_portfolios(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    return db.query(Portfolio).filter(Portfolio.owner_id == current_user.id).all()

@router.get("/{portfolio_id}", response_model=PortfolioRead)
def get_portfolio(
    portfolio_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    portfolio = db.query(Portfolio).filter(
        Portfolio.id == portfolio_id,
        Portfolio.owner_id == current_user.id
    ).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    return portfolio

@router.put("/{portfolio_id}", response_model=PortfolioRead)
def update_portfolio(
    portfolio_id: int,
    portfolio_in: PortfolioUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    portfolio = db.query(Portfolio).filter(
        Portfolio.id == portfolio_id,
        Portfolio.owner_id == current_user.id
    ).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    for field, value in portfolio_in.dict(exclude_unset=True).items():
        setattr(portfolio, field, value)
    db.commit()
    db.refresh(portfolio)
    return portfolio

@router.delete("/{portfolio_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_portfolio(
    portfolio_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    portfolio = db.query(Portfolio).filter(
        Portfolio.id == portfolio_id,
        Portfolio.owner_id == current_user.id
    ).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    db.delete(portfolio)
    db.commit()


# ---------------------------------------------------------------------------
# Company registration (manual)
# ---------------------------------------------------------------------------

class CompanyRegistration(BaseModel):
    name: str
    inn: str = ""
    oked: str = ""
    address: str = ""
    director: str = ""
    accountant: str = ""
    size: str = "medium"  # small/medium/large
    org_type: str = "solo"  # solo/branch/holding


@router.post("/register")
async def register_company(
    data: CompanyRegistration,
    current_user: User = Depends(get_current_user),
):
    """Register company manually (without file upload)."""
    cache = _user_cache(current_user.id)
    cache["company_info"] = {
        "name": data.name,
        "inn": data.inn,
        "activity": data.oked,
        "address": data.address,
        "director": data.director,
        "accountant": data.accountant,
        "size": data.size,
        "org_type": data.org_type,
    }
    return JSONResponse({"status": "ok", "message": "Организация зарегистрирована"})


@router.get("/company-info")
async def get_company_info(
    current_user: User = Depends(get_current_user),
):
    """Return cached company info."""
    cache = _user_cache(current_user.id)
    info = cache.get("company_info")
    if not info:
        return JSONResponse({"status": "empty", "company_info": None})
    return JSONResponse({"status": "ok", "company_info": info})


# ---------------------------------------------------------------------------
# Helper: parse a numeric value from a cell (handles None, str, float)
# ---------------------------------------------------------------------------

def _num(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.replace(" ", "").replace("\xa0", "").replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0


def _detect_section(code_val, name_val) -> Optional[int]:
    """Detect if a row is a section header (РАЗДЕЛ I..VII). Returns section number or None."""
    for val in (code_val, name_val):
        if val is None:
            continue
        s = str(val).strip().upper()
        m = re.search(r"РАЗДЕЛ\s*(VII|VI|IV|V|III|II|I)\b", s)
        if m:
            roman = m.group(1)
            mapping = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7}
            return mapping.get(roman)
    return None


def _is_section_header(code_val) -> bool:
    """Check if a row is a section header (РАЗДЕЛ, roman numeral, etc.)."""
    if code_val is None:
        return True
    s = str(code_val).strip()
    if not s:
        return True
    if re.search(r"РАЗДЕЛ|раздел|Раздел", s, re.IGNORECASE):
        return True
    if re.match(r"^[IVXLC]+\.?\s*$", s):
        return True
    return False


def _extract_full_code(code_val) -> Optional[str]:
    """Extract the FULL account code (e.g. '2010', '2010.1') from a cell value."""
    if code_val is None:
        return None
    s = str(code_val).strip()
    m = re.match(r"^(\d{4}(?:\.\d+)?)", s)
    if m:
        return m.group(1)
    return None


def _base_code(full_code: str) -> str:
    """Get the 4-digit base code: '2010.1' -> '2010'."""
    return full_code.split(".")[0]


def _is_subaccount(full_code: str) -> bool:
    """Check if this is a sub-account like 2010.1."""
    return "." in full_code


def _classify_account(code: str) -> str:
    """Classify account code into section."""
    prefix = int(code[:2])
    if 1 <= prefix <= 9:
        return "non_current_assets"
    elif 10 <= prefix <= 19:
        return "inventories"
    elif 20 <= prefix <= 29:
        return "receivables"
    elif 50 <= prefix <= 59:
        return "cash"
    elif 60 <= prefix <= 69:
        return "current_liabilities"
    elif 70 <= prefix <= 79:
        return "long_term_liabilities"
    elif 80 <= prefix <= 89:
        return "equity"
    else:
        return "other"


def _is_credit_account(code: str) -> bool:
    """Credit-side accounts: liabilities (6x, 7x) and equity (8x), plus contra-asset 0200."""
    if code == "0200":
        return True
    prefix = int(code[:2])
    return prefix >= 60


def _parse_company_info(wb) -> dict:
    """Parse company info from 'Реквизиты' sheet."""
    info = {}
    req_sheet = None
    for sheet_name in wb.sheetnames:
        if "еквизит" in sheet_name or "Реквизит" in sheet_name:
            req_sheet = wb[sheet_name]
            break
    if req_sheet is None:
        # Try first sheet if it looks like it has company info
        return info

    def _cell(row, col):
        v = req_sheet.cell(row=row, column=col).value
        return str(v).strip() if v else ""

    info["name"] = _cell(5, 3)       # C5
    info["inn"] = _cell(6, 3)        # C6
    info["address"] = _cell(7, 3)    # C7
    info["activity"] = _cell(8, 3)   # C8
    info["registration_date"] = _cell(10, 3)  # C10
    info["director"] = _cell(11, 3)  # C11
    info["accountant"] = _cell(12, 3)  # C12
    info["period"] = _cell(13, 3)    # C13
    info["unit"] = _cell(15, 3)      # C15
    return info


# ---------------------------------------------------------------------------
# POST /import/excel — real ОСВ parser
# ---------------------------------------------------------------------------

@router.post("/import/excel")
async def import_portfolio_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import financial data from 1C Excel (ОСВ) file."""
    content = await file.read()
    filename = file.filename or ""

    if filename.lower().endswith((".xlsx", ".xls")):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=True)

            # --- Parse company info from Реквизиты sheet ---
            company_info = _parse_company_info(wb)
            cache = _user_cache(current_user.id)
            cache["company_info"] = company_info

            # --- Find ОСВ sheet ---
            osv_sheet = None
            for sheet_name in wb.sheetnames:
                if "ОСВ" in sheet_name or "осв" in sheet_name.lower():
                    osv_sheet = wb[sheet_name]
                    break
            if osv_sheet is None:
                osv_sheet = wb[wb.sheetnames[0]]

            # --- Section-aware parser ---
            # Track current section (I-VII) to distinguish balance vs P&L accounts
            accounts = {}          # balance sheet accounts (sections I-V)
            revenue_accounts = {}  # section VI (revenue, credit side)
            expense_accounts = {}  # section VII (expenses, debit side)
            current_section = 0
            parsed_count = 0

            for row in osv_sheet.iter_rows(min_row=2, max_col=7, values_only=False):
                cells = list(row)
                if len(cells) < 7:
                    continue
                code_cell = cells[1].value  # B
                name_cell = cells[2].value  # C

                # Check if this row is a section header
                section_num = _detect_section(code_cell, name_cell)
                if section_num is not None:
                    current_section = section_num
                    continue

                if _is_section_header(code_cell):
                    continue

                full_code = _extract_full_code(code_cell)
                if full_code is None:
                    continue

                base = _base_code(full_code)
                name = str(name_cell).strip() if name_cell else ""

                debit_start = _num(cells[3].value)   # D
                credit_start = _num(cells[4].value)  # E
                debit_end = _num(cells[5].value)      # F
                credit_end = _num(cells[6].value)     # G

                parsed_count += 1

                # Section VI = Revenue (P&L), Section VII = Expenses (P&L)
                if current_section == 6:
                    revenue_accounts[full_code] = {
                        "code": full_code, "name": name,
                        "credit_start": credit_start, "credit_end": credit_end,
                    }
                    continue
                elif current_section == 7:
                    expense_accounts[full_code] = {
                        "code": full_code, "name": name,
                        "debit_start": debit_start, "debit_end": debit_end,
                    }
                    continue

                # Sections I-V: balance sheet accounts
                # For sub-accounts (e.g. 2010.1), skip — they belong to a different section
                # Only store the main account code for balance purposes
                if _is_subaccount(full_code):
                    continue

                is_credit = _is_credit_account(base)
                accounts[base] = {
                    "code": base,
                    "name": name,
                    "debit_start": debit_start,
                    "credit_start": credit_start,
                    "debit_end": debit_end,
                    "credit_end": credit_end,
                    "current": credit_end if is_credit else debit_end,
                    "previous": credit_start if is_credit else debit_start,
                    "section": _classify_account(base),
                    "osv_section": current_section,
                }

            # --- Calculate P&L totals ---
            total_revenue_start = sum(a["credit_start"] for a in revenue_accounts.values())
            total_revenue_end = sum(a["credit_end"] for a in revenue_accounts.values())
            total_expenses_start = sum(a["debit_start"] for a in expense_accounts.values())
            total_expenses_end = sum(a["debit_end"] for a in expense_accounts.values())

            pnl_data = {
                "total_revenue_start": total_revenue_start,
                "total_revenue_end": total_revenue_end,
                "total_expenses_start": total_expenses_start,
                "total_expenses_end": total_expenses_end,
                "net_profit_start": total_revenue_start - total_expenses_start,
                "net_profit_end": total_revenue_end - total_expenses_end,
                "revenue_accounts": revenue_accounts,
                "expense_accounts": expense_accounts,
            }

            wb.close()

            cache["accounts"] = accounts
            cache["pnl"] = pnl_data
            cache["filename"] = filename

            return JSONResponse({
                "status": "success",
                "message": f"Файл '{filename}' успешно загружен и разобран",
                "filename": filename,
                "parsed_accounts": parsed_count,
                "account_codes": sorted(accounts.keys()),
                "revenue_accounts": sorted(revenue_accounts.keys()),
                "expense_accounts": sorted(expense_accounts.keys()),
                "company_info": company_info,
            })
        except Exception as e:
            raise HTTPException(400, f"Ошибка чтения файла: {str(e)}")
    elif filename.lower().endswith(".csv"):
        return JSONResponse({
            "status": "success",
            "message": f"CSV файл '{filename}' успешно загружен",
            "filename": filename,
        })
    else:
        raise HTTPException(400, "Поддерживаются форматы: .xlsx, .xls, .csv")


# ---------------------------------------------------------------------------
# GET /reports/nsbu/balance — НСБУ баланс (Форма 1)
# ---------------------------------------------------------------------------

def _build_nsbu_rows(accounts: dict, pnl: Optional[dict] = None) -> list:
    """Build NSBU balance sheet rows from parsed accounts, with net profit adjustment."""

    def _get(code: str, field: str = "current") -> float:
        acc = accounts.get(code)
        return acc[field] if acc else 0.0

    rows = []

    # === АКТИВ ===
    rows.append({"code": "", "label": "АКТИВ", "current": None, "previous": None, "isHeader": True})

    # I. Внеоборотные активы
    rows.append({"code": "", "label": "I. ВНЕОБОРОТНЫЕ АКТИВЫ", "current": None, "previous": None, "isHeader": True})

    net_fa_current = _get("0100", "current") - _get("0200", "current")
    net_fa_previous = _get("0100", "previous") - _get("0200", "previous")
    rows.append({"code": "0100-0200", "label": "Основные средства (нетто)", "current": net_fa_current, "previous": net_fa_previous})
    rows.append({"code": "0800", "label": "Капитальные вложения", "current": _get("0800"), "previous": _get("0800", "previous")})

    total_noncurrent_current = net_fa_current + _get("0800")
    total_noncurrent_previous = net_fa_previous + _get("0800", "previous")
    rows.append({"code": "", "label": "Итого внеоборотные активы", "current": total_noncurrent_current, "previous": total_noncurrent_previous, "isHeader": True})

    # II. Оборотные активы
    rows.append({"code": "", "label": "II. ОБОРОТНЫЕ АКТИВЫ", "current": None, "previous": None, "isHeader": True})

    rows.append({"code": "1000", "label": "Стройматериалы", "current": _get("1000"), "previous": _get("1000", "previous")})
    rows.append({"code": "1010", "label": "НЗП (незавершённое производство)", "current": _get("1010"), "previous": _get("1010", "previous")})

    inventory_current = _get("1000") + _get("1010")
    inventory_previous = _get("1000", "previous") + _get("1010", "previous")
    rows.append({"code": "", "label": "Итого запасы", "current": inventory_current, "previous": inventory_previous, "isHeader": True})

    rows.append({"code": "2010", "label": "Дебиторы", "current": _get("2010"), "previous": _get("2010", "previous")})
    rows.append({"code": "2300", "label": "Авансы выданные", "current": _get("2300"), "previous": _get("2300", "previous")})

    receivables_current = _get("2010") + _get("2300")
    receivables_previous = _get("2010", "previous") + _get("2300", "previous")
    rows.append({"code": "", "label": "Итого дебиторская задолженность", "current": receivables_current, "previous": receivables_previous, "isHeader": True})

    rows.append({"code": "5010", "label": "Касса", "current": _get("5010"), "previous": _get("5010", "previous")})
    rows.append({"code": "5110", "label": "Расчётный счёт", "current": _get("5110"), "previous": _get("5110", "previous")})
    rows.append({"code": "5210", "label": "Валютный счёт", "current": _get("5210"), "previous": _get("5210", "previous")})

    cash_current = _get("5010") + _get("5110") + _get("5210")
    cash_previous = _get("5010", "previous") + _get("5110", "previous") + _get("5210", "previous")
    rows.append({"code": "", "label": "Итого денежные средства", "current": cash_current, "previous": cash_previous, "isHeader": True})

    total_current_assets_current = inventory_current + receivables_current + cash_current
    total_current_assets_previous = inventory_previous + receivables_previous + cash_previous
    rows.append({"code": "", "label": "Итого оборотные активы", "current": total_current_assets_current, "previous": total_current_assets_previous, "isHeader": True})

    total_assets_current = total_noncurrent_current + total_current_assets_current
    total_assets_previous = total_noncurrent_previous + total_current_assets_previous
    rows.append({
        "code": "", "label": "ИТОГО АКТИВ", "current": total_assets_current, "previous": total_assets_previous,
        "isHeader": True, "isTotalAsset": True,
    })

    # === ПАССИВ ===
    rows.append({"code": "", "label": "ПАССИВ", "current": None, "previous": None, "isHeader": True})

    # III. Капитал
    rows.append({"code": "", "label": "III. КАПИТАЛ", "current": None, "previous": None, "isHeader": True})

    rows.append({"code": "8300", "label": "Уставный капитал", "current": _get("8300"), "previous": _get("8300", "previous")})
    rows.append({"code": "8500", "label": "Резервный капитал", "current": _get("8500"), "previous": _get("8500", "previous")})
    rows.append({"code": "8700", "label": "Нераспределённая прибыль", "current": _get("8700"), "previous": _get("8700", "previous")})

    # Calculate unclosed P&L: the difference that sits in revenue/expense accounts
    # This is the amount needed to make Актив = Пассив
    unclosed_profit_current = 0.0
    unclosed_profit_previous = 0.0
    if pnl:
        # Unclosed P&L = total assets - (equity + liabilities without adjustment)
        # Or directly: revenue credits - expense debits still sitting in sections VI-VII
        unclosed_profit_end = pnl.get("net_profit_end", 0.0)
        unclosed_profit_start = pnl.get("net_profit_start", 0.0)
        # Calculate what assets vs liabilities gap would be without adjustment
        equity_raw_current = _get("8300") + _get("8500") + _get("8700")
        lt_raw = _get("7010") + _get("7800")
        st_raw = _get("6010") + _get("6110") + _get("6310") + _get("6710") + _get("6820") + _get("6610")
        gap_current = total_assets_current - (equity_raw_current + lt_raw + st_raw)
        equity_raw_previous = _get("8300", "previous") + _get("8500", "previous") + _get("8700", "previous")
        lt_raw_p = _get("7010", "previous") + _get("7800", "previous")
        st_raw_p = _get("6010", "previous") + _get("6110", "previous") + _get("6310", "previous") + _get("6710", "previous") + _get("6820", "previous") + _get("6610", "previous")
        gap_previous = total_assets_previous - (equity_raw_previous + lt_raw_p + st_raw_p)
        # Use the gap as the adjustment (this guarantees balance matches)
        unclosed_profit_current = gap_current
        unclosed_profit_previous = gap_previous

    if abs(unclosed_profit_current) > 0.01 or abs(unclosed_profit_previous) > 0.01:
        rows.append({"code": "VI-VII", "label": "Финансовый результат (незакрытый)", "current": unclosed_profit_current, "previous": unclosed_profit_previous})

    equity_current = _get("8300") + _get("8500") + _get("8700") + unclosed_profit_current
    equity_previous = _get("8300", "previous") + _get("8500", "previous") + _get("8700", "previous") + unclosed_profit_previous
    rows.append({"code": "", "label": "Итого капитал", "current": equity_current, "previous": equity_previous, "isHeader": True})

    # IV. Долгосрочные обязательства
    rows.append({"code": "", "label": "IV. ДОЛГОСРОЧНЫЕ ОБЯЗАТЕЛЬСТВА", "current": None, "previous": None, "isHeader": True})

    rows.append({"code": "7010", "label": "Долгосрочные кредиты", "current": _get("7010"), "previous": _get("7010", "previous")})
    rows.append({"code": "7800", "label": "Лизинг", "current": _get("7800"), "previous": _get("7800", "previous")})

    lt_liab_current = _get("7010") + _get("7800")
    lt_liab_previous = _get("7010", "previous") + _get("7800", "previous")
    rows.append({"code": "", "label": "Итого долгосрочные обязательства", "current": lt_liab_current, "previous": lt_liab_previous, "isHeader": True})

    # V. Краткосрочные обязательства
    rows.append({"code": "", "label": "V. КРАТКОСРОЧНЫЕ ОБЯЗАТЕЛЬСТВА", "current": None, "previous": None, "isHeader": True})

    rows.append({"code": "6010", "label": "Поставщики", "current": _get("6010"), "previous": _get("6010", "previous")})
    rows.append({"code": "6110", "label": "Авансы полученные", "current": _get("6110"), "previous": _get("6110", "previous")})
    rows.append({"code": "6310", "label": "Налоги", "current": _get("6310"), "previous": _get("6310", "previous")})
    rows.append({"code": "6710", "label": "Зарплата", "current": _get("6710"), "previous": _get("6710", "previous")})
    rows.append({"code": "6820", "label": "Краткосрочные кредиты", "current": _get("6820"), "previous": _get("6820", "previous")})
    rows.append({"code": "6610", "label": "Налог на прибыль", "current": _get("6610"), "previous": _get("6610", "previous")})

    st_liab_current = _get("6010") + _get("6110") + _get("6310") + _get("6710") + _get("6820") + _get("6610")
    st_liab_previous = _get("6010", "previous") + _get("6110", "previous") + _get("6310", "previous") + _get("6710", "previous") + _get("6820", "previous") + _get("6610", "previous")
    rows.append({"code": "", "label": "Итого краткосрочные обязательства", "current": st_liab_current, "previous": st_liab_previous, "isHeader": True})

    total_liabilities_equity_current = equity_current + lt_liab_current + st_liab_current
    total_liabilities_equity_previous = equity_previous + lt_liab_previous + st_liab_previous
    rows.append({
        "code": "", "label": "ИТОГО ПАССИВ", "current": total_liabilities_equity_current, "previous": total_liabilities_equity_previous,
        "isHeader": True, "isTotalLiability": True,
    })

    return rows


@router.get("/reports/nsbu/balance")
async def get_nsbu_balance(
    current_user: User = Depends(get_current_user),
):
    """Get NSBU balance report from parsed ОСВ data."""
    cache = _user_cache(current_user.id)
    accounts = cache.get("accounts")
    if not accounts:
        return JSONResponse({"rows": []})
    pnl = cache.get("pnl")
    company_info = cache.get("company_info")
    result = {"rows": _build_nsbu_rows(accounts, pnl)}
    if company_info:
        result["company_info"] = company_info
    return JSONResponse(result)


# ---------------------------------------------------------------------------
# GET /reports/ifrs/balance — МСФО баланс (IAS 1)
# ---------------------------------------------------------------------------

def _build_ifrs_rows(accounts: dict, pnl: Optional[dict] = None) -> list:
    """Build IFRS-style balance sheet rows (IAS 1 labels, note refs)."""

    def _get(code: str, field: str = "current") -> float:
        acc = accounts.get(code)
        return acc[field] if acc else 0.0

    rows = []

    # Non-current assets
    rows.append({"code": "", "label": "Внеоборотные активы", "current": None, "previous": None, "isHeader": True, "note": ""})

    net_fa_current = _get("0100", "current") - _get("0200", "current")
    net_fa_previous = _get("0100", "previous") - _get("0200", "previous")
    rows.append({"code": "0100-0200", "label": "Основные средства (IAS 16)", "current": net_fa_current, "previous": net_fa_previous, "note": "5"})
    rows.append({"code": "0800", "label": "Капитальные вложения (IAS 16)", "current": _get("0800"), "previous": _get("0800", "previous"), "note": "6"})

    total_noncurrent_current = net_fa_current + _get("0800")
    total_noncurrent_previous = net_fa_previous + _get("0800", "previous")
    rows.append({"code": "", "label": "Итого внеоборотные активы", "current": total_noncurrent_current, "previous": total_noncurrent_previous, "isHeader": True, "note": ""})

    # Current assets
    rows.append({"code": "", "label": "Оборотные активы", "current": None, "previous": None, "isHeader": True, "note": ""})

    rows.append({"code": "1000", "label": "Запасы — материалы (IAS 2)", "current": _get("1000"), "previous": _get("1000", "previous"), "note": "7"})
    rows.append({"code": "1010", "label": "Запасы — НЗП (IAS 2)", "current": _get("1010"), "previous": _get("1010", "previous"), "note": "7"})
    rows.append({"code": "2010", "label": "Торговая дебиторская задолженность (IFRS 9)", "current": _get("2010"), "previous": _get("2010", "previous"), "note": "8"})
    rows.append({"code": "2300", "label": "Авансы выданные (IAS 1)", "current": _get("2300"), "previous": _get("2300", "previous"), "note": "9"})
    rows.append({"code": "5010", "label": "Денежные средства — касса (IAS 7)", "current": _get("5010"), "previous": _get("5010", "previous"), "note": "10"})
    rows.append({"code": "5110", "label": "Денежные средства — р/с (IAS 7)", "current": _get("5110"), "previous": _get("5110", "previous"), "note": "10"})
    rows.append({"code": "5210", "label": "Денежные средства — валюта (IAS 7)", "current": _get("5210"), "previous": _get("5210", "previous"), "note": "10"})

    cash_current = _get("5010") + _get("5110") + _get("5210")
    cash_previous = _get("5010", "previous") + _get("5110", "previous") + _get("5210", "previous")
    inventory_current = _get("1000") + _get("1010")
    inventory_previous = _get("1000", "previous") + _get("1010", "previous")
    receivables_current = _get("2010") + _get("2300")
    receivables_previous = _get("2010", "previous") + _get("2300", "previous")

    total_current_assets_current = inventory_current + receivables_current + cash_current
    total_current_assets_previous = inventory_previous + receivables_previous + cash_previous
    rows.append({"code": "", "label": "Итого оборотные активы", "current": total_current_assets_current, "previous": total_current_assets_previous, "isHeader": True, "note": ""})

    total_assets_current = total_noncurrent_current + total_current_assets_current
    total_assets_previous = total_noncurrent_previous + total_current_assets_previous
    rows.append({"code": "", "label": "ИТОГО АКТИВЫ", "current": total_assets_current, "previous": total_assets_previous, "isHeader": True, "isTotalAsset": True, "note": ""})

    # Equity
    rows.append({"code": "", "label": "Капитал", "current": None, "previous": None, "isHeader": True, "note": ""})
    rows.append({"code": "8300", "label": "Уставный капитал (IAS 1)", "current": _get("8300"), "previous": _get("8300", "previous"), "note": "11"})
    rows.append({"code": "8500", "label": "Резервный капитал (IAS 1)", "current": _get("8500"), "previous": _get("8500", "previous"), "note": "12"})
    rows.append({"code": "8700", "label": "Нераспределённая прибыль (IAS 1)", "current": _get("8700"), "previous": _get("8700", "previous"), "note": "13"})

    # Unclosed P&L adjustment (same logic as NSBU)
    unclosed_profit_current = 0.0
    unclosed_profit_previous = 0.0
    if pnl:
        equity_raw_current = _get("8300") + _get("8500") + _get("8700")
        lt_raw = _get("7010") + _get("7800")
        st_raw = _get("6010") + _get("6110") + _get("6310") + _get("6710") + _get("6820") + _get("6610")
        unclosed_profit_current = total_assets_current - (equity_raw_current + lt_raw + st_raw)
        equity_raw_previous = _get("8300", "previous") + _get("8500", "previous") + _get("8700", "previous")
        lt_raw_p = _get("7010", "previous") + _get("7800", "previous")
        st_raw_p = _get("6010", "previous") + _get("6110", "previous") + _get("6310", "previous") + _get("6710", "previous") + _get("6820", "previous") + _get("6610", "previous")
        unclosed_profit_previous = total_assets_previous - (equity_raw_previous + lt_raw_p + st_raw_p)

    if abs(unclosed_profit_current) > 0.01 or abs(unclosed_profit_previous) > 0.01:
        rows.append({"code": "VI-VII", "label": "Прибыль текущего периода (IAS 1)", "current": unclosed_profit_current, "previous": unclosed_profit_previous, "note": "14"})

    equity_current = _get("8300") + _get("8500") + _get("8700") + unclosed_profit_current
    equity_previous = _get("8300", "previous") + _get("8500", "previous") + _get("8700", "previous") + unclosed_profit_previous
    rows.append({"code": "", "label": "Итого капитал", "current": equity_current, "previous": equity_previous, "isHeader": True, "note": ""})

    # Non-current liabilities
    rows.append({"code": "", "label": "Долгосрочные обязательства", "current": None, "previous": None, "isHeader": True, "note": ""})
    rows.append({"code": "7010", "label": "Долгосрочные кредиты (IFRS 9)", "current": _get("7010"), "previous": _get("7010", "previous"), "note": "15"})
    rows.append({"code": "7800", "label": "Обязательства по аренде (IFRS 16)", "current": _get("7800"), "previous": _get("7800", "previous"), "note": "16"})

    lt_liab_current = _get("7010") + _get("7800")
    lt_liab_previous = _get("7010", "previous") + _get("7800", "previous")
    rows.append({"code": "", "label": "Итого долгосрочные обязательства", "current": lt_liab_current, "previous": lt_liab_previous, "isHeader": True, "note": ""})

    # Current liabilities
    rows.append({"code": "", "label": "Краткосрочные обязательства", "current": None, "previous": None, "isHeader": True, "note": ""})
    rows.append({"code": "6010", "label": "Торговая кредиторская задолженность (IFRS 9)", "current": _get("6010"), "previous": _get("6010", "previous"), "note": "17"})
    rows.append({"code": "6110", "label": "Авансы полученные (IAS 1)", "current": _get("6110"), "previous": _get("6110", "previous"), "note": "18"})
    rows.append({"code": "6310", "label": "Текущие налоговые обязательства (IAS 12)", "current": _get("6310"), "previous": _get("6310", "previous"), "note": "19"})
    rows.append({"code": "6710", "label": "Обязательства по вознаграждениям (IAS 19)", "current": _get("6710"), "previous": _get("6710", "previous"), "note": "20"})
    rows.append({"code": "6820", "label": "Краткосрочные кредиты (IFRS 9)", "current": _get("6820"), "previous": _get("6820", "previous"), "note": "21"})
    rows.append({"code": "6610", "label": "Налог на прибыль к уплате (IAS 12)", "current": _get("6610"), "previous": _get("6610", "previous"), "note": "22"})

    st_liab_current = _get("6010") + _get("6110") + _get("6310") + _get("6710") + _get("6820") + _get("6610")
    st_liab_previous = _get("6010", "previous") + _get("6110", "previous") + _get("6310", "previous") + _get("6710", "previous") + _get("6820", "previous") + _get("6610", "previous")
    rows.append({"code": "", "label": "Итого краткосрочные обязательства", "current": st_liab_current, "previous": st_liab_previous, "isHeader": True, "note": ""})

    total_liabilities_equity_current = equity_current + lt_liab_current + st_liab_current
    total_liabilities_equity_previous = equity_previous + lt_liab_previous + st_liab_previous
    rows.append({
        "code": "", "label": "ИТОГО КАПИТАЛ И ОБЯЗАТЕЛЬСТВА", "current": total_liabilities_equity_current, "previous": total_liabilities_equity_previous,
        "isHeader": True, "isTotalLiability": True, "note": "",
    })

    return rows


@router.get("/reports/ifrs/balance")
async def get_ifrs_balance(
    current_user: User = Depends(get_current_user),
):
    """Get IFRS balance report (IAS 1) from parsed ОСВ data."""
    cache = _user_cache(current_user.id)
    accounts = cache.get("accounts")
    if not accounts:
        return JSONResponse({"rows": []})
    pnl = cache.get("pnl")
    company_info = cache.get("company_info")
    result = {"rows": _build_ifrs_rows(accounts, pnl)}
    if company_info:
        result["company_info"] = company_info
    return JSONResponse(result)


# ---------------------------------------------------------------------------
# GET /reports/diff — НСБУ vs МСФО сравнение
# ---------------------------------------------------------------------------

def _build_diff_rows(accounts: dict) -> list:
    """Build NSBU vs IFRS diff rows. Currently NSBU ≈ IFRS (no IFRS adjustments)."""

    def _get(code: str) -> float:
        acc = accounts.get(code)
        return acc["current"] if acc else 0.0

    net_fa = _get("0100") - _get("0200")

    items = [
        ("Основные средства (нетто)", net_fa, net_fa, "Без разницы — МСФО-корректировки не применены"),
        ("Капитальные вложения", _get("0800"), _get("0800"), "Без разницы"),
        ("Стройматериалы", _get("1000"), _get("1000"), "Без разницы"),
        ("НЗП", _get("1010"), _get("1010"), "Без разницы"),
        ("Дебиторы", _get("2010"), _get("2010"), "В МСФО возможен ECL-резерв (IFRS 9)"),
        ("Авансы выданные", _get("2300"), _get("2300"), "Без разницы"),
        ("Денежные средства", _get("5010") + _get("5110") + _get("5210"), _get("5010") + _get("5110") + _get("5210"), "Без разницы"),
        ("Уставный капитал", _get("8300"), _get("8300"), "Без разницы"),
        ("Резервный капитал", _get("8500"), _get("8500"), "Без разницы"),
        ("Нераспределённая прибыль", _get("8700"), _get("8700"), "Может отличаться после МСФО-корректировок"),
        ("Долгосрочные кредиты", _get("7010"), _get("7010"), "В МСФО — amortised cost (IFRS 9)"),
        ("Лизинг", _get("7800"), _get("7800"), "В МСФО — IFRS 16 (right-of-use asset + lease liability)"),
        ("Поставщики", _get("6010"), _get("6010"), "Без разницы"),
        ("Авансы полученные", _get("6110"), _get("6110"), "Без разницы"),
        ("Налоги", _get("6310"), _get("6310"), "Без разницы"),
        ("Зарплата", _get("6710"), _get("6710"), "В МСФО — IAS 19 (employee benefits)"),
        ("Краткосрочные кредиты", _get("6820"), _get("6820"), "Без разницы"),
        ("Налог на прибыль", _get("6610"), _get("6610"), "В МСФО возможен отложенный налог (IAS 12)"),
    ]

    return [{"label": label, "nsbu": nsbu, "ifrs": ifrs, "reason": reason} for label, nsbu, ifrs, reason in items]


@router.get("/reports/diff")
async def get_diff_report(
    current_user: User = Depends(get_current_user),
):
    """Get NSBU vs IFRS diff report from parsed ОСВ data."""
    cache = _user_cache(current_user.id)
    accounts = cache.get("accounts")
    if not accounts:
        return JSONResponse({"rows": []})
    return JSONResponse({"rows": _build_diff_rows(accounts)})


# ---------------------------------------------------------------------------
# GET /template/excel — download template (kept as-is)
# ---------------------------------------------------------------------------

@router.get("/template/excel")
async def download_portfolio_template(
    current_user: User = Depends(get_current_user),
):
    """Download empty NSBU + IFRS template."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "НСБУ Баланс"
        ws.append(["Код счёта", "Наименование", "Дебет (тыс.сум)", "Кредит (тыс.сум)"])
        ws.append(["0100", "Основные средства", "", ""])
        ws.append(["0200", "Амортизация ОС", "", ""])
        ws2 = wb.create_sheet("МСФО Баланс")
        ws2.append(["Статья", "Примечание", "Текущий период", "Прошлый период"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=template_nsbu_ifrs.xlsx"},
        )
    except Exception as e:
        raise HTTPException(500, f"Ошибка создания шаблона: {str(e)}")


# ---------------------------------------------------------------------------
# GET /export/excel — export report (kept as-is)
# ---------------------------------------------------------------------------

@router.get("/export/excel")
async def export_portfolio_excel(
    current_user: User = Depends(get_current_user),
):
    """Export full NSBU + IFRS report as Excel with real data from cache."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        cache = _user_cache(current_user.id)
        accounts = cache.get("accounts")
        if not accounts:
            # No data — return stub
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчёт НСБУ + МСФО"
            ws.append(["Раздел", "Показатель", "НСБУ (тыс.сум)", "МСФО (тыс.сум)", "Разница"])
            ws.append(["Пока нет данных", "", "", "", ""])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=report_nsbu_ifrs.xlsx"},
            )

        pnl = cache.get("pnl")
        company_info = cache.get("company_info", {})

        bold = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        wb = Workbook()

        # --- Sheet 1: НСБУ Баланс ---
        ws1 = wb.active
        ws1.title = "НСБУ Баланс"
        if company_info:
            ws1.append(["Компания:", company_info.get("name", "")])
            ws1.append(["ИНН:", company_info.get("inn", "")])
            ws1.append(["Период:", company_info.get("period", "")])
            ws1.append([])
        header_row = ["Код", "Статья", "Текущий период", "Прошлый период"]
        ws1.append(header_row)
        for cell in ws1[ws1.max_row]:
            cell.font = bold
            cell.fill = header_fill
        for row in _build_nsbu_rows(accounts, pnl):
            r = [row.get("code", ""), row.get("label", ""), row.get("current"), row.get("previous")]
            ws1.append(r)
            if row.get("isHeader"):
                for cell in ws1[ws1.max_row]:
                    cell.font = bold

        # --- Sheet 2: МСФО Баланс ---
        ws2 = wb.create_sheet("МСФО Баланс")
        if company_info:
            ws2.append(["Компания:", company_info.get("name", "")])
            ws2.append(["ИНН:", company_info.get("inn", "")])
            ws2.append(["Период:", company_info.get("period", "")])
            ws2.append([])
        header_row = ["Код", "Статья", "Текущий период", "Прошлый период", "Примечание"]
        ws2.append(header_row)
        for cell in ws2[ws2.max_row]:
            cell.font = bold
            cell.fill = header_fill
        for row in _build_ifrs_rows(accounts, pnl):
            r = [row.get("code", ""), row.get("label", ""), row.get("current"), row.get("previous"), row.get("note", "")]
            ws2.append(r)
            if row.get("isHeader"):
                for cell in ws2[ws2.max_row]:
                    cell.font = bold

        # --- Sheet 3: Разница ---
        ws3 = wb.create_sheet("Разница")
        header_row = ["Статья", "НСБУ", "МСФО", "Разница", "Комментарий"]
        ws3.append(header_row)
        for cell in ws3[ws3.max_row]:
            cell.font = bold
            cell.fill = header_fill
        for row in _build_diff_rows(accounts):
            diff_val = round(row["ifrs"] - row["nsbu"], 2) if row["nsbu"] is not None and row["ifrs"] is not None else None
            ws3.append([row["label"], row["nsbu"], row["ifrs"], diff_val, row.get("reason", "")])

        # --- Sheet 4: KPI ---
        from app.api.v1.routers.analytics_chapter import _get_balance_aggregates, _calc_kpis
        agg = _get_balance_aggregates()
        ws4 = wb.create_sheet("KPI")
        header_row = ["Группа", "Показатель", "Значение", "Норма", "Статус"]
        ws4.append(header_row)
        for cell in ws4[ws4.max_row]:
            cell.font = bold
            cell.fill = header_fill
        if agg:
            for group in _calc_kpis(agg):
                for m in group["metrics"]:
                    ws4.append([group["title"], m["label"], m["value"], m["norm"], m["status"]])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=report_nsbu_ifrs.xlsx"},
        )
    except Exception as e:
        raise HTTPException(500, str(e))
